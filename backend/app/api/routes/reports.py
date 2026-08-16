import hashlib
import json
import re
from datetime import UTC, datetime
from io import BytesIO
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from smart_commissioning_core.db.repositories import DiscoveryRepository
from smart_commissioning_core.db.sync_v2_repository import SyncV2Repository
from smart_commissioning_core.rbac import Role

from app.core.auth import AuthPrincipal, get_principal, require_role
from app.core.scopes import (
    allowed_scope_pairs,
    load_scoped_report,
    load_scoped_run,
    require_project_site_access,
)
from app.schemas.jobs import (
    ReportDeleteRequest,
    ReportDeleteResponse,
    ReportListResponse,
    ReportRequest,
    ReportSummary,
    RunRecord,
)
from app.services.report_artifacts import (
    ReportArtifactIntegrityError,
    delete_report_artifact,
    load_owned_report_artifact,
    store_report_artifact,
)
from app.services.report_naming import build_report_file_name, report_content_disposition
from app.services.report_pdf import PdfDocument
from app.services.run_service import (
    DISCOVERY_JOB_TYPES,
    VALIDATION_JOB_TYPES,
    ReportListIntegrityError,
    RunService,
)
from app.services.udmi_report_model import (
    ASSET_METRIC_LABELS,
    FAULT_METRIC_LABELS,
    ISSUE_METRIC_LABELS,
    PAYLOAD_METRIC_LABELS,
    build_udmi_report_model,
)
from app.services.validation_export import redact_export_value

router = APIRouter()
service = RunService()

# RBAC: listing/reading/downloading a report is viewer+; generating a report
# (creating a report run) is engineer+.
require_viewer = require_role(Role.VIEWER)
require_engineer = require_role(Role.ENGINEER)
_PROVENANCE_SHA256_RE = re.compile(r"^[0-9a-f]{64}$")


class ReportProvenanceError(ValueError):
    """Stored source identity cannot be represented in a signed report manifest."""


def _require_scoped_report(report_id: str, principal: AuthPrincipal) -> None:
    """Conceal foreign reports while preserving the caller-supplied ID in errors."""
    try:
        load_scoped_report(report_id, principal, engine=service.engine)
    except HTTPException as error:
        if error.status_code == 404:
            raise HTTPException(
                status_code=404,
                detail=f"Report '{report_id}' was not found.",
            ) from error
        raise


def _to_report_summary(report_id: str | RunRecord) -> ReportSummary:
    run = service.get_run(report_id) if isinstance(report_id, str) else report_id
    if run.job_type != "report_generation":
        raise FileNotFoundError(run.run_id)

    report_type = run.parameters.get("report_type", "evidence_pack")
    if not isinstance(report_type, str):
        report_type = "evidence_pack"
    output_format = run.parameters.get("output_format", "zip")
    if output_format not in {"docx", "pdf", "xlsx", "zip"}:
        output_format = "zip"
    # Scoped source runs, read back from the stored parameters. Same defensive
    # shape-checking as report_type/output_format above: a run record is
    # persisted JSON, so nothing guarantees the list survived as a list of str.
    # (Attribute access on run.parameters is unredacted — the field_serializer
    # only fires on serialization — and run ids are not sensitive.)
    raw_source_run_ids = run.parameters.get("source_run_ids")
    source_run_ids = (
        [item for item in raw_source_run_ids if isinstance(item, str)]
        if isinstance(raw_source_run_ids, list)
        else []
    )
    evidence_set_id = run.parameters.get("evidence_set_id")
    if not isinstance(evidence_set_id, str) or not evidence_set_id.strip():
        evidence_set_id = None
    stored_title = run.parameters.get("report_title")
    report_title = (
        stored_title.strip()
        if isinstance(stored_title, str) and stored_title.strip()
        else "UDMI Validation Report"
        if report_type == "udmi_validation"
        else _BRAND_DOC_TITLE
    )
    udmi_report_variant = run.parameters.get("udmi_report_variant", "technical")
    if udmi_report_variant not in {"client", "technical"}:
        udmi_report_variant = "technical"
    return ReportSummary(
        report_id=run.run_id,
        report_type=report_type,
        output_format=output_format,
        status=run.status,
        file_name=build_report_file_name(
            report_type,
            run.run_id,
            output_format,
            report_title=report_title,
            # Deliberately require the literal boolean marker. Runs persisted
            # before v0.1.25 keep their historical names.
            custom_title=run.parameters.get("report_title_custom") is True,
        ),
        created_at=run.created_at,
        source_run_ids=source_run_ids,
        evidence_set_id=evidence_set_id,
        report_title=report_title,
        udmi_report_variant=udmi_report_variant,
    )


@router.post("", response_model=ReportSummary, dependencies=[Depends(require_engineer)])
def create_report(
    request: ReportRequest,
    principal: AuthPrincipal = Depends(get_principal),
) -> ReportSummary:
    require_project_site_access(
        principal, request.project_id, request.site_id, engine=service.engine
    )
    for source_run_id in dict.fromkeys(request.source_run_ids):
        scoped = load_scoped_run(source_run_id, principal, engine=service.engine)
        if (scoped.project_id, scoped.site_id) != (
            request.project_id,
            request.site_id,
        ):
            raise HTTPException(status_code=404, detail="Run not found.")
    try:
        run, report = service.create_report_run(request)
    except ValueError as error:
        raise HTTPException(status_code=422, detail=str(error)) from error
    manifest: dict[str, object] | None = None
    try:
        render_run = run.model_copy(
            update={"status": "succeeded", "stage": "report_ready", "progress_percent": 100}
        )
        snapshot_hash = run.parameters.get("report_snapshot_sha256")
        generated_at = run.parameters.get("report_generated_at")
        if not isinstance(snapshot_hash, str) or not isinstance(generated_at, str):
            raise RuntimeError("Report snapshot provenance is incomplete.")
        content, media_type = _build_report_artifact(render_run, report.output_format)
        provenance = _report_provenance(render_run)
        application_version = provenance.get("application_version")
        source_run_ids = provenance.get("source_run_ids")
        source_provenance = provenance.get("sources")
        if (
            not isinstance(application_version, str)
            or not isinstance(source_run_ids, list)
            or not isinstance(source_provenance, list)
        ):
            raise RuntimeError("Report build provenance is incomplete.")
        manifest = store_report_artifact(
            report_id=run.run_id,
            snapshot_hash=snapshot_hash,
            file_name=report.file_name,
            media_type=media_type,
            artifact=content,
            origin=str(run.edge_id or "api"),
            signed_at=generated_at,
            evidence_set_id=report.evidence_set_id,
            application_version=application_version,
            source_run_ids=source_run_ids,
            source_provenance=source_provenance,
        )
        service.complete_report_run(run.run_id, manifest)
    except Exception as error:
        snapshot_hash = run.parameters.get("report_snapshot_sha256")
        discarded = False
        if isinstance(snapshot_hash, str):
            try:
                discarded = service.discard_unmaterialized_report_run(
                    run.run_id,
                    expected_snapshot_sha256=snapshot_hash,
                )
            except Exception:
                pass
        if discarded and manifest is not None:
            try:
                # Remove only bytes owned by the row that cleanup just proved
                # was still the pristine report created by this request.
                delete_report_artifact(manifest, report_id=run.run_id)
            except (OSError, RuntimeError, ValueError):
                pass
        if isinstance(error, ReportProvenanceError):
            raise HTTPException(status_code=422, detail=str(error)) from error
        raise HTTPException(status_code=500, detail="Report artifact materialization failed.") from error
    return _to_report_summary(render_run)


@router.get("", response_model=ReportListResponse, dependencies=[Depends(require_viewer)])
def list_reports(
    limit: int = Query(default=100, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    principal: AuthPrincipal = Depends(get_principal),
) -> ReportListResponse:
    scopes = allowed_scope_pairs(principal, engine=service.engine)
    try:
        records, total = service.page_verified_report_records(
            limit=limit,
            offset=offset,
            scope_pairs=scopes,
        )
    except ReportListIntegrityError as error:
        raise HTTPException(status_code=409, detail=str(error)) from error
    return ReportListResponse(
        reports=[_to_report_summary(run) for run in records],
        total=total,
        limit=limit,
        offset=offset,
        has_more=offset + limit < total,
    )


class ReportExportRequest(BaseModel):
    # Non-empty so an empty selection 422s instead of yielding an empty zip. The
    # ids ride in a JSON body, not repeated query params, so an unbounded
    # selection never hits the request-line limits uvicorn/h11 and proxies cap.
    report_ids: list[str] = Field(min_length=1)


@router.post(
    "/delete",
    response_model=ReportDeleteResponse,
    dependencies=[Depends(require_engineer)],
)
def delete_reports(
    request: ReportDeleteRequest,
    principal: AuthPrincipal = Depends(get_principal),
) -> ReportDeleteResponse:
    """Delete one or several generated reports after validating the full batch."""

    for report_id in dict.fromkeys(request.report_ids):
        _require_scoped_report(report_id, principal)

    try:
        deleted = service.delete_report_runs(request.report_ids)
    except FileNotFoundError as error:
        report_id = str(error.args[0]) if error.args else ""
        raise HTTPException(
            status_code=404,
            detail=f"Report '{report_id}' was not found.",
        ) from error

    warnings: list[str] = []
    for report_id, manifest in deleted:
        if manifest is None:
            continue
        try:
            delete_report_artifact(manifest, report_id=report_id)
        except (OSError, RuntimeError, ValueError):
            # The database deletion has committed. Keep the response honest while
            # avoiding raw filesystem paths or exception text in the public API.
            warnings.append(
                f"Report '{report_id}' was deleted, but its local artifact requires manual cleanup."
            )
    deleted_ids = [report_id for report_id, _manifest in deleted]
    return ReportDeleteResponse(
        deleted_report_ids=deleted_ids,
        deleted_count=len(deleted_ids),
        artifact_cleanup_warnings=warnings,
    )


# Declared BEFORE /{report_id}: kept ahead of the path route so the literal
# /export can never be swallowed as report_id="export" (defensive — a POST could
# not match a GET /{report_id} anyway). Multiple ticked reports become ONE zip so
# the browser fires a single download (its multiple-download throttle otherwise
# keeps only one file); a single report keeps downloading directly via
# /{report_id}/download.
@router.post("/export", dependencies=[Depends(require_viewer)])
def export_reports(
    request: ReportExportRequest,
    principal: AuthPrincipal = Depends(get_principal),
) -> Response:
    # Resolve every id up front (order-preserving dedupe, a twice-ticked report
    # yields one member) so an unknown id 404s the WHOLE request rather than
    # returning a silently partial archive (honesty rule). Each member reuses the
    # exact per-report path so its bytes equal that report's own download.
    seen: set[str] = set()
    runs: list[object] = []
    for candidate_id in request.report_ids:
        if candidate_id in seen:
            continue
        seen.add(candidate_id)
        _require_scoped_report(candidate_id, principal)
        try:
            run = service.get_report_for_serving(candidate_id)
        except (FileNotFoundError, RuntimeError, ValueError) as error:
            raise HTTPException(
                status_code=404, detail=f"Report '{candidate_id}' was not found."
            ) from error
        if run.job_type != "report_generation":
            raise HTTPException(status_code=404, detail=f"Report '{candidate_id}' was not found.")
        runs.append(run)

    buffer = BytesIO()
    manifest_members: list[dict[str, object]] = []
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        for run in runs:
            report = _to_report_summary(run)
            try:
                content, _media_type = _stored_or_legacy_artifact(
                    run, report.output_format
                )
            except ReportArtifactIntegrityError as error:
                raise HTTPException(status_code=409, detail=str(error)) from error
            # Pinned to the zip epoch (same as _normalize_zip_bytes) so the
            # bundle is byte-reproducible; member name embeds the run id, so
            # names are unique and never collide.
            info = ZipInfo(filename=report.file_name, date_time=_ZIP_EPOCH)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            archive.writestr(info, content)
            manifest_members.append(
                {
                    "report_id": report.report_id,
                    "file_name": report.file_name,
                    "output_format": report.output_format,
                    "sha256": hashlib.sha256(content).hexdigest(),
                    "byte_size": len(content),
                    "evidence_set_id": report.evidence_set_id,
                }
            )
        manifest = {
            "schema_version": "1.0",
            "manifest_type": "selected_reports_bundle",
            "report_ids": [member["report_id"] for member in manifest_members],
            "evidence_set_ids": sorted(
                {
                    member["evidence_set_id"]
                    for member in manifest_members
                    if isinstance(member["evidence_set_id"], str)
                    and member["evidence_set_id"]
                }
            ),
            "members": manifest_members,
        }
        manifest_info = ZipInfo(filename="report_set_manifest.json", date_time=_ZIP_EPOCH)
        manifest_info.compress_type = ZIP_DEFLATED
        manifest_info.external_attr = 0o600 << 16
        archive.writestr(
            manifest_info,
            json.dumps(manifest, indent=2, sort_keys=True).encode("utf-8"),
        )
    return Response(
        content=buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="reports_export.zip"'},
    )


@router.get("/{report_id}", response_model=ReportSummary, dependencies=[Depends(require_viewer)])
def get_report(
    report_id: str,
    principal: AuthPrincipal = Depends(get_principal),
) -> ReportSummary:
    _require_scoped_report(report_id, principal)
    try:
        return _to_report_summary(service.get_report_for_serving(report_id))
    except (FileNotFoundError, RuntimeError, ValueError) as error:
        raise HTTPException(status_code=404, detail=f"Report '{report_id}' was not found.") from error


@router.get("/{report_id}/download", dependencies=[Depends(require_viewer)])
def download_report(
    report_id: str,
    principal: AuthPrincipal = Depends(get_principal),
) -> Response:
    _require_scoped_report(report_id, principal)
    try:
        run = service.get_report_for_serving(report_id)
    except (FileNotFoundError, RuntimeError, ValueError) as error:
        raise HTTPException(status_code=404, detail=f"Report '{report_id}' was not found.") from error
    if run.job_type != "report_generation":
        raise HTTPException(status_code=404, detail=f"Report '{report_id}' was not found.")

    report = _to_report_summary(run)
    try:
        content, media_type = _stored_or_legacy_artifact(run, report.output_format)
    except ReportArtifactIntegrityError as error:
        raise HTTPException(status_code=409, detail=str(error)) from error
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": report_content_disposition(report.file_name)},
    )


def _stored_or_legacy_artifact(run: object, output_format: str) -> tuple[bytes, str]:
    content, media_type, _manifest = _report_artifact_for_serving(run, output_format)
    return content, media_type


def _report_artifact_for_serving(
    run: object,
    output_format: str,
    *,
    require_valid_signature: bool = True,
) -> tuple[bytes, str, dict[str, object] | None]:
    """Load one report through its verified seal and owned signed manifest."""

    run_id = getattr(run, "run_id", None)
    if not isinstance(run_id, str):
        raise ReportArtifactIntegrityError("Stored report id is invalid.")
    try:
        serving_run = service.get_report_for_serving(run_id)
    except (FileNotFoundError, RuntimeError, ValueError) as error:
        raise ReportArtifactIntegrityError(
            "Stored report evidence failed sealed integrity verification."
        ) from error
    summary = (
        serving_run.result_summary
        if isinstance(serving_run.result_summary, dict)
        else {}
    )
    manifest = summary.get("artifact_manifest")
    if isinstance(manifest, dict):
        artifact, media_type = load_owned_report_artifact(
            report_id=run_id,
            manifest=manifest,
            synchronized=SyncV2Repository(service.engine).get_artifact(run_id),
            require_valid_signature=require_valid_signature,
        )
        return artifact, media_type, dict(manifest)
    # Pre-v0.1.26 records remain downloadable, but the compatibility path is
    # strictly read-only and never creates provenance during GET.
    try:
        artifact, media_type = _build_report_artifact(serving_run, output_format)
    except (RuntimeError, ValueError) as error:
        raise ReportArtifactIntegrityError(
            "Historical report evidence could not be reconstructed."
        ) from error
    return artifact, media_type, None


def _generated_at(run: object) -> str:
    """Stable artifact-generation timestamp for the run.

    Reports MUST be derivable from the stored run record (the audit
    requirement), so the "Generated" field cannot be ``datetime.now`` at
    download time — that would make the bytes non-reproducible and break
    hash-based verification. The first download persists a fixed timestamp in
    result_summary; every later regeneration reuses it, yielding byte-identical
    artifacts that the verify endpoint can re-hash.
    """
    parameters = run.parameters if isinstance(run.parameters, dict) else {}
    frozen = parameters.get("report_generated_at")
    if isinstance(frozen, str) and frozen:
        return frozen
    summary = run.result_summary if isinstance(run.result_summary, dict) else {}
    existing = summary.get("report_generated_at")
    if isinstance(existing, str) and existing:
        return existing
    created_at = getattr(run, "created_at", None)
    if isinstance(created_at, datetime):
        return created_at.isoformat()
    raise RuntimeError("Stored report generation timestamp is invalid.")


# Fixed member timestamp (1980-01-01, the zip epoch) so regenerated artifacts
# are byte-identical. Reports are derived from the stored run record, so the
# bytes must be reproducible for hash-based verification — embedded "now"
# timestamps (zipfile uses localtime; openpyxl stamps created/modified) would
# otherwise make every regeneration hash differently.
_ZIP_EPOCH = (1980, 1, 1, 0, 0, 0)
# Fixed instant pinned into openpyxl core properties (docProps/core.xml).
_ARTIFACT_PROPERTIES_EPOCH = datetime(1980, 1, 1, tzinfo=UTC)


def _build_report_artifact(run: object, output_format: str) -> tuple[bytes, str]:
    # PDF is NOT a zip container, so it must skip the zip normalisation pass the
    # OOXML/zip formats need; it is deterministic by construction instead (no
    # /CreationDate, no /ID — see app.services.report_pdf).
    if output_format == "pdf":
        return _build_pdf_report(run), "application/pdf"
    if output_format == "xlsx":
        artifact = _normalize_zip_bytes(_build_xlsx_report(run))
        return artifact, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if output_format == "docx":
        artifact = _normalize_zip_bytes(_build_docx_report(run))
        return artifact, "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    return _normalize_zip_bytes(_build_zip_report(run)), "application/zip"


def _normalize_zip_bytes(data: bytes) -> bytes:
    """Rewrite a zip container with deterministic member order + timestamps.

    All three report formats are ZIP containers (xlsx/docx are OOXML zips). The
    rewrite sorts entries by name and pins every date_time to the zip epoch so
    the same run record always produces byte-identical artifacts, which is what
    lets the verify endpoint re-hash a regenerated artifact.
    """
    source = BytesIO(data)
    out = BytesIO()
    with ZipFile(source, "r") as reader, ZipFile(out, "w", ZIP_DEFLATED) as writer:
        for name in sorted(reader.namelist()):
            info = ZipInfo(filename=name, date_time=_ZIP_EPOCH)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            writer.writestr(info, _pin_member_timestamps(name, reader.read(name)))
    return out.getvalue()


# openpyxl overwrites dcterms:modified with "now" on save regardless of the
# workbook properties, so pin it (and created) to the epoch in core.xml content.
_MODIFIED_RE = re.compile(rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)")


def _pin_member_timestamps(name: str, payload: bytes) -> bytes:
    """Pin OOXML core-property timestamps so the artifact bytes are reproducible."""
    if name == "docProps/core.xml":
        return _MODIFIED_RE.sub(rb"\g<1>1980-01-01T00:00:00Z\g<2>", payload)
    return payload


def _report_rows(
    run: object,
    *,
    udmi_data: dict[str, object] | None = None,
) -> list[tuple[str, str]]:
    parameters = run.parameters
    if udmi_data is not None:
        return [
            ("Project", str(udmi_data.get("project_label") or "Not recorded")),
            ("Site", str(udmi_data.get("site_label") or "Not recorded")),
            ("Report ID", str(run.run_id)),
            ("Evidence set ID", str(udmi_data.get("evidence_set_id") or "Not recorded")),
            (
                "Acceptance",
                "Eligible for field acceptance"
                if udmi_data.get("acceptance_eligible") is True
                else "Not eligible for field acceptance",
            ),
            ("Generated", _generated_at(run)),
        ]
    rows = [
        ("Report title", _report_title(run)),
        ("Report type", str(parameters.get("report_type", "evidence_pack"))),
        ("Output format", str(parameters.get("output_format", "zip")).upper()),
        ("Project", str(run.project_id)),
        ("Site", str(run.site_id)),
        ("Status", str(run.status)),
        (
            "Source runs",
            ", ".join(str(item) for item in parameters.get("source_run_ids", []))
            or "None selected (no run findings included)",
        ),
        ("Generated", _generated_at(run)),
    ]
    return rows


def _report_title(run: object) -> str:
    parameters = run.parameters if isinstance(run.parameters, dict) else {}
    value = parameters.get("report_title")
    if isinstance(value, str) and value.strip():
        return value.strip()
    if parameters.get("report_type") == "udmi_validation":
        return "UDMI Validation Report"
    return _BRAND_DOC_TITLE


def _is_udmi_report(run: object) -> bool:
    parameters = run.parameters if isinstance(run.parameters, dict) else {}
    return parameters.get("report_type") == "udmi_validation"


def _udmi_report_variant(run: object) -> str:
    parameters = run.parameters if isinstance(run.parameters, dict) else {}
    value = parameters.get("udmi_report_variant", "technical")
    return value if value in {"client", "technical"} else "technical"


def _udmi_report_data(run: object) -> dict[str, object] | None:
    if not _is_udmi_report(run):
        return None
    parameters = run.parameters if isinstance(run.parameters, dict) else {}
    if "udmi_report_snapshot" in parameters:
        snapshot = parameters.get("udmi_report_snapshot")
        return dict(snapshot) if isinstance(snapshot, dict) else None
    return build_udmi_report_model(
        _source_runs(run),
        parameters.get("udmi_scope"),
    )


def _display_timestamp(value: object) -> str:
    if not isinstance(value, str) or not value:
        return _NO_VALUE
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return value
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC).strftime("%Y-%m-%d %H:%M UTC")


# Report branding furniture (field ask 2026-07-15, ITP witnessing packs). One
# wordmark, one look across pdf/docx/xlsx page furniture: header = wordmark +
# document title with a thin rule; footer = wordmark + page number + the run id
# (traceability back to a run record on a printed page). The run id is stable
# per run, so the bytes stay reproducible. Phase 1 is text only — logo image
# embedding is deferred to a later release. The zip and topics.xlsx export are
# deliberately NOT branded (zip has no page concept; the export is not a report).
_BRAND_NAME = "ELECTRACOM"
_BRAND_DOC_TITLE = "Smart Commissioning Report"

_UDMI_EXECUTIVE_TITLE = "Executive Summary"
_UDMI_SYSTEM_TITLE = "Metrics by System"
_UDMI_ASSET_TITLE = "Asset Validation Schedule"
_UDMI_FAULT_MATRIX_TITLE = "Fault Matrix"
_UDMI_FAULT_DETAIL_TITLE = "Faults in Detail"
_UDMI_DEFINITIONS_TITLE = "Metric Definitions"
_UDMI_INCOMPLETE_SCOPE_TITLE = "Validation Scope Incomplete"
_UDMI_WRONG_TOPIC_TITLE = "Registered Assets on Wrong Topics"
_UDMI_WRONG_TOPIC_SHEET_TITLE = "Wrong Topic Assets"
_UDMI_ANNOTATED_REGISTER_TITLE = "Annotated Input Register"

# Section titles for the end-to-end validation report (field ask 2026-07-14).
# The frontend references these strings — keep them stable across formats.
_SUMMARY_SECTION_TITLE = "Summary"
_FAILURE_SECTION_TITLE = "Failure detail"
_SILENT_SECTION_TITLE = "Silent systems"
_SILENT_NOTE = (
    "Silent systems are devices that published nothing within the allowed capture window; "
    "they are neither validated nor failed."
)
_NO_SILENT_NOTE = "No silent systems: every expected device published within the capture window."
_NO_FINDINGS_NOTE = "No findings in the scoped source runs."
# Fallback Device ID for pre-upgrade validation runs that recorded only the
# silent-device COUNT (not_publishing) and not the ids themselves.
_SILENT_IDS_NOT_RECORDED = "(ids not recorded by this run's app version)"
# Rendered where a run's record simply does not carry a value.
_NO_VALUE = "—"

_FINDING_COLUMNS = (
    "Source Run",
    "Issue ID",
    "Asset",
    "Severity",
    "Type",
    "Point",
    "Expected",
    "Observed",
    "Mismatch",
    "Suggested Action",
    "Description",
)

_VALIDATION_SUMMARY_COLUMNS = (
    "Source Run",
    "Type",
    "Status",
    "Expected Devices",
    "Publishing",
    "Silent",
    "Blocking Issues",
    "Compliance %",
)

_SILENT_COLUMNS = ("Source Run", "Device ID")

# Discovery inventory sections (field ask 2026-07-15, per-head handover packs).
# The frontend / API may reference these titles — keep them stable across
# formats, same convention as the validation section titles above.
_INVENTORY_SUMMARY_TITLE = "Discovery summary"
_INVENTORY_IP_TITLE = "Discovered IP hosts"
_INVENTORY_BACNET_DEVICES_TITLE = "Discovered BACnet devices"
_INVENTORY_BACNET_POINTS_TITLE = "Discovered BACnet points"
_INVENTORY_BACNET_SILENT_TITLE = "Expected BACnet devices not responding"
_INVENTORY_MQTT_TITLE = "Discovered MQTT topics"

_INVENTORY_SUMMARY_COLUMNS = ("Source Run", "Type", "Status", "Counts")
_IP_INVENTORY_COLUMNS = (
    "Source Run",
    "Address",
    "Hostname",
    "MAC",
    "Open Ports",
    "Forbidden Open",
    "Unexpected Open",
    "Missing Expected",
)
_BACNET_DEVICE_COLUMNS = (
    "Source Run",
    "Instance",
    "Address",
    "Name",
    "Vendor",
    "Model",
    "Register Asset",
    "Points",
)
_BACNET_POINT_COLUMNS = ("Source Run", "Device", "Point ID", "Point Name", "Value", "Units")
_BACNET_SILENT_COLUMNS = ("Source Run", "Register Asset", "Instance", "Address", "Directed Who-Is")
_MQTT_TOPIC_COLUMNS = ("Source Run", "Topic", "Messages", "Device Ref")

# Honesty-rule wording for the expected-but-silent section: it mirrors the
# engine's own framing (bacnet_discovery: "amber, never a failure, never device
# absent") and the BACnet-135 rationale. It must never read "fail" or "absent".
_BACNET_SILENT_NOTE = (
    "Expected devices that did not answer during the scan window. Directed-Who-Is silence is "
    "inconclusive under BACnet-135 (an off-subnet device may reply with a local broadcast we "
    "cannot hear); these rows are neither confirmed present nor absent."
)
# Shown when a discovery head was scoped but recorded no rows: an empty scan is
# a recorded result, not a gap in the report.
_INVENTORY_EMPTY_NOTE = (
    "No rows recorded by the scoped discovery runs (an empty scan is a recorded result, "
    "not an omission)."
)

# Excel caps sheet names at 31 chars; only the silent title exceeds it, so map
# it to a short unique name (the full title is surfaced in the sheet's row 1).
_INVENTORY_SHEET_NAMES = {_INVENTORY_BACNET_SILENT_TITLE: "Expected not responding"}

# xlsx column widths by column name (reused across every inventory sheet — the
# names are unique enough that one map covers all sections). Wide free-text
# columns (Topic/Point Name/Counts) get room; id-like columns a fixed width.
_INVENTORY_COLUMN_WIDTHS = {
    "Source Run": 26,
    "Address": 22,
    "Hostname": 24,
    "MAC": 20,
    "Open Ports": 24,
    "Forbidden Open": 16,
    "Unexpected Open": 16,
    "Missing Expected": 16,
    "Instance": 12,
    "Name": 28,
    "Vendor": 18,
    "Model": 18,
    "Register Asset": 26,
    "Points": 10,
    "Device": 20,
    "Point ID": 18,
    "Point Name": 40,
    "Value": 20,
    "Units": 12,
    "Directed Who-Is": 16,
    "Topic": 40,
    "Messages": 12,
    "Device Ref": 24,
    "Type": 18,
    "Status": 14,
    "Counts": 60,
}


def _source_runs(run: object) -> list[object]:
    """The report's source runs, in the order they were scoped.

    Order-preserving dedupe: a run id scoped twice must contribute one Summary
    row and one set of findings, not doubled device/blocking totals.
    """
    parameters = run.parameters if isinstance(run.parameters, dict) else {}
    raw_snapshots = parameters.get("source_run_snapshots")
    if isinstance(raw_snapshots, list):
        return [
            RunRecord.model_validate(snapshot)
            for snapshot in raw_snapshots
            if isinstance(snapshot, dict)
        ]

    source_ids = parameters.get("source_run_ids", [])
    if not isinstance(source_ids, list):
        return []
    sources: list[object] = []
    seen: set[str] = set()
    for source_id in source_ids:
        key = str(source_id)
        if key in seen:
            continue
        seen.add(key)
        # Creation validates every id. If retention or manual database damage
        # later removes one, fail the download instead of silently weakening the
        # report's evidence scope.
        sources.append(service.get_run(key))
    return sources


def _report_provenance(run: object) -> dict[str, object]:
    """Return source and build identity from the report's frozen snapshot."""

    parameters = run.parameters if isinstance(run.parameters, dict) else {}
    snapshot = parameters.get("report_snapshot_v2")
    snapshot = snapshot if isinstance(snapshot, dict) else {}
    raw_source_ids = snapshot.get("source_run_ids", parameters.get("source_run_ids", []))
    if not isinstance(raw_source_ids, list) or any(
        not isinstance(source_id, str) or not 1 <= len(source_id) <= 64
        for source_id in raw_source_ids
    ):
        raise ReportProvenanceError("Stored report source-run provenance is malformed.")
    source_ids = list(raw_source_ids)
    if len(source_ids) != len(set(source_ids)):
        raise ReportProvenanceError("Stored report provenance contains duplicate source run IDs.")

    raw_snapshots = snapshot.get(
        "source_run_snapshots", parameters.get("source_run_snapshots", [])
    )
    raw_snapshots = raw_snapshots if isinstance(raw_snapshots, list) else []
    snapshots_by_id = {
        source["run_id"]: source
        for source in raw_snapshots
        if isinstance(source, dict) and isinstance(source.get("run_id"), str)
    }
    raw_contexts = snapshot.get("configuration_provenance", {})
    contexts = raw_contexts if isinstance(raw_contexts, dict) else {}
    raw_hashes = snapshot.get("source_result_hashes", {})
    result_hashes = raw_hashes if isinstance(raw_hashes, dict) else {}
    raw_seals = snapshot.get("source_run_seals", parameters.get("source_run_seals", {}))
    seals = raw_seals if isinstance(raw_seals, dict) else {}

    def recorded_text(value: object, *, field: str, max_length: int) -> str | None:
        if value is None or value == "":
            return None
        if not isinstance(value, str):
            raise ReportProvenanceError(
                f"Stored report provenance field '{field}' must be text."
            )
        normalized = value.strip()
        if not normalized:
            return None
        if len(normalized) > max_length:
            raise ReportProvenanceError(
                f"Stored report provenance field '{field}' exceeds {max_length} characters."
            )
        return normalized

    def recorded_sha256(value: object, *, field: str) -> str | None:
        normalized = recorded_text(value, field=field, max_length=64)
        if normalized is None:
            return None
        normalized = normalized.casefold()
        if not _PROVENANCE_SHA256_RE.fullmatch(normalized):
            raise ReportProvenanceError(
                f"Stored report provenance field '{field}' is not a SHA-256 digest."
            )
        return normalized

    sources: list[dict[str, object]] = []
    for source_id in source_ids:
        source = snapshots_by_id.get(source_id, {})
        source_parameters = source.get("parameters", {}) if isinstance(source, dict) else {}
        source_parameters = source_parameters if isinstance(source_parameters, dict) else {}
        context = contexts.get(source_id, {})
        context = context if isinstance(context, dict) else {}
        seal = seals.get(source_id, {})
        seal = seal if isinstance(seal, dict) else {}
        sources.append(
            {
                "source_run_id": source_id,
                "application_version": recorded_text(
                    context.get("application_version") or source_parameters.get("application_version"),
                    field="application_version",
                    max_length=128,
                ),
                "source_commit": recorded_text(
                    source_parameters.get("source_commit")
                    or source_parameters.get("application_source_commit"),
                    field="source_commit",
                    max_length=128,
                ),
                "portable_exe_sha256": recorded_sha256(
                    source_parameters.get("portable_exe_sha256")
                    or source_parameters.get("exe_sha256")
                    or source_parameters.get("portable_hash"),
                    field="portable_exe_sha256",
                ),
                "context_sha256": recorded_sha256(
                    context.get("context_sha256") or seal.get("context_sha256"),
                    field="context_sha256",
                ),
                "result_sha256": recorded_sha256(
                    result_hashes.get(source_id) or seal.get("result_sha256"),
                    field="result_sha256",
                ),
            }
        )

    application_version = recorded_text(
        snapshot.get("renderer_version") or parameters.get("renderer_version"),
        field="application_version",
        max_length=64,
    )
    return {
        "schema_version": "1.0",
        "application_version": application_version,
        "source_run_ids": source_ids,
        "sources": sources,
    }


def _asset_topic_discovery_export(run: object) -> dict[str, object] | None:
    """Return retained per-run asset-topic ledgers for a ZIP evidence pack."""

    source_runs: list[dict[str, object]] = []
    for source in _source_runs(run):
        if getattr(source, "job_type", None) != "udmi_validation":
            continue
        summary = getattr(source, "result_summary", None)
        summary = summary if isinstance(summary, dict) else {}
        ledger = summary.get("asset_topic_discovery")
        if not isinstance(ledger, dict):
            continue
        source_runs.append(
            {
                "source_run_id": str(getattr(source, "run_id", "")),
                "asset_topic_discovery": dict(ledger),
            }
        )
    return {"source_runs": source_runs} if source_runs else None


def _raw_evidence_export(run: object) -> dict[str, object]:
    """Build the redacted, portable raw-evidence area for a report ZIP."""
    report_parameters = run.parameters if isinstance(run.parameters, dict) else {}
    portable_snapshots = report_parameters.get("source_raw_evidence")
    portable_snapshots = portable_snapshots if isinstance(portable_snapshots, dict) else {}
    records: list[dict[str, object]] = []
    source_runs: list[dict[str, object]] = []
    captured_count = 0
    retained_bytes = 0
    truncated = False
    for source in _source_runs(run):
        if getattr(source, "job_type", None) != "udmi_validation":
            continue
        summary = getattr(source, "result_summary", None)
        summary = summary if isinstance(summary, dict) else {}
        raw = portable_snapshots.get(str(getattr(source, "run_id", "")))
        if not isinstance(raw, dict):
            raw = summary.get("raw_evidence")
        if not isinstance(raw, dict):
            continue
        source_id = str(getattr(source, "run_id", ""))
        raw_records = raw.get("records")
        if not isinstance(raw_records, list):
            raw_records = []
        source_runs.append(
            {
                "source_run_id": source_id,
                "record_count": len(raw_records),
                "captured_record_count": int(raw.get("captured_record_count") or 0),
                "truncated": bool(raw.get("truncated")),
            }
        )
        captured_count += int(raw.get("captured_record_count") or 0)
        retained_bytes += int(raw.get("retained_bytes") or 0)
        truncated = truncated or bool(raw.get("truncated"))
        for raw_record in raw_records:
            if not isinstance(raw_record, dict):
                continue
            record = dict(raw_record)
            record["source_run_id"] = source_id
            records.append(record)
    records.sort(
        key=lambda record: (
            str(record.get("source_run_id") or ""),
            str(record.get("asset_id") or ""),
            str(record.get("broker_received_at") or record.get("received_at") or ""),
            str(record.get("topic") or ""),
            str(record.get("content_sha256") or ""),
        )
    )
    for index, record in enumerate(records, start=1):
        fingerprint = hashlib.sha256(
            json.dumps(
                record,
                ensure_ascii=True,
                sort_keys=True,
                separators=(",", ":"),
            ).encode("utf-8")
        ).hexdigest()[:24]
        record["evidence_id"] = f"raw-{fingerprint}-{index:06d}"
    return redact_export_value(
        {
            "schema_version": "1.0",
            "evidence_id_field": "evidence_id",
            "source_runs": source_runs,
            "records": records,
            "record_count": len(records),
            "captured_record_count": captured_count,
            "retained_bytes": retained_bytes,
            "truncated": truncated,
            "redaction_policy": (
                "JSON bodies are recursively redacted by credential-shaped keys; "
                "non-JSON bodies are represented by hash and metadata only."
            ),
        }
    )


def _safe_evidence_file_component(value: object, *, fallback: str) -> str:
    """Return a deterministic, filesystem-safe evidence-file component."""

    text = str(value or "").strip()
    text = re.sub(r"[^A-Za-z0-9._-]+", "-", text).strip(".-")
    return (text[:48] or fallback).casefold()


def _raw_evidence_payload_files(
    records: list[dict[str, object]],
) -> tuple[list[tuple[str, bytes]], dict[str, object]]:
    """Project latest redacted JSON payloads into individually addressable files.

    The source record remains in records.jsonl for compatibility. The payload
    file body is the JSON payload itself, while the manifest carries its topic,
    asset, timestamp and integrity metadata. Non-JSON records are metadata-only
    because synthesising a JSON payload would misrepresent captured evidence.
    """

    files: list[tuple[str, bytes]] = []
    entries: list[dict[str, object]] = []
    used_names: set[str] = set()
    latest_by_topic: dict[tuple[str, str], dict[str, object]] = {}
    for record in records:
        source_run_id = str(record.get("source_run_id") or "source")
        topic = str(record.get("topic") or "")
        key = (source_run_id, topic or str(record.get("evidence_id") or "raw"))
        previous = latest_by_topic.get(key)
        # Latest evidence follows broker receipt order, never an embedded
        # payload timestamp that may be stale or supplied out of order.
        timestamp = str(record.get("broker_received_at") or record.get("received_at") or "")
        previous_timestamp = str(
            (previous or {}).get("broker_received_at") or (previous or {}).get("received_at") or ""
        )
        if previous is None or (timestamp, str(record.get("evidence_id") or "")) > (
            previous_timestamp,
            str(previous.get("evidence_id") or ""),
        ):
            latest_by_topic[key] = record

    for record in latest_by_topic.values():
        evidence_id = str(record.get("evidence_id") or "raw")
        source_run_id = str(record.get("source_run_id") or "source")
        asset_id = str(record.get("asset_id") or "asset")
        payload_type = str(record.get("payload_type") or "payload")
        topic = str(record.get("topic") or "")
        identity_hash = hashlib.sha256(
            f"{evidence_id}\0{source_run_id}\0{asset_id}\0{payload_type}\0{topic}".encode()
        ).hexdigest()
        base_name = "-".join(
            (
                _safe_evidence_file_component(source_run_id, fallback="source"),
                _safe_evidence_file_component(asset_id, fallback="asset"),
                _safe_evidence_file_component(payload_type, fallback="payload"),
                identity_hash[:20],
            )
        )
        payload = record.get("payload")
        is_json_payload = record.get("payload_encoding") == "json" and (
            payload is None or isinstance(payload, (dict, list, str, int, float, bool))
        )
        entry: dict[str, object] = {
            "evidence_id": evidence_id,
            "source_run_id": source_run_id,
            "topic": topic,
            "asset_id": record.get("asset_id"),
            "payload_type": record.get("payload_type"),
            "timestamp": record.get("payload_timestamp")
            or record.get("broker_received_at")
            or record.get("received_at"),
            "payload_timestamp": record.get("payload_timestamp"),
            "broker_received_at": record.get("broker_received_at") or record.get("received_at"),
            "content_sha256": record.get("content_sha256"),
        }
        if not is_json_payload:
            entry.update(
                {
                    "filename": None,
                    "byte_count": None,
                    "sha256": None,
                    "export_status": "metadata_only_non_json",
                }
            )
            entries.append(entry)
            continue
        filename = f"raw_evidence/payloads/{base_name}.json"
        suffix = 2
        while filename in used_names:
            filename = f"raw_evidence/payloads/{base_name}-{suffix}.json"
            suffix += 1
        used_names.add(filename)
        content = json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True).encode("utf-8")
        files.append((filename, content))
        entry.update(
            {
                "filename": filename,
                "byte_count": len(content),
                "sha256": hashlib.sha256(content).hexdigest(),
                "export_status": "payload_json",
            }
        )
        entries.append(entry)
    return files, {
        "schema_version": "1.0",
        "record_id_field": "evidence_id",
        "selection_policy": "Latest retained record per source run and concrete topic.",
        "redaction_policy": "Payload files contain redacted JSON only; non-JSON evidence is metadata-only.",
        "entries": entries,
    }


def _raw_finding_index(
    findings: list[dict[str, object]],
    raw_records: list[dict[str, object]],
) -> dict[str, object]:
    """Map each exported finding to portable raw-record ids.

    The URI is retained as provenance, but it is often a capture-scope URI
    rather than a file path. Asset, payload type, and topic are therefore used
    as deterministic fallback keys so a clean-machine unzip can still locate
    the relevant raw record without access to the original runtime database.
    """

    entries: list[dict[str, object]] = []
    for finding in findings:
        source_id = str(finding.get("source_run_id") or finding.get("Source Run") or "")
        issue_id = str(finding.get("issue_id") or finding.get("Issue ID") or "")
        asset_id = str(finding.get("asset_id") or finding.get("Asset") or "")
        payload_type = str(
            finding.get("payload_type") or finding.get("Payload") or ""
        ).strip().casefold()
        raw_uri = str(finding.get("raw_evidence_uri") or "")
        matching_ids: list[str] = []
        for record in raw_records:
            if str(record.get("source_run_id") or "") != source_id:
                continue
            record_asset = str(record.get("asset_id") or "")
            record_payload = str(record.get("payload_type") or "").strip().casefold()
            record_topic = str(record.get("topic") or "")
            uri_match = bool(raw_uri and record_topic and record_topic in raw_uri)
            identity_match = (
                (not asset_id or not record_asset or asset_id == record_asset)
                and (not payload_type or not record_payload or payload_type == record_payload)
            )
            if uri_match or identity_match:
                evidence_id = record.get("evidence_id")
                if isinstance(evidence_id, str) and evidence_id not in matching_ids:
                    matching_ids.append(evidence_id)
        entries.append(
            {
                "source_run_id": source_id,
                "issue_id": issue_id,
                "raw_evidence_uri": raw_uri or None,
                "evidence_ids": matching_ids,
                "match_status": "matched" if matching_ids else "unmatched",
            }
        )
    entries.sort(key=lambda entry: (str(entry["source_run_id"]), str(entry["issue_id"])))
    return {
        "schema_version": "1.0",
        "record_id_field": "evidence_id",
        "matching_policy": "source_run_id plus asset_id/payload_type, with topic URI match when available",
        "entries": entries,
    }


def _source_run_findings(run: object) -> list[dict[str, str]]:
    """Issues from the report's source runs, flattened + deterministically sorted.

    A report scoped to ``source_run_ids`` must carry the ACTUAL findings of those
    runs (not just an id label) to be a usable MSI handover deliverable. Terminal
    runs' issues are immutable, so the sorted output keeps the artifact
    byte-reproducible (required by the integrity verify endpoint).
    """
    findings: list[dict[str, str]] = []
    for source in _source_runs(run):
        for issue in source.issues:
            findings.append(
                {
                    "Source Run": str(source.run_id),
                    "Issue ID": str(getattr(issue, "issue_id", "") or ""),
                    "Asset": str(getattr(issue, "asset_id", "") or ""),
                    "Severity": str(getattr(issue, "severity", "") or ""),
                    "Type": str(getattr(issue, "issue_type", "") or ""),
                    "Point": str(getattr(issue, "point_name", "") or ""),
                    "Expected": str(getattr(issue, "expected_value", "") or ""),
                    "Observed": str(getattr(issue, "observed_value", "") or ""),
                    "Mismatch": (
                        "Yes"
                        if getattr(issue, "expected_value", None) is not None
                        and getattr(issue, "observed_value", None) is not None
                        and str(issue.expected_value)
                        != str(issue.observed_value)
                        else "No"
                    ),
                    "Topic Payload Type": "",
                    "Raw Evidence URI": str(getattr(issue, "raw_evidence_uri", "") or ""),
                    "Suggested Action": str(getattr(issue, "suggested_action", "") or ""),
                    "Description": str(getattr(issue, "description", "") or ""),
                }
            )
    findings.sort(key=lambda finding: (finding["Source Run"], finding["Issue ID"]))
    return findings


def _int_or_none(value: object) -> int | None:
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    return None


# Severities that block a clean handover. Mirrors core's
# udmi_validation._BLOCKING_SEVERITIES (critical + high/medium, which the
# workbench renders as "Fail") so a count derived from persisted issue records
# can never disagree with a run's own persisted blocking_issue_count.
_BLOCKING_SEVERITIES = frozenset({"critical", "high", "medium"})


def _blocking_issue_count(source: object) -> int:
    """Blocking issues derived from the run's own persisted issue records.

    Pre-upgrade runs never recorded blocking_issue_count, and without a count
    the ≤99 compliance clamp cannot fire — the report could print "100%
    (liveness)" beside critical findings.
    """
    return sum(
        1
        for issue in source.issues
        if str(getattr(issue, "severity", "") or "").casefold() in _BLOCKING_SEVERITIES
    )


def _validation_summary(run: object) -> dict[str, object] | None:
    """Summary + Silent-systems data for a report scoped to validation runs.

    Returns None when no source run is a validation run, so non-validation
    reports omit the validation sections entirely. Discovery source runs are
    excluded from the rows (they are inventory, not validation outcomes).
    """
    sources = [source for source in _source_runs(run) if source.job_type in VALIDATION_JOB_TYPES]
    if not sources:
        return None

    rows: list[dict[str, str]] = []
    silent_rows: list[dict[str, str]] = []
    total_expected = total_silent = total_blocking = 0
    # Device-percent units: overall % = floor(sum(percent_i * expected_i) / sum(expected_i)).
    weighted_percent_sum = 0
    scored_expected_sum = 0
    liveness_used = False
    for source in sources:
        summary = source.result_summary if isinstance(source.result_summary, dict) else {}
        expected = _int_or_none(summary.get("expected_devices"))
        publishing = _int_or_none(summary.get("publishing_seen"))
        silent = _int_or_none(summary.get("not_publishing"))
        blocking = _int_or_none(summary.get("blocking_issue_count"))
        if blocking is None:
            # Pre-upgrade run that never persisted the count: derive it from
            # the run's issue records so the ≤99 clamps below still fire.
            blocking = _blocking_issue_count(source)

        compliance = _NO_VALUE
        if "payload_conformance_percent" in summary:
            percent = _int_or_none(summary.get("payload_conformance_percent"))
            if percent is not None:
                compliance = f"{percent}%"
                if expected:
                    weighted_percent_sum += percent * expected
                    scored_expected_sum += expected
        elif expected and publishing is not None:
            # Pre-upgrade source run without the conformance fields: fall back to
            # publishing liveness and SAY SO — liveness is not conformance.
            percent = round(100 * publishing / expected)
            if blocking > 0:
                # Mirror core's per-run clamp: 100% beside a blocking issue is a lie.
                percent = min(percent, 99)
            compliance = f"{percent}% (liveness)"
            weighted_percent_sum += 100 * publishing
            scored_expected_sum += expected
            liveness_used = True

        rows.append(
            {
                "Source Run": str(source.run_id),
                "Type": str(source.job_type),
                "Status": str(source.status),
                "Expected Devices": str(expected) if expected is not None else _NO_VALUE,
                "Publishing": str(publishing) if publishing is not None else _NO_VALUE,
                "Silent": str(silent) if silent is not None else _NO_VALUE,
                "Blocking Issues": str(blocking),
                "Compliance %": compliance,
            }
        )
        total_expected += expected or 0
        total_silent += silent or 0
        total_blocking += blocking

        device_ids = summary.get("not_publishing_devices")
        if isinstance(device_ids, list):
            for device_id in sorted(str(device) for device in device_ids):
                silent_rows.append({"Source Run": str(source.run_id), "Device ID": device_id})
        elif (silent or 0) > 0:
            silent_rows.append({"Source Run": str(source.run_id), "Device ID": _SILENT_IDS_NOT_RECORDED})

    # Overall compliance = floor(100 * conforming-devices-sum / expected-devices-sum),
    # where each run contributes conforming ~= percent_i * expected_i / 100 devices
    # (publishing_seen for pre-upgrade liveness runs). min() across runs is dishonest
    # (one dirty run would mask nothing / one clean run everything) and an unweighted
    # mean of percents ignores fleet size; the device-weighted ratio does neither.
    if scored_expected_sum:
        overall_percent = weighted_percent_sum // scored_expected_sum
        if total_blocking > 0:
            # Mirror the per-run clamp: 100% next to a blocking issue is a lie.
            overall_percent = min(overall_percent, 99)
        overall_compliance = f"{overall_percent}%" + (" (liveness)" if liveness_used else "")
    else:
        overall_compliance = _NO_VALUE

    overall = {
        "Total Devices": total_expected,
        "Total Silent": total_silent,
        "Total Blocking Issues": total_blocking,
        "Overall Compliance %": overall_compliance,
    }
    overall_row = {
        "Source Run": "Overall",
        "Type": "",
        "Status": "",
        "Expected Devices": str(total_expected),
        "Publishing": "",
        "Silent": str(total_silent),
        "Blocking Issues": str(total_blocking),
        "Compliance %": overall_compliance,
    }
    overall_text = (
        f"Overall: {total_expected} devices, {total_silent} silent, "
        f"{total_blocking} blocking issues, {overall_compliance} compliance."
    )
    return {
        "rows": rows,
        "silent_rows": silent_rows,
        "overall": overall,
        "overall_row": overall_row,
        "overall_text": overall_text,
    }


def _inv_cell(value: object) -> str:
    """Render a single inventory cell: str() a real value, _NO_VALUE for blank/None.

    Never fabricates: an absent value or an empty list is shown as a blank, never
    a synthesised verdict (honesty rule).
    """
    if value is None:
        return _NO_VALUE
    text = str(value)
    return text if text else _NO_VALUE


def _inv_ports(value: object) -> str:
    """Join a persisted port list; an empty/absent list renders as _NO_VALUE.

    The engine stamps these lists (open_ports, forbidden_open_ports, …); the
    report renders the recorded fact, never a recomputed "fail" for an empty
    list (a host with no open ports is a recorded result).
    """
    if isinstance(value, list) and value:
        return ", ".join(str(port) for port in value)
    return _NO_VALUE


def _inv_count(value: object) -> str:
    """A count fragment for the Discovery-summary Counts cell.

    A missing key renders as _NO_VALUE — never an invented number.
    """
    count = _int_or_none(value)
    return str(count) if count is not None else _NO_VALUE


def _discovery_inventory(run: object) -> list[dict[str, object]] | None:
    """Ordered inventory sections for a report scoped to discovery runs.

    Returns None when no source run is a discovery run (non-discovery reports
    omit the inventory entirely), otherwise a list of section dicts
    ``{title, columns, rows, note}`` so every format builder is one generic loop.

    Determinism: device/point/topic rows inherit DiscoveryRepository's
    ``ORDER BY (position, id)`` over the immutable terminal-run rows; the only
    rows not already DB-ordered are the expected-not-responding rows, sorted here
    by (asset_id, device_instance, source run). No repository ``id`` or
    ``created_at`` field is ever placed in a rendered row — including one would
    break byte-reproducibility of the signed artifact.
    """
    sources = [source for source in _source_runs(run) if source.job_type in DISCOVERY_JOB_TYPES]
    if not sources:
        return None

    repo = DiscoveryRepository(service.engine)
    parameters = run.parameters if isinstance(run.parameters, dict) else {}
    frozen_rows = parameters.get("source_discovery_snapshots")

    def discovery_rows(source_run_id: str, name: str) -> list[dict[str, object]]:
        if isinstance(frozen_rows, dict):
            snapshot = frozen_rows.get(source_run_id)
            if isinstance(snapshot, dict):
                rows = snapshot.get(name)
                if isinstance(rows, list):
                    return [dict(item) for item in rows if isinstance(item, dict)]
        if name == "devices":
            return repo.list_devices(source_run_id)
        if name == "points":
            return repo.list_points(source_run_id)
        return repo.list_topics(source_run_id)

    summary_rows: list[dict[str, str]] = []
    ip_rows: list[dict[str, str]] = []
    device_rows: list[dict[str, str]] = []
    point_rows: list[dict[str, str]] = []
    topic_rows: list[dict[str, str]] = []
    # (sort key, row) so the silent rows can be ordered independently of scope.
    silent_entries: list[tuple[tuple[str, str, str], dict[str, str]]] = []

    has_ip = has_bacnet = has_mqtt = has_silent = False

    for source in sources:
        run_id = str(source.run_id)
        summary = source.result_summary if isinstance(source.result_summary, dict) else {}

        if source.job_type == "ip_discovery":
            has_ip = True
            devices = [
                device
                for device in discovery_rows(source.run_id, "devices")
                if device.get("device_type") == "ip_host"
            ]
            for device in devices:
                attributes = device.get("attributes") or {}
                ip_rows.append(
                    {
                        "Source Run": run_id,
                        "Address": _inv_cell(device.get("address")),
                        "Hostname": _inv_cell(device.get("name")),
                        "MAC": _inv_cell(attributes.get("mac_address")),
                        "Open Ports": _inv_ports(attributes.get("open_ports")),
                        "Forbidden Open": _inv_ports(attributes.get("forbidden_open_ports")),
                        "Unexpected Open": _inv_ports(attributes.get("unexpected_open_ports")),
                        "Missing Expected": _inv_ports(attributes.get("missing_expected_ports")),
                    }
                )
            counts = f"{_inv_count(summary.get('hosts_scanned'))} hosts scanned, "
            counts += f"{_inv_count(summary.get('hosts_responsive'))} responsive"

        elif source.job_type == "bacnet_discovery":
            has_bacnet = True
            devices = [
                device
                for device in discovery_rows(source.run_id, "devices")
                if device.get("device_type") == "bacnet_device"
            ]
            points = discovery_rows(source.run_id, "points")
            # Points per device: the point's device_ref is the device's asset_id.
            points_by_asset: dict[object, int] = {}
            for point in points:
                points_by_asset[point.get("device_ref")] = points_by_asset.get(point.get("device_ref"), 0) + 1
            for device in devices:
                attributes = device.get("attributes") or {}
                register = attributes.get("register_asset_name") or attributes.get("register_asset_id")
                device_rows.append(
                    {
                        "Source Run": run_id,
                        "Instance": _inv_cell(attributes.get("device_instance")),
                        "Address": _inv_cell(device.get("address")),
                        "Name": _inv_cell(device.get("name")),
                        "Vendor": _inv_cell(device.get("vendor")),
                        "Model": _inv_cell(device.get("model")),
                        "Register Asset": _inv_cell(register),
                        "Points": str(points_by_asset.get(attributes.get("asset_id"), 0)),
                    }
                )
            for point in points:
                attributes = point.get("attributes") or {}
                read_error = attributes.get("read_error")
                if read_error:
                    # A read failure is rendered as the error, never a value.
                    value = str(read_error)
                else:
                    observed = point.get("observed_value") or {}
                    value = _inv_cell(observed.get("value"))
                point_rows.append(
                    {
                        "Source Run": run_id,
                        "Device": _inv_cell(point.get("device_ref")),
                        "Point ID": _inv_cell(point.get("point_id")),
                        "Point Name": _inv_cell(point.get("point_name")),
                        "Value": value,
                        "Units": _inv_cell(point.get("units")),
                    }
                )
            counts = f"{len(devices)} devices, {len(points)} points read"
            if "expected_responding_count" in summary and "expected_device_count" in summary:
                responding = _inv_count(summary.get("expected_responding_count"))
                expected = _inv_count(summary.get("expected_device_count"))
                counts += f", {responding}/{expected} expected devices responding"
            # Pre-v0.1.12 bacnet runs never persisted expected_not_responding;
            # gate the silent section on the key's presence, not truthiness.
            if "expected_not_responding" in summary:
                has_silent = True
                silent_list = summary.get("expected_not_responding")
                if isinstance(silent_list, list):
                    for entry in silent_list:
                        if not isinstance(entry, dict):
                            continue
                        asset_id = entry.get("asset_id")
                        asset_name = entry.get("asset_name")
                        instance = entry.get("device_instance")
                        if asset_name and asset_id is not None:
                            register = f"{asset_name} ({asset_id})"
                        else:
                            register = asset_name or asset_id
                        directed = "sent" if entry.get("directed_probe_sent") else "not sent"
                        silent_entries.append(
                            (
                                (str(asset_id), str(instance), run_id),
                                {
                                    "Source Run": run_id,
                                    "Register Asset": _inv_cell(register),
                                    "Instance": _inv_cell(instance),
                                    "Address": _inv_cell(entry.get("address")),
                                    "Directed Who-Is": directed,
                                },
                            )
                        )

        else:  # mqtt_discovery
            has_mqtt = True
            for topic in discovery_rows(source.run_id, "topics"):
                attributes = topic.get("attributes") or {}
                # last_payload is deliberately excluded — non-tabular, and a
                # signed shareable artifact should not embed captured payloads.
                topic_rows.append(
                    {
                        "Source Run": run_id,
                        "Topic": _inv_cell(topic.get("topic")),
                        "Messages": _inv_cell(topic.get("message_count")),
                        "Device Ref": _inv_cell(attributes.get("device_ref")),
                    }
                )
            counts = f"{_inv_count(summary.get('topics_discovered'))} topics, "
            counts += f"{_inv_count(summary.get('messages_captured'))} messages"

        summary_rows.append(
            {
                "Source Run": run_id,
                "Type": str(source.job_type),
                "Status": str(source.status),
                "Counts": counts,
            }
        )

    silent_entries.sort(key=lambda item: item[0])
    silent_rows = [row for _, row in silent_entries]

    def _section(
        title: str,
        columns: tuple[str, ...],
        rows: list[dict[str, str]],
        *,
        note: str | None = None,
    ) -> dict[str, object]:
        # Non-silent sections with no rows carry the empty note so every builder
        # renders "empty scan is a recorded result" instead of a bare heading.
        effective_note = note if note is not None else (None if rows else _INVENTORY_EMPTY_NOTE)
        return {"title": title, "columns": columns, "rows": rows, "note": effective_note}

    sections: list[dict[str, object]] = [
        _section(_INVENTORY_SUMMARY_TITLE, _INVENTORY_SUMMARY_COLUMNS, summary_rows)
    ]
    if has_ip:
        sections.append(_section(_INVENTORY_IP_TITLE, _IP_INVENTORY_COLUMNS, ip_rows))
    if has_bacnet:
        sections.append(_section(_INVENTORY_BACNET_DEVICES_TITLE, _BACNET_DEVICE_COLUMNS, device_rows))
        sections.append(_section(_INVENTORY_BACNET_POINTS_TITLE, _BACNET_POINT_COLUMNS, point_rows))
    if has_silent:
        sections.append(
            _section(
                _INVENTORY_BACNET_SILENT_TITLE,
                _BACNET_SILENT_COLUMNS,
                silent_rows,
                note=_BACNET_SILENT_NOTE,
            )
        )
    if has_mqtt:
        sections.append(_section(_INVENTORY_MQTT_TITLE, _MQTT_TOPIC_COLUMNS, topic_rows))
    return sections


_UDMI_SYSTEM_ASSET_COLUMNS = (
    "System",
    "Expected Assets",
    "Observed Assets",
    "Not Observed Assets",
    "Assets With Issues",
    "Successfully Validated Assets",
    "Completion",
)
_UDMI_SYSTEM_PAYLOAD_COLUMNS = (
    "System",
    "Expected Payloads",
    "Received Payloads",
    "Not Received Payloads",
    "Payloads With Issues",
    "Successfully Validated Payloads",
    "Issues",
    "Warnings",
)
_UDMI_SYSTEM_FAULT_COLUMNS = ("System",) + tuple(label for _key, label in FAULT_METRIC_LABELS)
_UDMI_ASSET_COLUMNS = (
    "Asset ID",
    "System",
    "Observed",
    "All Payloads Received",
    "All Payloads Validated",
    "Evidence Timestamp",
)
_UDMI_FAULT_MATRIX_COLUMNS = (
    "Asset ID",
    "System",
    "Payload Formatting",
    "Missing Points",
    "Point Naming",
    "Additional Points",
    "Stale/Cadence",
    "Other",
)
_UDMI_FAULT_DETAIL_COLUMNS = (
    "Issue ID",
    "Asset ID",
    "System",
    "Payload",
    "Category",
    "Point",
    "Expected",
    "Observed",
    "Mismatch",
    "Suggested Action",
    "Description",
)
_UDMI_WRONG_TOPIC_COLUMNS = (
    "Source Run",
    "Asset ID",
    "System",
    "Payload",
    "Expected Topic",
    "Actual Topic",
    "Last Seen",
)
_UDMI_CATEGORY_LABELS = dict(FAULT_METRIC_LABELS)


def _content_weights(
    columns: tuple[str, ...],
    rows: list[dict[str, str]],
    *,
    minimum: int = 7,
    maximum: int = 48,
) -> tuple[float, ...]:
    """Size table columns from retained content while bounding extreme prose."""

    return tuple(
        float(
            max(
                minimum,
                min(
                    maximum,
                    max(
                        [len(column), *(len(str(row.get(column, ""))) for row in rows)],
                    ),
                ),
            )
        )
        for column in columns
    )


def _docx_content_widths(
    columns: tuple[str, ...],
    rows: list[dict[str, str]],
    *,
    total: int = 15120,
) -> tuple[int, ...]:
    weights = _content_weights(columns, rows)
    weight_total = sum(weights) or 1.0
    widths = [max(480, round(total * weight / weight_total)) for weight in weights]
    widths[-1] += total - sum(widths)
    return tuple(widths)


def _xlsx_content_widths(
    columns: tuple[str, ...],
    rows: list[dict[str, str]],
) -> dict[str, int]:
    weights = _content_weights(columns, rows, minimum=10, maximum=64)
    return {
        get_column_letter(index): int(weight) + 2
        for index, weight in enumerate(weights, start=1)
    }


def _yes_no(value: object) -> str:
    return "Yes" if value is True else "No"


def _completion(metrics: dict[str, object]) -> str:
    expected = metrics.get("expected")
    validated = metrics.get("successfully_validated")
    if not isinstance(expected, int) or isinstance(expected, bool) or expected <= 0:
        return "N/A"
    validated_count = validated if isinstance(validated, int) and not isinstance(validated, bool) else 0
    return f"{validated_count}/{expected} ({round(100 * validated_count / expected)}%)"


def _payload_correctness(metrics: dict[str, object]) -> tuple[str, str]:
    expected = metrics.get("expected")
    validated = metrics.get("successfully_validated")
    if not isinstance(expected, int) or isinstance(expected, bool) or expected <= 0:
        return "N/A", "N/A"
    validated_count = validated if isinstance(validated, int) and not isinstance(validated, bool) else 0
    incorrect_count = expected - validated_count
    correct = f"{validated_count}/{expected} ({round(100 * validated_count / expected)}%)"
    incorrect = f"{incorrect_count}/{expected} ({round(100 * incorrect_count / expected)}%)"
    return correct, incorrect


def _udmi_metric_rows(
    metrics: dict[str, int], labels: tuple[tuple[str, str], ...]
) -> list[dict[str, str]]:
    return [{"Metric": label, "Value": str(metrics.get(key, 0))} for key, label in labels]


def _udmi_supporting_metric_rows(
    run: object,
    data: dict[str, object],
) -> list[dict[str, str]]:
    rows = _udmi_metric_rows(data["issue_metrics"], ISSUE_METRIC_LABELS)
    payloads_correct, payloads_incorrect = _payload_correctness(data["payload_metrics"])
    rows.extend(
        [
            {
                "Metric": "Overall Compliance",
                "Value": _completion(data["asset_metrics"]),
            },
            {"Metric": "Payloads Correct %", "Value": payloads_correct},
            {"Metric": "Payloads Incorrect %", "Value": payloads_incorrect},
            {
                "Metric": "Last Validation Run",
                "Value": _display_timestamp(data.get("last_validation_run_at")),
            },
            {"Metric": "Report Generated", "Value": _display_timestamp(_generated_at(run))},
        ]
    )
    return rows


def _udmi_client_metric_rows(run: object, data: dict[str, object]) -> list[dict[str, str]]:
    """Small metrics-only projection shared by the client renderers."""

    rows: list[dict[str, str]] = []
    for section, labels, key in (
        ("Assets", ASSET_METRIC_LABELS, "asset_metrics"),
        ("Payloads", PAYLOAD_METRIC_LABELS, "payload_metrics"),
        ("Findings", FAULT_METRIC_LABELS, "fault_metrics"),
        ("Issues", ISSUE_METRIC_LABELS, "issue_metrics"),
    ):
        metrics = data.get(key)
        if not isinstance(metrics, dict):
            continue
        for row in _udmi_metric_rows(metrics, labels):
            rows.append({"Metric": f"{section}: {row['Metric']}", "Value": row["Value"]})
    rows.extend(
        [
            {"Metric": "Overall Compliance", "Value": _completion(data["asset_metrics"])},
            {"Metric": "Payloads Correct %", "Value": _payload_correctness(data["payload_metrics"])[0]},
            {"Metric": "Payloads Incorrect %", "Value": _payload_correctness(data["payload_metrics"])[1]},
            {"Metric": "Validation Scope", "Value": str(data.get("scope_status", "unknown"))},
            {
                "Metric": "Acceptance",
                "Value": "Eligible" if data.get("acceptance_eligible") is True else "Not eligible",
            },
            {
                "Metric": "Floor reporting",
                "Value": str(
                    (data.get("floor_reporting") or {}).get("status", "not_provided")
                    if isinstance(data.get("floor_reporting"), dict)
                    else "not_provided"
                ),
            },
            {"Metric": "Last Validation Run", "Value": _display_timestamp(data.get("last_validation_run_at"))},
            {"Metric": "Report Generated", "Value": _display_timestamp(_generated_at(run))},
        ]
    )
    return rows


def _udmi_system_tables(data: dict[str, object]) -> tuple[list[dict[str, str]], ...]:
    asset_rows: list[dict[str, str]] = []
    payload_rows: list[dict[str, str]] = []
    fault_rows: list[dict[str, str]] = []
    for raw in data["system_metrics"]:
        row = raw if isinstance(raw, dict) else {}
        system = str(row.get("system", "Unspecified"))
        asset = row.get("asset_metrics") if isinstance(row.get("asset_metrics"), dict) else {}
        payload = row.get("payload_metrics") if isinstance(row.get("payload_metrics"), dict) else {}
        fault = row.get("fault_metrics") if isinstance(row.get("fault_metrics"), dict) else {}
        issue = row.get("issue_metrics") if isinstance(row.get("issue_metrics"), dict) else {}
        asset_rows.append(
            {
                "System": system,
                **{label: str(asset.get(key, 0)) for key, label in ASSET_METRIC_LABELS},
                "Completion": _completion(asset),
            }
        )
        payload_rows.append(
            {
                "System": system,
                **{label: str(payload.get(key, 0)) for key, label in PAYLOAD_METRIC_LABELS},
                **{label: str(issue.get(key, 0)) for key, label in ISSUE_METRIC_LABELS},
            }
        )
        fault_rows.append(
            {
                "System": system,
                **{label: str(fault.get(key, 0)) for key, label in FAULT_METRIC_LABELS},
            }
        )
    return asset_rows, payload_rows, fault_rows


def _udmi_asset_rows(data: dict[str, object]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for raw in data["asset_results"]:
        asset = raw if isinstance(raw, dict) else {}
        rows.append(
            {
                "Asset ID": str(asset.get("asset_id", "")),
                "System": str(asset.get("system", "Unspecified")),
                "Observed": _yes_no(asset.get("observed")),
                "All Payloads Received": _yes_no(asset.get("all_expected_payloads_received")),
                # The report-level verdict requires every expected payload to
                # have arrived as well as the retained checks to pass. Using
                # all_received_payloads_successfully_validated alone would let
                # a missing payload render Received=No beside Validated=Yes.
                "All Payloads Validated": _yes_no(asset.get("successfully_validated")),
                # This is the latest retained evidence timestamp. It may be the
                # event time carried by a payload, so do not imply receipt time.
                "Evidence Timestamp": _display_timestamp(asset.get("last_observed_at")),
            }
        )
    return rows


def _udmi_wrong_topic_rows(data: dict[str, object]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    raw_assets = data.get("wrong_topic_assets")
    if not isinstance(raw_assets, list):
        return rows
    for raw_asset in raw_assets:
        asset = raw_asset if isinstance(raw_asset, dict) else {}
        payloads = asset.get("payloads")
        if not isinstance(payloads, list):
            continue
        for raw_payload in payloads:
            payload = raw_payload if isinstance(raw_payload, dict) else {}
            rows.append(
                {
                    "Source Run": str(asset.get("source_run_id", "")),
                    "Asset ID": str(asset.get("asset_id", "")),
                    "System": str(asset.get("system", "Unspecified")),
                    "Payload": str(payload.get("payload_type", "")),
                    "Expected Topic": str(payload.get("expected_topic", "")),
                    "Actual Topic": str(payload.get("actual_topic", "")),
                    "Last Seen": _display_timestamp(asset.get("last_seen")),
                }
            )
    return rows


def _udmi_annotated_register(
    data: dict[str, object],
) -> tuple[tuple[str, ...], list[dict[str, str]]]:
    raw_columns = data.get("annotated_register_columns")
    columns = tuple(
        str(column)
        for column in raw_columns
        if str(column)
    ) if isinstance(raw_columns, list) else ()
    raw_rows = data.get("annotated_register_rows")
    rows = (
        [
            {str(key): "" if value is None else str(value) for key, value in raw.items()}
            for raw in raw_rows
            if isinstance(raw, dict)
        ]
        if isinstance(raw_rows, list)
        else []
    )
    return columns, rows


def _udmi_fault_matrix_rows(data: dict[str, object]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for raw in data["fault_matrix"]:
        item = raw if isinstance(raw, dict) else {}
        rows.append(
            {
                "Asset ID": str(item.get("asset_id", "")),
                "System": str(item.get("system", "Unspecified")),
                "Payload Formatting": _yes_no(item.get("payload_formatting_issues")),
                "Missing Points": _yes_no(item.get("missing_points")),
                "Point Naming": _yes_no(item.get("point_naming_issues")),
                "Additional Points": _yes_no(item.get("additional_points")),
                "Stale/Cadence": _yes_no(item.get("stale_or_cadence")),
                "Other": _yes_no(item.get("other_issues")),
            }
        )
    return rows


def _udmi_fault_detail_rows(data: dict[str, object]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for raw in data["fault_rows"]:
        fault = raw if isinstance(raw, dict) else {}
        category = str(fault.get("category", "other_issues"))
        rows.append(
            {
                "Issue ID": str(fault.get("issue_id", "")),
                "Asset ID": str(fault.get("asset_id") or "Run-wide"),
                "System": str(fault.get("system", "Unspecified")),
                "Payload": str(fault.get("payload_type", "")),
                "Category": _UDMI_CATEGORY_LABELS.get(category, "Other Issues"),
                "Point": str(fault.get("point_name") or ""),
                "Expected": str(fault.get("expected_value") or ""),
                "Observed": str(fault.get("observed_value") or ""),
                "Mismatch": "Yes" if fault.get("mismatch") is True else "No",
                "Suggested Action": str(fault.get("suggested_action") or ""),
                "Description": str(fault.get("description") or ""),
            }
        )
    return rows


def _udmi_fault_json_rows(data: dict[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for raw in data["fault_rows"]:
        fault = raw if isinstance(raw, dict) else {}
        rows.append(
            {
                "issue_id": str(fault.get("issue_id", "")),
                "asset_id": fault.get("asset_id"),
                "system": str(fault.get("system", "Unspecified")),
                "payload_type": str(fault.get("payload_type", "")),
                "category": str(fault.get("category", "other_issues")),
                "point_name": fault.get("point_name"),
                "expected_value": fault.get("expected_value"),
                "observed_value": fault.get("observed_value"),
                "mismatch": fault.get("mismatch") is True,
                "suggested_action": fault.get("suggested_action"),
                "description": str(fault.get("description", "")),
            }
        )
    return rows


def _udmi_audit_finding_rows(data: dict[str, object]) -> list[dict[str, object]]:
    """Scoped machine-audit findings, retaining provenance omitted from tables."""

    rows: list[dict[str, object]] = []
    for raw in data["fault_rows"]:
        fault = raw if isinstance(raw, dict) else {}
        rows.append(
            {
                "source_run_id": str(fault.get("source_run_id", "")),
                "issue_id": str(fault.get("issue_id", "")),
                "asset_id": fault.get("asset_id"),
                "system": str(fault.get("system", "Unspecified")),
                "payload_type": str(fault.get("payload_type", "")),
                "category": str(fault.get("category", "other_issues")),
                "severity": str(fault.get("severity", "")),
                "point_name": fault.get("point_name"),
                "expected_value": fault.get("expected_value"),
                "observed_value": fault.get("observed_value"),
                "mismatch": fault.get("mismatch") is True,
                "topic_compliance_payload_type": fault.get("topic_compliance_payload_type"),
                "suggested_action": fault.get("suggested_action"),
                "description": str(fault.get("description", "")),
                "raw_evidence_uri": fault.get("raw_evidence_uri"),
            }
        )
    return rows


_XLSX_DARK_FILL = PatternFill("solid", fgColor="203746")
_XLSX_ACCENT_FILL = PatternFill("solid", fgColor="2A7B9B")
_XLSX_POSITIVE_FILL = PatternFill("solid", fgColor="E2F0D9")
_XLSX_CAUTION_FILL = PatternFill("solid", fgColor="FFF2CC")
_XLSX_NEGATIVE_FILL = PatternFill("solid", fgColor="FCE4D6")
_XLSX_WHITE_FONT = Font(color="FFFFFF", bold=True)
_XLSX_GRID_SIDE = Side(style="thin", color="B0C1C9")
_XLSX_GRID_BORDER = Border(
    left=_XLSX_GRID_SIDE,
    right=_XLSX_GRID_SIDE,
    top=_XLSX_GRID_SIDE,
    bottom=_XLSX_GRID_SIDE,
)
_XML_ILLEGAL_CHAR_RE = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F\uD800-\uDFFF\uFFFE\uFFFF]"
)


def _sanitize_report_text(value: object) -> str:
    """Remove XML 1.0-illegal characters while preserving tab/newline/CR."""

    return _XML_ILLEGAL_CHAR_RE.sub("", str(value))


def _xlsx_safe_text(value: object) -> object:
    """Keep untrusted strings as literal cells, never spreadsheet formulas."""

    if isinstance(value, str):
        value = _sanitize_report_text(value)
        candidate = value.lstrip(" \t\r\n")
        if candidate.startswith(("=", "+", "-", "@")):
            return f"'{value}"
    return value


def _append_xlsx_row(sheet: object, values: list[object] | tuple[object, ...]) -> None:
    sheet.append([_xlsx_safe_text(value) for value in values])


def _configure_xlsx_sheet(
    sheet: object,
    *,
    widths: dict[str, int],
    freeze_panes: str | None = "A2",
    auto_filter: bool = True,
) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.freeze_panes = freeze_panes
    sheet.print_title_rows = "1:1"
    sheet.sheet_view.showGridLines = False
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    if auto_filter and sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions


def _style_xlsx_header(sheet: object, row_number: int = 1) -> None:
    for cell in sheet[row_number]:
        if cell.value is None:
            continue
        cell.fill = _XLSX_DARK_FILL
        cell.font = _XLSX_WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _XLSX_GRID_BORDER
    sheet.row_dimensions[row_number].height = 30


def _style_xlsx_table(
    sheet: object,
    *,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
) -> None:
    final_row = max_row if max_row is not None else sheet.max_row
    final_col = max_col if max_col is not None else sheet.max_column
    for row_number in range(min_row, final_row + 1):
        estimated_lines = 1
        for column_number in range(min_col, final_col + 1):
            cell = sheet.cell(row=row_number, column=column_number)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = _XLSX_GRID_BORDER
            width = sheet.column_dimensions[get_column_letter(column_number)].width or 10
            value_length = len(str(cell.value or ""))
            estimated_lines = max(
                estimated_lines,
                max(1, (value_length + max(int(width), 1) - 1) // max(int(width), 1)),
            )
        sheet.row_dimensions[row_number].height = min(150, max(20, estimated_lines * 15))


def _style_xlsx_statuses(sheet: object) -> None:
    headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value is not None}
    for heading in ("Observed", "All Payloads Received", "All Payloads Validated"):
        column = headers.get(heading)
        if column is None:
            continue
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=column)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.fill = _XLSX_POSITIVE_FILL if cell.value == "Yes" else _XLSX_CAUTION_FILL
    topic_status_column = headers.get("Topic status")
    if topic_status_column is not None:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=topic_status_column)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            if cell.value == "Expected topic":
                cell.fill = _XLSX_POSITIVE_FILL
            elif cell.value == "Wrong topic":
                cell.fill = _XLSX_CAUTION_FILL
    for heading in (
        "Payload Formatting",
        "Missing Points",
        "Point Naming",
        "Additional Points",
        "Stale/Cadence",
        "Other",
    ):
        column = headers.get(heading)
        if column is None:
            continue
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=column)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.fill = _XLSX_NEGATIVE_FILL if cell.value == "Yes" else _XLSX_POSITIVE_FILL


def _append_xlsx_summary_section(
    sheet: object,
    title: str,
    rows: list[dict[str, str]],
) -> None:
    _append_xlsx_row(sheet, [])
    title_row = sheet.max_row + 1
    _append_xlsx_row(sheet, [title])
    sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=2)
    title_cell = sheet.cell(title_row, 1)
    title_cell.fill = _XLSX_ACCENT_FILL
    title_cell.font = _XLSX_WHITE_FONT
    _append_xlsx_row(sheet, ["Metric", "Value"])
    header_row = sheet.max_row
    _style_xlsx_header(sheet, header_row)
    for row in rows:
        _append_xlsx_row(sheet, [row["Metric"], row["Value"]])
    _style_xlsx_table(
        sheet,
        min_row=header_row,
        max_row=sheet.max_row,
        min_col=1,
        max_col=2,
    )
    _style_xlsx_header(sheet, header_row)


def _build_udmi_client_xlsx_report(run: object, data: dict[str, object]) -> bytes:
    workbook = Workbook()
    workbook.properties.created = _ARTIFACT_PROPERTIES_EPOCH
    workbook.properties.modified = _ARTIFACT_PROPERTIES_EPOCH
    sheet = workbook.active
    sheet.title = "Client Metrics"
    _append_xlsx_row(sheet, ["Metric", "Value"])
    for row in _udmi_client_metric_rows(run, data):
        _append_xlsx_row(sheet, [row["Metric"], row["Value"]])
    _configure_xlsx_sheet(
        sheet,
        widths={"A": 42, "B": 34},
        freeze_panes="A2",
        auto_filter=True,
    )
    _style_xlsx_table(sheet)
    _style_xlsx_header(sheet)
    _apply_xlsx_branding(workbook, str(run.run_id), _report_title(run))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_udmi_xlsx_report(run: object, data: dict[str, object]) -> bytes:
    if _udmi_report_variant(run) == "client":
        return _build_udmi_client_xlsx_report(run, data)
    workbook = Workbook()
    workbook.properties.created = _ARTIFACT_PROPERTIES_EPOCH
    workbook.properties.modified = _ARTIFACT_PROPERTIES_EPOCH
    title = _report_title(run)

    definition_rows = [
        {
            "Metric": str(definition["metric"]),
            "Definition": str(definition["definition"]),
        }
        for definition in data["metric_definitions"]
    ]
    definitions = workbook.active
    definitions.title = _UDMI_DEFINITIONS_TITLE
    _append_xlsx_row(definitions, ["Metric", "Definition"])
    for row in definition_rows:
        _append_xlsx_row(definitions, [row["Metric"], row["Definition"]])
    _configure_xlsx_sheet(
        definitions,
        widths=_xlsx_content_widths(("Metric", "Definition"), definition_rows),
    )
    _style_xlsx_table(definitions)
    _style_xlsx_header(definitions)

    executive = workbook.create_sheet(_UDMI_EXECUTIVE_TITLE)
    executive.title = _UDMI_EXECUTIVE_TITLE
    _append_xlsx_row(executive, [title])
    executive.merge_cells("A1:B1")
    executive["A1"].fill = _XLSX_DARK_FILL
    executive["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    executive["A1"].alignment = Alignment(horizontal="center", vertical="center")
    executive.row_dimensions[1].height = 34
    for label, value in _report_rows(run, udmi_data=data):
        _append_xlsx_row(executive, [label, value])
    if data["scope_complete"] is not True:
        warning_row = executive.max_row + 1
        _append_xlsx_row(
            executive,
            [_UDMI_INCOMPLETE_SCOPE_TITLE, str(data["scope_summary"])],
        )
        for cell in executive[warning_row]:
            cell.fill = _XLSX_NEGATIVE_FILL
            cell.font = Font(bold=True, color="9C0006")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    asset_metrics = data["asset_metrics"]
    payload_metrics = data["payload_metrics"]
    fault_metrics = data["fault_metrics"]
    _append_xlsx_summary_section(
        executive,
        "Asset Level Metrics",
        _udmi_metric_rows(asset_metrics, ASSET_METRIC_LABELS),
    )
    _append_xlsx_summary_section(
        executive,
        "Payload Level Metrics",
        _udmi_metric_rows(payload_metrics, PAYLOAD_METRIC_LABELS),
    )
    _append_xlsx_summary_section(
        executive,
        "Fault Metrics",
        _udmi_metric_rows(fault_metrics, FAULT_METRIC_LABELS),
    )
    supporting = _udmi_supporting_metric_rows(run, data)
    _append_xlsx_summary_section(executive, "Supporting Metrics", supporting)
    for note in data["notes"]:
        _append_xlsx_row(executive, ["Note", str(note)])
    _configure_xlsx_sheet(
        executive,
        widths={"A": 34, "B": 92},
        freeze_panes="A2",
        auto_filter=False,
    )

    system_assets, system_payloads, system_faults = _udmi_system_tables(data)
    systems = workbook.create_sheet(_UDMI_SYSTEM_TITLE)
    system_columns = (
        _UDMI_SYSTEM_ASSET_COLUMNS
        + _UDMI_SYSTEM_PAYLOAD_COLUMNS[1:]
        + _UDMI_SYSTEM_FAULT_COLUMNS[1:]
    )
    _append_xlsx_row(systems, list(system_columns))
    payload_by_system = {row["System"]: row for row in system_payloads}
    fault_by_system = {row["System"]: row for row in system_faults}
    for asset_row in system_assets:
        system = asset_row["System"]
        combined = {**asset_row, **payload_by_system[system], **fault_by_system[system]}
        _append_xlsx_row(systems, [combined[column] for column in system_columns])
    _configure_xlsx_sheet(
        systems,
        widths=_xlsx_content_widths(system_columns, [
            {column: str(combined[column]) for column in system_columns}
            for asset_row in system_assets
            for system in [asset_row["System"]]
            for combined in [{**asset_row, **payload_by_system[system], **fault_by_system[system]}]
        ]),
    )
    _style_xlsx_table(systems)
    _style_xlsx_header(systems)

    asset_rows = _udmi_asset_rows(data)
    assets = workbook.create_sheet(_UDMI_ASSET_TITLE)
    _append_xlsx_row(assets, list(_UDMI_ASSET_COLUMNS))
    for row in asset_rows:
        _append_xlsx_row(assets, [row[column] for column in _UDMI_ASSET_COLUMNS])
    _configure_xlsx_sheet(
        assets,
        widths=_xlsx_content_widths(_UDMI_ASSET_COLUMNS, asset_rows),
    )
    _style_xlsx_table(assets)
    _style_xlsx_header(assets)
    _style_xlsx_statuses(assets)

    wrong_topic_rows = _udmi_wrong_topic_rows(data)
    if wrong_topic_rows:
        wrong_topics = workbook.create_sheet(_UDMI_WRONG_TOPIC_SHEET_TITLE)
        _append_xlsx_row(wrong_topics, list(_UDMI_WRONG_TOPIC_COLUMNS))
        for row in wrong_topic_rows:
            _append_xlsx_row(
                wrong_topics,
                [row[column] for column in _UDMI_WRONG_TOPIC_COLUMNS],
            )
        _configure_xlsx_sheet(
            wrong_topics,
            widths=_xlsx_content_widths(_UDMI_WRONG_TOPIC_COLUMNS, wrong_topic_rows),
        )
        _style_xlsx_table(wrong_topics)
        _style_xlsx_header(wrong_topics)

    register_columns, register_rows = _udmi_annotated_register(data)
    if register_columns and register_rows:
        annotated = workbook.create_sheet(_UDMI_ANNOTATED_REGISTER_TITLE)
        _append_xlsx_row(annotated, list(register_columns))
        for row in register_rows:
            _append_xlsx_row(annotated, [row.get(column, "") for column in register_columns])
        _configure_xlsx_sheet(
            annotated,
            widths=_xlsx_content_widths(register_columns, register_rows),
        )
        _style_xlsx_table(annotated)
        _style_xlsx_header(annotated)
        _style_xlsx_statuses(annotated)

    matrix_rows = _udmi_fault_matrix_rows(data)
    matrix = workbook.create_sheet(_UDMI_FAULT_MATRIX_TITLE)
    _append_xlsx_row(matrix, list(_UDMI_FAULT_MATRIX_COLUMNS))
    for row in matrix_rows:
        _append_xlsx_row(matrix, [row[column] for column in _UDMI_FAULT_MATRIX_COLUMNS])
    _configure_xlsx_sheet(
        matrix,
        widths=_xlsx_content_widths(_UDMI_FAULT_MATRIX_COLUMNS, matrix_rows),
    )
    _style_xlsx_table(matrix)
    _style_xlsx_header(matrix)
    _style_xlsx_statuses(matrix)

    detail_rows = _udmi_fault_detail_rows(data)
    details = workbook.create_sheet(_UDMI_FAULT_DETAIL_TITLE)
    _append_xlsx_row(details, list(_UDMI_FAULT_DETAIL_COLUMNS))
    for row in detail_rows:
        _append_xlsx_row(details, [row[column] for column in _UDMI_FAULT_DETAIL_COLUMNS])
    _configure_xlsx_sheet(
        details,
        widths=_xlsx_content_widths(_UDMI_FAULT_DETAIL_COLUMNS, detail_rows),
    )
    _style_xlsx_table(details)
    _style_xlsx_header(details)

    _apply_xlsx_branding(workbook, str(run.run_id), title)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _apply_xlsx_branding(
    workbook: Workbook,
    run_id: str,
    document_title: str = _BRAND_DOC_TITLE,
) -> None:
    """Apply the text-only branding band to every sheet's page header/footer.

    Header: wordmark (left) + document title (right). Footer: wordmark (left) +
    "Page N of M" (center) + run id (right). These surface in Excel's Page
    Layout view and on print — exactly the ITP-pack use. openpyxl serializes
    header/footer as a static <headerFooter> element per sheet, so the bytes are
    deterministic and survive _normalize_zip_bytes unchanged. Called once, after
    all sheets exist, so every sheet is covered regardless of which conditional
    validation sections were created.
    """
    safe_run_id = _sanitize_report_text(run_id)
    safe_document_title = _sanitize_report_text(document_title)
    for sheet in workbook.worksheets:
        sheet.oddHeader.left.text = _BRAND_NAME
        sheet.oddHeader.right.text = safe_document_title
        sheet.oddFooter.left.text = _BRAND_NAME
        # openpyxl's friendly page tokens: &[Page] -> &P, &[Pages] would -> &N;
        # &N is written directly. Serialized into each sheet XML as &amp;P/&amp;N.
        sheet.oddFooter.center.text = "Page &[Page] of &N"
        sheet.oddFooter.right.text = safe_run_id


def _build_xlsx_report(run: object) -> bytes:
    udmi = _udmi_report_data(run)
    if udmi is not None:
        return _build_udmi_xlsx_report(run, udmi)

    workbook = Workbook()
    # openpyxl stamps docProps/core.xml with the current time on save; pin the
    # core properties to a fixed instant so the artifact bytes are reproducible
    # from the run record (required for hash-based verification).
    workbook.properties.created = _ARTIFACT_PROPERTIES_EPOCH
    workbook.properties.modified = _ARTIFACT_PROPERTIES_EPOCH
    sheet = workbook.active
    sheet.title = "Report Summary"
    _append_xlsx_row(sheet, ["Field", "Value"])
    for row in _report_rows(run):
        _append_xlsx_row(sheet, list(row))
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 56
    # Validation sections (Summary / Silent systems) only when the scoped source
    # runs include validation runs; other report types keep their prior shape.
    validation = _validation_summary(run)
    if validation is not None:
        summary_sheet = workbook.create_sheet(_SUMMARY_SECTION_TITLE)
        _append_xlsx_row(summary_sheet, list(_VALIDATION_SUMMARY_COLUMNS))
        for row in validation["rows"]:
            _append_xlsx_row(
                summary_sheet,
                [row[column] for column in _VALIDATION_SUMMARY_COLUMNS],
            )
        _append_xlsx_row(
            summary_sheet,
            [validation["overall_row"][column] for column in _VALIDATION_SUMMARY_COLUMNS],
        )
        for column, width in {"A": 26, "B": 18, "C": 12, "D": 16, "E": 12, "F": 10, "G": 16, "H": 18}.items():
            summary_sheet.column_dimensions[column].width = width
    # Discovery inventory sheets (one per section) when the scoped source runs
    # include discovery runs; other report types keep their prior shape.
    inventory = _discovery_inventory(run)
    if inventory is not None:
        for section in inventory:
            columns = section["columns"]
            sheet_title = _INVENTORY_SHEET_NAMES.get(section["title"], section["title"])
            inventory_sheet = workbook.create_sheet(sheet_title)
            # A remapped (truncated) sheet name loses the full title — surface it
            # as row 1 so the abbreviation is never the only record of the head.
            if sheet_title != section["title"]:
                _append_xlsx_row(inventory_sheet, [section["title"]])
            if section["note"]:
                _append_xlsx_row(inventory_sheet, [section["note"]])
            _append_xlsx_row(inventory_sheet, list(columns))
            for row in section["rows"]:
                _append_xlsx_row(inventory_sheet, [row[column] for column in columns])
            for index, column in enumerate(columns, start=1):
                width = _INVENTORY_COLUMN_WIDTHS.get(column)
                if width:
                    inventory_sheet.column_dimensions[get_column_letter(index)].width = width
    # Failure detail: findings from the scoped source runs (the actual report
    # content, not just the metadata above). Empty source runs -> header-only.
    findings = _source_run_findings(run)
    findings_sheet = workbook.create_sheet(_FAILURE_SECTION_TITLE)
    _append_xlsx_row(findings_sheet, list(_FINDING_COLUMNS))
    for finding in findings:
        _append_xlsx_row(findings_sheet, [finding[column] for column in _FINDING_COLUMNS])
    findings_widths = {"A": 26, "B": 16, "C": 18, "D": 12, "E": 22, "F": 24, "G": 18, "H": 18, "I": 40, "J": 70}
    for column, width in findings_widths.items():
        findings_sheet.column_dimensions[column].width = width
    if validation is not None:
        silent_sheet = workbook.create_sheet(_SILENT_SECTION_TITLE)
        _append_xlsx_row(silent_sheet, [_SILENT_NOTE])
        _append_xlsx_row(silent_sheet, list(_SILENT_COLUMNS))
        for row in validation["silent_rows"]:
            _append_xlsx_row(silent_sheet, [row[column] for column in _SILENT_COLUMNS])
        silent_sheet.column_dimensions["A"].width = 30
        silent_sheet.column_dimensions["B"].width = 46
    # Branding must run after every create_sheet so all sheets carry the band.
    _apply_xlsx_branding(workbook, str(run.run_id), _report_title(run))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _docx_paragraph(
    text: str,
    *,
    bold: bool = False,
    alignment: str | None = None,
    page_break_before: bool = False,
    keep_with_next: bool = False,
) -> str:
    run_properties = "<w:rPr><w:b/></w:rPr>" if bold else ""
    paragraph_settings = ""
    if page_break_before:
        paragraph_settings += "<w:pageBreakBefore/>"
    if keep_with_next:
        paragraph_settings += "<w:keepNext/>"
    if alignment:
        paragraph_settings += f'<w:jc w:val="{alignment}"/>'
    paragraph_properties = (
        f"<w:pPr>{paragraph_settings}</w:pPr>" if paragraph_settings else ""
    )
    safe_text = _sanitize_report_text(text)
    return (
        f'<w:p>{paragraph_properties}<w:r>{run_properties}'
        f'<w:t xml:space="preserve">{escape(safe_text)}</w:t></w:r></w:p>'
    )


# Single hairline borders so the hand-rolled tables read as tables in Word.
_DOCX_TABLE_BORDERS = "".join(
    f'<w:{edge} w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
    for edge in ("top", "left", "bottom", "right", "insideH", "insideV")
)


def _docx_table(
    columns: tuple[str, ...],
    rows: list[dict[str, str]],
    *,
    widths: tuple[int, ...] | None = None,
    center_cells: bool = False,
) -> str:
    effective_widths = (
        widths if widths is not None and len(widths) == len(columns) else None
    )

    def cell(text: str, *, bold: bool = False, width: int | None = None) -> str:
        width_property = f'<w:tcW w:w="{width}" w:type="dxa"/>' if width else ""
        cell_properties = f"<w:tcPr>{width_property}<w:vAlign w:val=\"center\"/></w:tcPr>"
        paragraph = _docx_paragraph(
            text,
            bold=bold,
            alignment="center" if center_cells else None,
        )
        return (
            f"<w:tc>{cell_properties}"
            f"{paragraph}"
            "</w:tc>"
        )

    grid = "".join(
        f'<w:gridCol w:w="{effective_widths[index]}"/>' if effective_widths else "<w:gridCol/>"
        for index in range(len(columns))
    )
    header = (
        "<w:tr><w:trPr><w:tblHeader/><w:cantSplit/></w:trPr>"
        + "".join(
            cell(
                column,
                bold=True,
                width=effective_widths[index] if effective_widths else None,
            )
            for index, column in enumerate(columns)
        )
        + "</w:tr>"
    )
    body = "".join(
        "<w:tr><w:trPr><w:cantSplit/></w:trPr>"
        + "".join(
            cell(
                row.get(column, ""),
                width=effective_widths[index] if effective_widths else None,
            )
            for index, column in enumerate(columns)
        )
        + "</w:tr>"
        for row in rows
    )
    return (
        f'<w:tbl><w:tblPr><w:tblW w:w="0" w:type="auto"/><w:tblBorders>{_DOCX_TABLE_BORDERS}</w:tblBorders>'
        f"</w:tblPr><w:tblGrid>{grid}</w:tblGrid>{header}{body}</w:tbl>"
    )


# --- DOCX branding parts (real OOXML header/footer) --------------------------
# Word requires the header/footer as separate parts referenced from the section
# properties via relationship ids. The header is a constant; the footer carries
# the run id so it is built per run. fldSimple PAGE/NUMPAGES fields hold a "1"
# placeholder that Word recomputes on open/print (kept static so the bytes stay
# reproducible — do NOT precompute it into a real page count).
def _build_docx_header_xml(document_title: str, *, landscape: bool = False) -> str:
    right_tab = 15120 if landscape else 9026
    safe_document_title = _sanitize_report_text(document_title)
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:hdr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:p><w:pPr>"
        '<w:pBdr><w:bottom w:val="single" w:sz="4" w:space="1" w:color="auto"/></w:pBdr>'
        f'<w:tabs><w:tab w:val="right" w:pos="{right_tab}"/></w:tabs>'
        "</w:pPr>"
        f'<w:r><w:rPr><w:b/></w:rPr><w:t xml:space="preserve">{escape(_BRAND_NAME)}</w:t></w:r>'
        "<w:r><w:tab/></w:r>"
        f'<w:r><w:t xml:space="preserve">{escape(safe_document_title)}</w:t></w:r>'
        "</w:p></w:hdr>"
    )


_DOCX_HEADER_XML = _build_docx_header_xml(_BRAND_DOC_TITLE)

# Two relationships from word/document.xml to the header and footer parts. This
# file does not exist in the base 3-member docx — it must be created.
_DOCX_DOCUMENT_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rIdHdr1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/header" Target="header1.xml"/>
  <Relationship Id="rIdFtr1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/footer" Target="footer1.xml"/>
</Relationships>"""

# Section properties wiring the header/footer references in and declaring A4 (to
# match the PDF; a sectPr-less docx defaults to Letter in Word). Child order —
# headerReference, footerReference, pgSz, pgMar — follows the ECMA-376 sequence.
def _docx_sectpr(*, landscape: bool = False) -> str:
    page_size = (
        '<w:pgSz w:w="16838" w:h="11906" w:orient="landscape"/>'
        if landscape
        else '<w:pgSz w:w="11906" w:h="16838"/>'
    )
    horizontal_margin = 720 if landscape else 1440
    top_margin = 720 if landscape else 1440
    bottom_margin = 1080 if landscape else 1440
    return (
        "<w:sectPr>"
        '<w:headerReference w:type="default" r:id="rIdHdr1"/>'
        '<w:footerReference w:type="default" r:id="rIdFtr1"/>'
        f"{page_size}"
        f'<w:pgMar w:top="{top_margin}" w:right="{horizontal_margin}" '
        f'w:bottom="{bottom_margin}" w:left="{horizontal_margin}" '
        'w:header="540" w:footer="540" w:gutter="0"/>'
        "</w:sectPr>"
    )


_DOCX_SECTPR = _docx_sectpr()


def _build_docx_footer_xml(run_id: str, *, landscape: bool = False) -> str:
    """word/footer1.xml: wordmark + Page N of M (PAGE/NUMPAGES fields) + run id."""
    center_tab = 7560 if landscape else 4513
    right_tab = 15120 if landscape else 9026
    safe_run_id = _sanitize_report_text(run_id)
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:ftr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:p><w:pPr>"
        f'<w:tabs><w:tab w:val="center" w:pos="{center_tab}"/>'
        f'<w:tab w:val="right" w:pos="{right_tab}"/></w:tabs>'
        "</w:pPr>"
        f'<w:r><w:t xml:space="preserve">{escape(_BRAND_NAME)}</w:t></w:r>'
        "<w:r><w:tab/></w:r>"
        '<w:r><w:t xml:space="preserve">Page </w:t></w:r>'
        '<w:fldSimple w:instr=" PAGE "><w:r><w:t>1</w:t></w:r></w:fldSimple>'
        '<w:r><w:t xml:space="preserve"> of </w:t></w:r>'
        '<w:fldSimple w:instr=" NUMPAGES "><w:r><w:t>1</w:t></w:r></w:fldSimple>'
        "<w:r><w:tab/></w:r>"
        f'<w:r><w:t xml:space="preserve">{escape(safe_run_id)}</w:t></w:r>'
        "</w:p></w:ftr>"
    )


def _docx_page_break() -> str:
    return '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'


def _assemble_docx(
    run: object,
    blocks: list[str],
    *,
    document_title: str,
    landscape: bool,
) -> bytes:
    body_xml = "\n    ".join(blocks)
    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <w:body>
    {body_xml}
    {_docx_sectpr(landscape=landscape)}
  </w:body>
</w:document>"""
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
  <Override PartName="/word/header1.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.header+xml"/>
  <Override PartName="/word/footer1.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.footer+xml"/>
</Types>""",
        )
        archive.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>""",
        )
        archive.writestr("word/document.xml", document_xml)
        archive.writestr("word/_rels/document.xml.rels", _DOCX_DOCUMENT_RELS)
        archive.writestr(
            "word/header1.xml",
            _build_docx_header_xml(document_title, landscape=landscape),
        )
        archive.writestr(
            "word/footer1.xml",
            _build_docx_footer_xml(str(run.run_id), landscape=landscape),
        )
    return buffer.getvalue()


def _build_udmi_client_docx_report(run: object, data: dict[str, object]) -> bytes:
    title = _report_title(run)
    blocks: list[str] = [_docx_paragraph(title, bold=True)]
    blocks.extend(
        _docx_paragraph(f"{label}: {value}")
        for label, value in _report_rows(run, udmi_data=data)
    )
    if data["scope_complete"] is not True:
        blocks.append(_docx_paragraph(_UDMI_INCOMPLETE_SCOPE_TITLE, bold=True))
        blocks.append(_docx_paragraph(str(data["scope_summary"]), bold=True))
    rows = _udmi_client_metric_rows(run, data)
    blocks.extend(
        [
            _docx_paragraph("Client Metrics", bold=True, keep_with_next=True),
            _docx_table(
                ("Metric", "Value"),
                rows,
                widths=_docx_content_widths(("Metric", "Value"), rows),
                center_cells=True,
            ),
            _docx_paragraph(
                "This client product contains metrics and acceptance status only. "
                "Use the technical product for asset and finding detail."
            ),
        ]
    )
    return _assemble_docx(run, blocks, document_title=title, landscape=False)


def _build_udmi_docx_report(run: object, data: dict[str, object]) -> bytes:
    if _udmi_report_variant(run) == "client":
        return _build_udmi_client_docx_report(run, data)
    title = _report_title(run)
    blocks: list[str] = [_docx_paragraph(title, bold=True)]
    blocks.extend(
        _docx_paragraph(f"{label}: {value}")
        for label, value in _report_rows(run, udmi_data=data)
    )
    if data["scope_complete"] is not True:
        blocks.append(_docx_paragraph(_UDMI_INCOMPLETE_SCOPE_TITLE, bold=True))
        blocks.append(_docx_paragraph(str(data["scope_summary"]), bold=True))
    definition_rows = [
        {
            "Metric": str(definition["metric"]),
            "Definition": str(definition["definition"]),
        }
        for definition in data["metric_definitions"]
    ]
    definition_columns = ("Metric", "Definition")
    blocks.append(_docx_paragraph(_UDMI_DEFINITIONS_TITLE, bold=True, keep_with_next=True))
    blocks.append(
        _docx_table(
            definition_columns,
            definition_rows,
            widths=_docx_content_widths(definition_columns, definition_rows),
            center_cells=True,
        )
    )
    blocks.extend(
        [
            _docx_page_break(),
            _docx_paragraph(_UDMI_EXECUTIVE_TITLE, bold=True, keep_with_next=True),
        ]
    )

    for section_title, labels, key in (
        ("Asset Level Metrics", ASSET_METRIC_LABELS, "asset_metrics"),
        ("Payload Level Metrics", PAYLOAD_METRIC_LABELS, "payload_metrics"),
        ("Fault Metrics", FAULT_METRIC_LABELS, "fault_metrics"),
    ):
        blocks.append(
            _docx_paragraph(
                section_title,
                bold=True,
                page_break_before=section_title == "Fault Metrics",
                keep_with_next=True,
            )
        )
        rows = _udmi_metric_rows(data[key], labels)
        blocks.append(
            _docx_table(
                ("Metric", "Value"),
                rows,
                widths=_docx_content_widths(("Metric", "Value"), rows),
                center_cells=True,
            )
        )
    blocks.append(_docx_paragraph("Supporting Metrics", bold=True, keep_with_next=True))
    blocks.append(
        _docx_table(
            ("Metric", "Value"),
            _udmi_supporting_metric_rows(run, data),
            widths=_docx_content_widths(
                ("Metric", "Value"),
                _udmi_supporting_metric_rows(run, data),
            ),
            center_cells=True,
        )
    )
    notes = [str(note) for note in data["notes"]]
    if notes:
        blocks.append(_docx_paragraph(f"Notes: {' '.join(notes)}"))

    wrong_topic_rows = _udmi_wrong_topic_rows(data)
    if wrong_topic_rows:
        blocks.extend(
            [
                _docx_paragraph(
                    _UDMI_WRONG_TOPIC_TITLE,
                    bold=True,
                    page_break_before=True,
                    keep_with_next=True,
                ),
                _docx_table(
                    _UDMI_WRONG_TOPIC_COLUMNS,
                    wrong_topic_rows,
                    widths=_docx_content_widths(
                        _UDMI_WRONG_TOPIC_COLUMNS,
                        wrong_topic_rows,
                    ),
                    center_cells=True,
                ),
            ]
        )

    system_assets, system_payloads, system_faults = _udmi_system_tables(data)
    blocks.extend(
        [
            _docx_page_break(),
            _docx_paragraph(_UDMI_SYSTEM_TITLE, bold=True, keep_with_next=True),
        ]
    )
    for section_title, columns, rows in (
        ("Asset Metrics by System", _UDMI_SYSTEM_ASSET_COLUMNS, system_assets),
        ("Payload and Issue Metrics by System", _UDMI_SYSTEM_PAYLOAD_COLUMNS, system_payloads),
        ("Fault Metrics by System", _UDMI_SYSTEM_FAULT_COLUMNS, system_faults),
    ):
        blocks.append(_docx_paragraph(section_title, bold=True, keep_with_next=True))
        blocks.append(
            _docx_table(
                columns,
                rows,
                widths=_docx_content_widths(columns, rows),
                center_cells=True,
            )
        )

    asset_rows = _udmi_asset_rows(data)
    blocks.extend(
        [
            _docx_page_break(),
            _docx_paragraph(_UDMI_ASSET_TITLE, bold=True, keep_with_next=True),
        ]
    )
    if asset_rows:
        blocks.append(
            _docx_table(
                _UDMI_ASSET_COLUMNS,
                asset_rows,
                widths=_docx_content_widths(_UDMI_ASSET_COLUMNS, asset_rows),
                center_cells=True,
            )
        )
    else:
        blocks.append(_docx_paragraph("No asset results were retained for the selected runs."))

    matrix_rows = _udmi_fault_matrix_rows(data)
    blocks.extend(
        [
            _docx_page_break(),
            _docx_paragraph(_UDMI_FAULT_MATRIX_TITLE, bold=True, keep_with_next=True),
        ]
    )
    if matrix_rows:
        blocks.append(
            _docx_table(
                _UDMI_FAULT_MATRIX_COLUMNS,
                matrix_rows,
                widths=_docx_content_widths(_UDMI_FAULT_MATRIX_COLUMNS, matrix_rows),
                center_cells=True,
            )
        )
    else:
        blocks.append(_docx_paragraph("No faults were retained for the selected runs."))

    detail_rows = _udmi_fault_detail_rows(data)
    blocks.extend(
        [
            _docx_page_break(),
            _docx_paragraph(_UDMI_FAULT_DETAIL_TITLE, bold=True, keep_with_next=True),
        ]
    )
    if detail_rows:
        blocks.append(
            _docx_table(
                _UDMI_FAULT_DETAIL_COLUMNS,
                detail_rows,
                widths=_docx_content_widths(_UDMI_FAULT_DETAIL_COLUMNS, detail_rows),
                center_cells=True,
            )
        )
    else:
        blocks.append(_docx_paragraph("No faults were retained for the selected runs."))

    return _assemble_docx(run, blocks, document_title=title, landscape=True)


def _build_docx_report(run: object) -> bytes:
    udmi = _udmi_report_data(run)
    if udmi is not None:
        return _build_udmi_docx_report(run, udmi)

    blocks: list[str] = [_docx_paragraph(_report_title(run), bold=True)]
    blocks.extend(_docx_paragraph(f"{label}: {value}") for label, value in _report_rows(run))

    validation = _validation_summary(run)
    if validation is not None:
        blocks.append(_docx_paragraph(_SUMMARY_SECTION_TITLE, bold=True))
        blocks.append(_docx_table(_VALIDATION_SUMMARY_COLUMNS, validation["rows"]))
        # The paragraph after each table doubles as the Word-required trailing
        # paragraph (a body may not end <w:tbl><w:sectPr/>).
        blocks.append(_docx_paragraph(validation["overall_text"]))

    inventory = _discovery_inventory(run)
    if inventory is not None:
        for section in inventory:
            blocks.append(_docx_paragraph(section["title"], bold=True))
            if section["note"]:
                blocks.append(_docx_paragraph(section["note"]))
            if section["rows"]:
                blocks.append(_docx_table(section["columns"], section["rows"]))
                # Trailing paragraph so a body never ends on <w:tbl> (Word rule).
                blocks.append(_docx_paragraph(""))

    blocks.append(_docx_paragraph(_FAILURE_SECTION_TITLE, bold=True))
    findings = _source_run_findings(run)
    if findings:
        blocks.append(_docx_table(_FINDING_COLUMNS, findings))
        blocks.append(_docx_paragraph(""))
    else:
        blocks.append(_docx_paragraph(_NO_FINDINGS_NOTE))

    if validation is not None:
        blocks.append(_docx_paragraph(_SILENT_SECTION_TITLE, bold=True))
        blocks.append(_docx_paragraph(_SILENT_NOTE))
        if validation["silent_rows"]:
            blocks.append(_docx_table(_SILENT_COLUMNS, validation["silent_rows"]))
            blocks.append(_docx_paragraph(""))
        else:
            blocks.append(_docx_paragraph(_NO_SILENT_NOTE))

    return _assemble_docx(
        run,
        blocks,
        document_title=_report_title(run),
        landscape=False,
    )


def _build_udmi_client_zip_report(run: object, data: dict[str, object]) -> bytes:
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr(
            "provenance.json",
            json.dumps(_report_provenance(run), indent=2, sort_keys=True),
        )
        archive.writestr(
            "summary.json",
            json.dumps(dict(_report_rows(run, udmi_data=data)), indent=2),
        )
        archive.writestr(
            "client_metrics.json",
            json.dumps(
                {
                    "schema_version": data["schema_version"],
                    "report_product": "client_metrics",
                    "report_title": _report_title(run),
                    "evidence_set_id": data.get("evidence_set_id"),
                    "scope_status": data.get("scope_status"),
                    "scope_summary": data.get("scope_summary"),
                    "acceptance_eligible": data.get("acceptance_eligible") is True,
                    "metrics": _udmi_client_metric_rows(run, data),
                    "floor_reporting": data.get("floor_reporting"),
                    "evidence_provenance": data.get("evidence_provenance"),
                    "notes": data.get("notes", []),
                },
                indent=2,
            ),
        )
    return buffer.getvalue()


def _build_zip_report(run: object) -> bytes:
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        udmi = _udmi_report_data(run)
        if udmi is not None and _udmi_report_variant(run) == "client":
            return _build_udmi_client_zip_report(run, udmi)
        archive.writestr(
            "provenance.json",
            json.dumps(_report_provenance(run), indent=2, sort_keys=True),
        )
        archive.writestr(
            "summary.json",
            json.dumps(dict(_report_rows(run, udmi_data=udmi)), indent=2),
        )
        # UDMI findings come from the exact retained report model, including an
        # optional selected-payload scope. Other report types retain the generic
        # source-run projection.
        findings = (
            _udmi_audit_finding_rows(udmi)
            if udmi is not None
            else _source_run_findings(run)
        )
        archive.writestr("findings.json", json.dumps(findings, indent=2))
        asset_topic_discovery = _asset_topic_discovery_export(run)
        if asset_topic_discovery is not None:
            archive.writestr(
                "asset_topic_discovery.json",
                json.dumps(asset_topic_discovery, indent=2),
            )
        # Parity with the document formats: the validation sections ship as
        # their own JSON members when the source runs include validation runs.
        if udmi is not None:
            payloads_correct, payloads_incorrect = _payload_correctness(udmi["payload_metrics"])
            archive.writestr(
                "validation_summary.json",
                json.dumps(
                    {
                        "schema_version": udmi["schema_version"],
                        "report_title": _report_title(run),
                        "report_job_status": str(run.status),
                        "scope_complete": udmi["scope_complete"],
                        "acceptance_eligible": udmi["acceptance_eligible"],
                        "scope_status": udmi["scope_status"],
                        "scope_summary": udmi["scope_summary"],
                        "incomplete_source_runs": udmi["incomplete_source_runs"],
                        "last_validation_run_at": udmi["last_validation_run_at"],
                        "report_generated_at": _generated_at(run),
                        "source_runs": udmi["source_runs"],
                        "asset_metrics": udmi["asset_metrics"],
                        "payload_metrics": udmi["payload_metrics"],
                        "fault_metrics": udmi["fault_metrics"],
                        "issue_metrics": udmi["issue_metrics"],
                        "overall_compliance": _completion(udmi["asset_metrics"]),
                        "payloads_correct": payloads_correct,
                        "payloads_incorrect": payloads_incorrect,
                        "system_metrics": udmi["system_metrics"],
                        "unexpected_devices": udmi["unexpected_devices"],
                        "wrong_topic_assets": udmi.get("wrong_topic_assets", []),
                        "unexpected_devices_measured": udmi["unexpected_devices_measured"],
                        "unexpected_devices_measurement_scope": udmi[
                            "unexpected_devices_measurement_scope"
                        ],
                        "report_scope": udmi["report_scope"],
                        "evidence_set_id": udmi.get("evidence_set_id"),
                        "evidence_provenance": udmi.get("evidence_provenance"),
                        "filter_provenance": udmi["filter_provenance"],
                        "notes": udmi["notes"],
                    },
                    indent=2,
                ),
            )
            archive.writestr(
                "asset_validation_schedule.json",
                json.dumps(
                    {
                        "schema_version": udmi["schema_version"],
                        "rows": udmi["asset_results"],
                    },
                    indent=2,
                ),
            )
            wrong_topic_assets = udmi.get("wrong_topic_assets")
            if isinstance(wrong_topic_assets, list) and wrong_topic_assets:
                archive.writestr(
                    "wrong_topic_assets.json",
                    json.dumps(
                        {
                            "schema_version": udmi["schema_version"],
                            "rows": wrong_topic_assets,
                        },
                        indent=2,
                    ),
                )
            register_columns, register_rows = _udmi_annotated_register(udmi)
            if register_columns and register_rows:
                archive.writestr(
                    "annotated_input_register.json",
                    json.dumps(
                        {
                            "schema_version": udmi["schema_version"],
                            "columns": list(register_columns),
                            "rows": register_rows,
                        },
                        indent=2,
                    ),
                )
            archive.writestr(
                "fault_matrix.json",
                json.dumps(
                    {
                        "schema_version": udmi["schema_version"],
                        "rows": udmi["fault_matrix"],
                    },
                    indent=2,
                ),
            )
            archive.writestr(
                "fault_details.json",
                json.dumps(
                    {
                        "schema_version": udmi["schema_version"],
                        "rows": _udmi_fault_json_rows(udmi),
                    },
                    indent=2,
                ),
            )
            archive.writestr(
                "metric_definitions.json",
                json.dumps({"rows": udmi["metric_definitions"]}, indent=2),
            )

        validation = None if udmi is not None else _validation_summary(run)
        if validation is not None:
            archive.writestr(
                "validation_summary.json",
                json.dumps(
                    {
                        "columns": list(_VALIDATION_SUMMARY_COLUMNS),
                        "rows": validation["rows"],
                        "overall": validation["overall"],
                    },
                    indent=2,
                ),
            )
            archive.writestr(
                "silent_systems.json",
                json.dumps({"note": _SILENT_NOTE, "rows": validation["silent_rows"]}, indent=2),
            )
        # Discovery inventory (parity with the document formats): its own member
        # when the source runs include discovery runs, absent otherwise. Key
        # order is fixed by construction so the member bytes stay reproducible.
        inventory = None if udmi is not None else _discovery_inventory(run)
        if inventory is not None:
            archive.writestr(
                "discovery_inventory.json",
                json.dumps(
                    {
                        "sections": [
                            {
                                "title": section["title"],
                                "columns": list(section["columns"]),
                                "rows": section["rows"],
                                **({"note": section["note"]} if section["note"] else {}),
                            }
                            for section in inventory
                        ]
                    },
                    indent=2,
                ),
            )
        raw_evidence = _raw_evidence_export(run)
        raw_records = raw_evidence.get("records")
        raw_records = raw_records if isinstance(raw_records, list) else []
        portable_raw_records = [record for record in raw_records if isinstance(record, dict)]
        payload_files, payload_manifest = _raw_evidence_payload_files(portable_raw_records)
        raw_manifest = dict(raw_evidence)
        raw_manifest.pop("records", None)
        raw_manifest["payload_file_manifest"] = "raw_evidence/payload_manifest.json"
        archive.writestr(
            "raw_evidence/manifest.json",
                json.dumps(raw_manifest, indent=2, sort_keys=True),
            )
        for filename, content in payload_files:
            archive.writestr(filename, content)
        archive.writestr(
            "raw_evidence/payload_manifest.json",
            json.dumps(payload_manifest, indent=2, sort_keys=True),
        )
        archive.writestr(
            "raw_evidence/records.jsonl",
            "".join(
                json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n"
                for record in portable_raw_records
            ),
        )
        archive.writestr(
            "raw_evidence/finding_index.json",
            json.dumps(
                _raw_finding_index(
                    findings if isinstance(findings, list) else [],
                    portable_raw_records,
                ),
                indent=2,
                sort_keys=True,
            ),
        )
    return buffer.getvalue()


# Relative column weights for the fixed-width PDF tables (long text columns get
# more room; every cell truncates with an ellipsis rather than overflowing).
# The dense tables render at 9pt and the summary weights are sized so the
# column headers and the compliance cell — including its honesty-critical
# "(liveness)" marker — always fit; run ids may truncate (they appear in full
# in the header rows and the Silent systems table).
_PDF_SUMMARY_WEIGHTS = (71, 68, 49, 82, 52, 30, 74, 69)
_PDF_SILENT_WEIGHTS = (1.0, 1.6)
_PDF_DENSE_TABLE_SIZE = 9.0

# PDF Failure detail renders in two parts: a slim identity table (short,
# id-like fields only — weighted so every header and a full run id fit at 9pt)
# followed per finding by word-wrapped paragraphs for the long free-text
# fields, which a 10-column table truncated into uselessness.
_PDF_FINDING_IDENTITY_COLUMNS = ("Source Run", "Issue ID", "Asset", "Severity", "Type", "Point")
_PDF_FINDING_IDENTITY_WEIGHTS = (136, 62, 72, 47, 86, 92)
_PDF_FINDING_DETAIL_FIELDS = ("Expected", "Observed", "Suggested Action", "Description")

# Relative column weights for the fixed-width inventory PDF tables. Long text
# columns (Counts, Topic, Point Name) get more room; every cell still truncates
# with an ellipsis. Truncation is acceptable ONLY because the xlsx and zip
# artifacts carry the full untruncated values — same honesty trade-off as
# _PDF_SUMMARY_WEIGHTS; do not "fix" truncation by dropping columns.
_PDF_INVENTORY_SUMMARY_WEIGHTS = (80, 60, 45, 220)
_PDF_IP_WEIGHTS = (92, 74, 70, 58, 62, 46, 46, 46)
_PDF_BACNET_DEVICE_WEIGHTS = (92, 42, 62, 74, 52, 52, 74, 30)
_PDF_BACNET_POINT_WEIGHTS = (86, 60, 56, 104, 78, 40)
_PDF_BACNET_SILENT_WEIGHTS = (86, 96, 44, 72, 58)
_PDF_MQTT_WEIGHTS = (72, 214, 46, 80)
_PDF_INVENTORY_WEIGHTS = {
    _INVENTORY_SUMMARY_TITLE: _PDF_INVENTORY_SUMMARY_WEIGHTS,
    _INVENTORY_IP_TITLE: _PDF_IP_WEIGHTS,
    _INVENTORY_BACNET_DEVICES_TITLE: _PDF_BACNET_DEVICE_WEIGHTS,
    _INVENTORY_BACNET_POINTS_TITLE: _PDF_BACNET_POINT_WEIGHTS,
    _INVENTORY_BACNET_SILENT_TITLE: _PDF_BACNET_SILENT_WEIGHTS,
    _INVENTORY_MQTT_TITLE: _PDF_MQTT_WEIGHTS,
}

_PDF_UDMI_METRIC_WEIGHTS = (3.3, 1.0, 3.3, 1.0)
_PDF_UDMI_DETAIL_COLUMNS = _UDMI_FAULT_DETAIL_COLUMNS


def _pdf_metric_pairs(rows: list[dict[str, str]]) -> list[list[str]]:
    pairs: list[list[str]] = []
    for index in range(0, len(rows), 2):
        left = rows[index]
        right = rows[index + 1] if index + 1 < len(rows) else {"Metric": "", "Value": ""}
        pairs.append([left["Metric"], left["Value"], right["Metric"], right["Value"]])
    return pairs


def _build_udmi_client_pdf_report(run: object, data: dict[str, object]) -> bytes:
    title = _report_title(run)
    document = PdfDocument(
        header_left=_BRAND_NAME,
        header_right=title,
        footer_left=_BRAND_NAME,
        footer_right=str(run.run_id),
    )
    document.add_heading(title, level=1)
    for label, value in _report_rows(run, udmi_data=data):
        document.add_paragraph(f"{label}: {value}")
    if data["scope_complete"] is not True:
        document.add_heading(_UDMI_INCOMPLETE_SCOPE_TITLE)
        document.add_paragraph(str(data["scope_summary"]), bold=True)
    rows = _udmi_client_metric_rows(run, data)
    document.add_heading("Client Metrics", level=1)
    document.add_table(
        ("Metric", "Value"),
        [[row["Metric"], row["Value"]] for row in rows],
        widths=(3.4, 2.6),
        size=9,
        wrap_cells=True,
        center_cells=True,
        draw_grid=True,
    )
    return document.render()


def _build_udmi_pdf_report(run: object, data: dict[str, object]) -> bytes:
    if _udmi_report_variant(run) == "client":
        return _build_udmi_client_pdf_report(run, data)
    title = _report_title(run)
    document = PdfDocument(
        header_left=_BRAND_NAME,
        header_right=title,
        footer_left=_BRAND_NAME,
        footer_right=str(run.run_id),
        landscape=True,
    )
    document.add_heading(title, level=1)
    for label, value in _report_rows(run, udmi_data=data):
        document.add_paragraph(f"{label}: {value}")
    if data["scope_complete"] is not True:
        document.add_heading(_UDMI_INCOMPLETE_SCOPE_TITLE)
        document.add_paragraph(str(data["scope_summary"]), bold=True)
    definition_rows = [
        {
            "Metric": str(definition["metric"]),
            "Definition": str(definition["definition"]),
        }
        for definition in data["metric_definitions"]
    ]
    definition_columns = ("Metric", "Definition")
    document.add_heading(_UDMI_DEFINITIONS_TITLE, level=1)
    document.add_table(
        definition_columns,
        [[row[column] for column in definition_columns] for row in definition_rows],
        widths=_content_weights(definition_columns, definition_rows),
        size=8,
        wrap_cells=True,
        center_cells=True,
        draw_grid=True,
    )
    document.add_page_break()
    document.add_heading(_UDMI_EXECUTIVE_TITLE)

    for section_title, labels, key in (
        ("Asset Level Metrics", ASSET_METRIC_LABELS, "asset_metrics"),
        ("Payload Level Metrics", PAYLOAD_METRIC_LABELS, "payload_metrics"),
        ("Fault Metrics", FAULT_METRIC_LABELS, "fault_metrics"),
    ):
        document.add_heading(section_title)
        rows = _udmi_metric_rows(data[key], labels)
        document.add_table(
            ("Metric", "Value", "Metric", "Value"),
            _pdf_metric_pairs(rows),
            widths=_PDF_UDMI_METRIC_WEIGHTS,
            size=9,
            wrap_cells=True,
            center_cells=True,
            draw_grid=True,
        )
    supporting = _udmi_supporting_metric_rows(run, data)
    document.add_heading("Supporting Metrics")
    document.add_table(
        ("Metric", "Value", "Metric", "Value"),
        _pdf_metric_pairs(supporting),
        widths=_PDF_UDMI_METRIC_WEIGHTS,
        size=9,
        wrap_cells=True,
        center_cells=True,
        draw_grid=True,
    )
    for note in data["notes"]:
        document.add_paragraph(f"Note: {note}")

    wrong_topic_rows = _udmi_wrong_topic_rows(data)
    if wrong_topic_rows:
        document.add_page_break()
        document.add_heading(_UDMI_WRONG_TOPIC_TITLE, level=1)
        document.add_table(
            _UDMI_WRONG_TOPIC_COLUMNS,
            [
                [row[column] for column in _UDMI_WRONG_TOPIC_COLUMNS]
                for row in wrong_topic_rows
            ],
            widths=_content_weights(_UDMI_WRONG_TOPIC_COLUMNS, wrong_topic_rows),
            size=7.5,
            wrap_cells=True,
            center_cells=True,
            draw_grid=True,
        )

    system_assets, system_payloads, system_faults = _udmi_system_tables(data)
    document.add_page_break()
    document.add_heading(_UDMI_SYSTEM_TITLE, level=1)
    for section_title, columns, rows in (
        (
            "Asset Metrics by System",
            _UDMI_SYSTEM_ASSET_COLUMNS,
            system_assets,
        ),
        (
            "Payload and Issue Metrics by System",
            _UDMI_SYSTEM_PAYLOAD_COLUMNS,
            system_payloads,
        ),
        (
            "Fault Metrics by System",
            _UDMI_SYSTEM_FAULT_COLUMNS,
            system_faults,
        ),
    ):
        document.add_heading(section_title)
        if rows:
            document.add_table(
                columns,
                [[row[column] for column in columns] for row in rows],
                widths=_content_weights(columns, rows),
                size=8,
                wrap_cells=True,
                center_cells=True,
                draw_grid=True,
            )
        else:
            document.add_paragraph("No per-system metrics were retained for the selected runs.")

    asset_rows = _udmi_asset_rows(data)
    document.add_page_break()
    document.add_heading(_UDMI_ASSET_TITLE, level=1)
    if asset_rows:
        document.add_table(
            _UDMI_ASSET_COLUMNS,
            [[row[column] for column in _UDMI_ASSET_COLUMNS] for row in asset_rows],
            widths=_content_weights(_UDMI_ASSET_COLUMNS, asset_rows),
            size=8,
            wrap_cells=True,
            center_cells=True,
            draw_grid=True,
        )
    else:
        document.add_paragraph("No asset results were retained for the selected runs.")

    matrix_rows = _udmi_fault_matrix_rows(data)
    document.add_page_break()
    document.add_heading(_UDMI_FAULT_MATRIX_TITLE, level=1)
    if matrix_rows:
        document.add_table(
            _UDMI_FAULT_MATRIX_COLUMNS,
            [[row[column] for column in _UDMI_FAULT_MATRIX_COLUMNS] for row in matrix_rows],
            widths=_content_weights(_UDMI_FAULT_MATRIX_COLUMNS, matrix_rows),
            size=7.5,
            wrap_cells=True,
            center_cells=True,
            draw_grid=True,
        )
    else:
        document.add_paragraph("No faults were retained for the selected runs.")

    detail_rows = _udmi_fault_detail_rows(data)
    document.add_page_break()
    document.add_heading(_UDMI_FAULT_DETAIL_TITLE, level=1)
    if detail_rows:
        document.add_table(
            _PDF_UDMI_DETAIL_COLUMNS,
            [[row[column] for column in _PDF_UDMI_DETAIL_COLUMNS] for row in detail_rows],
            widths=_content_weights(_PDF_UDMI_DETAIL_COLUMNS, detail_rows),
            size=7,
            wrap_cells=True,
            center_cells=True,
            draw_grid=True,
        )
    else:
        document.add_paragraph("No faults were retained for the selected runs.")

    return document.render()


def _build_pdf_report(run: object) -> bytes:
    udmi = _udmi_report_data(run)
    if udmi is not None:
        return _build_udmi_pdf_report(run, udmi)

    document = PdfDocument(
        header_left=_BRAND_NAME,
        header_right=_report_title(run),
        footer_left=_BRAND_NAME,
        footer_right=str(run.run_id),
    )
    document.add_heading(_report_title(run), level=1)
    for label, value in _report_rows(run):
        document.add_paragraph(f"{label}: {value}")

    validation = _validation_summary(run)
    if validation is not None:
        document.add_heading(_SUMMARY_SECTION_TITLE)
        document.add_table(
            _VALIDATION_SUMMARY_COLUMNS,
            [[row[column] for column in _VALIDATION_SUMMARY_COLUMNS] for row in validation["rows"]],
            widths=_PDF_SUMMARY_WEIGHTS,
            size=_PDF_DENSE_TABLE_SIZE,
        )
        document.add_paragraph(validation["overall_text"])

    inventory = _discovery_inventory(run)
    if inventory is not None:
        for section in inventory:
            document.add_heading(section["title"])
            if section["note"]:
                document.add_paragraph(section["note"])
            rows = section["rows"]
            if rows:
                columns = section["columns"]
                document.add_table(
                    columns,
                    [[row[column] for column in columns] for row in rows],
                    widths=_PDF_INVENTORY_WEIGHTS.get(section["title"]),
                    size=_PDF_DENSE_TABLE_SIZE,
                )

    document.add_heading(_FAILURE_SECTION_TITLE)
    findings = _source_run_findings(run)
    if findings:
        document.add_table(
            _PDF_FINDING_IDENTITY_COLUMNS,
            [[finding[column] for column in _PDF_FINDING_IDENTITY_COLUMNS] for finding in findings],
            widths=_PDF_FINDING_IDENTITY_WEIGHTS,
            size=_PDF_DENSE_TABLE_SIZE,
        )
        # Long free-text fields render as full-width wrapped paragraphs (the
        # existing deterministic wrap), never truncated table cells.
        for finding in findings:
            identity = " — ".join(
                part
                for part in (finding["Issue ID"], finding["Asset"], finding["Point"])
                if part
            )
            document.add_paragraph(identity or finding["Source Run"], bold=True)
            for field in _PDF_FINDING_DETAIL_FIELDS:
                if finding[field]:
                    document.add_paragraph(f"{field}: {finding[field]}")
    else:
        document.add_paragraph(_NO_FINDINGS_NOTE)

    if validation is not None:
        document.add_heading(_SILENT_SECTION_TITLE)
        document.add_paragraph(_SILENT_NOTE)
        if validation["silent_rows"]:
            document.add_table(
                _SILENT_COLUMNS,
                [[row[column] for column in _SILENT_COLUMNS] for row in validation["silent_rows"]],
                widths=_PDF_SILENT_WEIGHTS,
            )
        else:
            document.add_paragraph(_NO_SILENT_NOTE)
    return document.render()
