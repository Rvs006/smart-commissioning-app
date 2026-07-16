import codecs
import csv
import io
import ipaddress
import re
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from secrets import token_hex

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from smart_commissioning_core.db.repositories import ImportRepository
from smart_commissioning_core.engines.comparison_common import parse_tolerance
from sqlalchemy.engine import Engine

from app.core.db import get_engine
from app.core.runtime import IMPORT_FILES_ROOT, ensure_runtime_directories
from app.schemas.imports import (
    ImportBatchSummary,
    ImportErrorRecord,
    ImportErrorReport,
    ImportProfileSummary,
    ImportType,
)

ALLOWED_UNITS = {
    "",
    "degrees-celsius",
    "percent",
    "percent-relative-humidity",
    "parts-per-million",
    "minutes",
    "seconds",
    "no-units",
    "pa",
    "cfm",
    "kwh",
    "kw",
}


@dataclass(frozen=True)
class ImportProfile:
    import_type: ImportType
    description: str
    required_columns: tuple[str, ...]
    duplicate_key_fields: tuple[str, ...]
    # Extra row checks run after the required-column emptiness check.
    extra_checks: tuple[Callable[[dict[str, str], int], list[ImportErrorRecord]], ...] = ()
    # Recognised-but-not-required columns. Included in the downloadable template
    # and canonicalised into accepted rows when present, but an empty/missing
    # value never rejects a row (e.g. "Expected hostname", which most sites do
    # not populate).
    optional_columns: tuple[str, ...] = ()
    # Informational row checks: their records surface as ImportBatchSummary
    # warnings and never affect acceptance (unlike extra_checks).
    warning_checks: tuple[Callable[[dict[str, str], int], list[ImportErrorRecord]], ...] = ()

    def validate_row(self, row: dict[str, str], row_number: int) -> list[ImportErrorRecord]:
        errors = _base_row_validation(self.required_columns, row, row_number)
        for check in self.extra_checks:
            errors.extend(check(row, row_number))
        return errors

    def collect_warnings(self, row: dict[str, str], row_number: int) -> list[ImportErrorRecord]:
        warnings: list[ImportErrorRecord] = []
        for check in self.warning_checks:
            warnings.extend(check(row, row_number))
        return warnings

    def as_summary(self) -> ImportProfileSummary:
        return ImportProfileSummary(
            import_type=self.import_type,
            description=self.description,
            required_columns=list(self.required_columns),
            optional_columns=list(self.optional_columns),
            duplicate_key_fields=list(self.duplicate_key_fields),
        )

    @property
    def template_columns(self) -> tuple[str, ...]:
        """Required columns followed by optional ones, for template + mapping."""
        return (*self.required_columns, *self.optional_columns)


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().casefold())


def _validate_ip(row: dict[str, str], row_number: int, field: str) -> list[ImportErrorRecord]:
    value = row.get(field, "").strip()
    if not value:
        return []
    try:
        ipaddress.ip_address(value)
    except ValueError:
        return [ImportErrorRecord(row_number=row_number, field=field, code="invalid_ip", message=f"{field} is not a valid IP address.")]
    return []


def _validate_numeric(row: dict[str, str], row_number: int, field: str) -> list[ImportErrorRecord]:
    """Non-negative integers only — do not loosen to accept decimals.

    This guards identity/count fields (BACnet device + object instance,
    reporting interval) where a decimal is meaningless. Decimal-valued fields
    have their own validators (see :func:`_validate_tolerance`).
    """
    value = row.get(field, "").strip()
    if not value:
        return []
    if not value.isdigit():
        return [ImportErrorRecord(row_number=row_number, field=field, code="invalid_numeric", message=f"{field} must be numeric.")]
    return []


def _validate_tolerance(row: dict[str, str], row_number: int, field: str) -> list[ImportErrorRecord]:
    """Accept exactly the tolerance forms the comparison engines understand.

    A tolerance is a decimal ("0.5") or a percentage ("5%"), never an integer.
    Delegating to ``parse_tolerance`` rather than re-stating its grammar keeps
    the import gate from drifting out of step with the engines that parse these
    cells later (comparison / point_validation, via build_tolerance_index).
    """
    value = row.get(field, "").strip()
    if not value:
        return []
    if parse_tolerance(value) is None:
        return [
            ImportErrorRecord(
                row_number=row_number,
                field=field,
                code="invalid_tolerance",
                message=f"{field} must be a number (e.g. 0.5) or a percentage (e.g. 5%).",
            )
        ]
    return []


def _validate_units(row: dict[str, str], row_number: int, field: str) -> list[ImportErrorRecord]:
    value = row.get(field, "").strip().casefold()
    if value in ALLOWED_UNITS:
        return []
    return [ImportErrorRecord(row_number=row_number, field=field, code="invalid_unit", message=f"{field} is not a recognized unit.")]


def _validate_topic(row: dict[str, str], row_number: int, field: str) -> list[ImportErrorRecord]:
    value = row.get(field, "").strip()
    if not value:
        return []
    # A topic field may hold one topic, a "prefix/#" wildcard (covers an asset's
    # metadata/state/events topics in one entry), or a comma-separated list.
    # Wildcards must occupy a complete MQTT level; a bare wildcard is too broad
    # for an asset register and cannot be mapped to state/metadata/pointset.
    for topic in (part.strip() for part in value.split(",")):
        if not topic:
            continue
        levels = topic.split("/")
        wildcard_invalid = any(
            ("#" in level and (level != "#" or index != len(levels) - 1))
            or ("+" in level and level != "+")
            for index, level in enumerate(levels)
        )
        if " " in topic or "/" not in topic or topic in ("#", "+") or wildcard_invalid:
            return [
                ImportErrorRecord(
                    row_number=row_number,
                    field=field,
                    code="invalid_topic",
                    message=f"{field} must be a valid MQTT topic path (use '/', optional '#'/'+' wildcards).",
                )
            ]
    return []


# A whole number, optionally written with an integral decimal part: Excel
# routinely saves a numeric column as "60.0", which asserts the same cadence as
# "60". Deliberately narrower than float(): "1e2", "+60", "inf" and "60.5" stay
# rejected. Used ONLY by _validate_positive_numeric (mqtt_register's Expected
# reporting interval) — _validate_numeric above stays strict because it guards
# genuinely-integer BACnet fields shared by other profiles.
_WHOLE_NUMBER_RE = re.compile(r"\d+(\.0+)?")


def _validate_positive_numeric(row: dict[str, str], row_number: int, field: str) -> list[ImportErrorRecord]:
    value = row.get(field, "").strip()
    if not value:
        # Blank is the required-column check's business, not ours.
        return []
    if not _WHOLE_NUMBER_RE.fullmatch(value):
        return [
            ImportErrorRecord(
                row_number=row_number,
                field=field,
                code="invalid_numeric",
                message=f"{field} must be a whole number of seconds (e.g. 60 — Excel decimals like 60.0 are accepted, 60.5 is not).",
            )
        ]
    if float(value) <= 0:
        return [ImportErrorRecord(row_number=row_number, field=field, code="invalid_number", message=f"{field} must be greater than zero.")]
    return []


# Recognised trailing payload suffixes of an asset topic; stripping one yields
# the asset's topic root (its device prefix).
_ASSET_TOPIC_SUFFIXES = ("/#", "/state", "/metadata", "/event/pointset", "/events/pointset")


def _topic_roots(value: str) -> frozenset[str]:
    """Topic roots for a comma-separated topic cell.

    Strips exactly ONE trailing payload suffix per topic — never chained, so a
    root that itself ends in a payload-like level (e.g. ``a/metadata/state`` ->
    ``a/metadata``) keeps that level.
    """
    roots: set[str] = set()
    for part in value.split(","):
        topic = part.strip()
        if not topic:
            continue
        for suffix in _ASSET_TOPIC_SUFFIXES:
            if topic.endswith(suffix):
                topic = topic.removesuffix(suffix)
                break
        roots.add(topic)
    return frozenset(roots)


def _validate_mqtt_asset_topic(
    row: dict[str, str], row_number: int, field: str
) -> list[ImportErrorRecord]:
    value = row.get(field, "").strip()
    if not value or _validate_topic(row, row_number, field):
        return []
    topics = [topic.strip() for topic in value.split(",") if topic.strip()]
    if not any("+" in topic.split("/") for topic in topics) and all(
        topic.endswith(_ASSET_TOPIC_SUFFIXES) for topic in topics
    ):
        if not row.get("Payload type", "").strip():
            if len(_topic_roots(value)) != 1:
                return [ImportErrorRecord(row_number=row_number, field=field, code="invalid_topic", message=f"{field} blank payload type must reference one asset root.")]
        return []
    return [
        ImportErrorRecord(
            row_number=row_number,
            field=field,
            code="invalid_topic",
            # Suffixes are interpolated from the tuple the gate above tests, so
            # the message can never drift from the rule it explains.
            message=(
                f"{field} must end with one of {', '.join(_ASSET_TOPIC_SUFFIXES)} — e.g. "
                "'site/b1/fcu-04/#' covers that asset's state, metadata and events/pointset "
                "topics; '+' wildcards are not allowed."
            ),
        )
    ]


def _validate_payload_type(row: dict[str, str], row_number: int) -> list[ImportErrorRecord]:
    value = row.get("Payload type", "").strip().casefold()
    if value in {"", "state", "metadata", "pointset"}:
        return []
    return [
        ImportErrorRecord(
            row_number=row_number,
            field="Payload type",
            code="invalid_payload_type",
            message="Payload type must be blank, state, metadata, or pointset.",
        )
    ]


def _validate_mqtt_point_unit_pairs(row: dict[str, str], row_number: int) -> list[ImportErrorRecord]:
    points_value = row.get("Expected points", "").strip()
    units_value = row.get("Expected units", "").strip()
    if not points_value or not units_value:
        return []  # Required-field validation already reports blanks.
    point_count = len([value for value in points_value.split(",") if value.strip()])
    unit_slots = [value.strip() for value in units_value.split(",")]
    last_unit_index = max((index for index, value in enumerate(unit_slots) if value), default=-1)
    if last_unit_index < point_count:
        return []
    return [
        ImportErrorRecord(
            row_number=row_number,
            field="Expected units",
            code="unit_without_point",
            message=(
                "Expected units has an entry without a corresponding Expected point."
            ),
        )
    ]


def _validate_asset_identity(row: dict[str, str], row_number: int) -> list[ImportErrorRecord]:
    """Accept the row as long as Asset ID OR Asset name is present (one-of)."""
    if row.get("Asset ID", "").strip() or row.get("Asset name", "").strip():
        return []
    return [
        ImportErrorRecord(
            row_number=row_number,
            field="Asset ID",
            code="missing_asset_identity",
            message="Provide an Asset ID or an Asset name (at least one is required).",
        )
    ]


def _conflicting_asset_topic_error(
    row: dict[str, str],
    row_number: int,
    roots_by_identity: dict[str, frozenset[str]],
) -> ImportErrorRecord | None:
    """Cross-row mqtt_register rule: one asset identity means ONE device.

    On-site 2026-07-13 a register row reused another asset's ID for a different
    device's topics; the import accepted every row and the run then grouped two
    devices under one ID — one looked missing, its neighbour doubled its issue
    list. The first errorless row registers the identity's topic root(s); a
    later row whose roots share nothing with them points at a different device
    and is rejected, naming both roots so the operator can tell which row is
    the copy-paste error. Rows sharing identity AND a root (e.g. one row per
    payload type) stay accepted — the run-time merge relies on them.
    """
    identity = row.get("Asset ID", "").strip() or row.get("Asset name", "").strip()
    if not identity:
        return None  # Row already rejected by _validate_asset_identity.
    identity_label = "Asset ID" if row.get("Asset ID", "").strip() else "Asset name"
    roots = _topic_roots(row.get("Expected topic", ""))
    if not roots:
        return None
    known_roots = roots_by_identity.get(identity)
    if known_roots is None:
        roots_by_identity[identity] = roots
        return None
    if known_roots & roots:
        # Same device: widen the registered set so a multi-root topic list
        # cannot smuggle a new root past later rows or falsely reject a
        # follow-up row that matches only the newly-listed root.
        roots_by_identity[identity] = known_roots | roots
        return None
    return ImportErrorRecord(
        row_number=row_number,
        field="Expected topic",
        code="conflicting_asset_topic",
        message=(
            f"{identity_label} '{identity}' is already registered for topic root "
            f"'{', '.join(sorted(known_roots))}' but this row points at "
            f"'{', '.join(sorted(roots))}' — a different device. "
            "Give each device its own unique Asset ID."
        ),
    )


# One "<port>/udp" entry of a comma-separated ports cell (each token is
# comma-split and stripped first; tolerate spaces around the slash and any
# casing of "udp").
_UDP_PORT_ENTRY = re.compile(r"(.+?)\s*/\s*udp", re.IGNORECASE)


def _warn_udp_ports(row: dict[str, str], row_number: int, field: str) -> list[ImportErrorRecord]:
    """Warn (never reject) about UDP entries in an ip_register ports cell.

    The IP scan engine is TCP-connect only (see engines/ip_scan.py): its port
    parser strips the /udp suffix and probes over TCP, so a UDP entry is
    tolerated but never actually verified. Without this note an "accepted"
    import reads as "my 47808/udp line gets checked".
    """
    warnings: list[ImportErrorRecord] = []
    for token in (part.strip() for part in row.get(field, "").split(",")):
        match = _UDP_PORT_ENTRY.fullmatch(token)
        if match is None:
            continue
        port = match.group(1).strip()
        if port == "47808":
            message = (
                "47808/udp is a UDP service — the IP scan verifies TCP ports only. "
                "UDP 47808 (BACnet/IP) is verified by the BACnet discovery run."
            )
        else:
            message = (
                f"{port}/udp is a UDP service — the IP scan verifies TCP ports only "
                "and cannot check this entry."
            )
        warnings.append(
            ImportErrorRecord(
                row_number=row_number,
                field=field,
                code="udp_port_not_verified",
                message=message,
            )
        )
    return warnings


def _base_row_validation(required_columns: tuple[str, ...], row: dict[str, str], row_number: int) -> list[ImportErrorRecord]:
    errors: list[ImportErrorRecord] = []
    for column in required_columns:
        if not row.get(column, "").strip():
            errors.append(
                ImportErrorRecord(
                    row_number=row_number,
                    field=column,
                    code="empty_required_field",
                    message=f"{column} must not be empty.",
                )
            )
    return errors


def _field_check(field: str, fn: Callable[[dict[str, str], int, str], list[ImportErrorRecord]]) -> Callable[[dict[str, str], int], list[ImportErrorRecord]]:
    def wrapper(row: dict[str, str], row_number: int) -> list[ImportErrorRecord]:
        return fn(row, row_number, field)

    return wrapper


PROFILES: dict[ImportType, ImportProfile] = {
    "ip_register": ImportProfile(
        import_type="ip_register",
        description="Expected network-connected device register for IP discovery.",
        required_columns=(
            "Project/site",
            "System",
            "Expected IP address",
            "Expected services/ports",
        ),
        # Asset ID / Asset name are one-of (see _validate_asset_identity), and
        # hostname is rarely used on commissioning networks. "Ports that should
        # not be enabled" lists ports flagged if found open (same syntax as
        # Expected services/ports). All are offered in the template and preserved
        # when present, but the base "must not be empty" check never applies.
        optional_columns=(
            "Asset ID",
            "Asset name",
            "Expected hostname",
            "Ports that should not be enabled",
        ),
        duplicate_key_fields=("Asset ID", "Expected IP address"),
        extra_checks=(_validate_asset_identity, _field_check("Expected IP address", _validate_ip)),
        warning_checks=(
            _field_check("Expected services/ports", _warn_udp_ports),
            _field_check("Ports that should not be enabled", _warn_udp_ports),
        ),
    ),
    "bacnet_register": ImportProfile(
        import_type="bacnet_register",
        description="Expected BACnet device register for BACnet discovery.",
        required_columns=(
            "Project/site",
            "System",
            "Asset ID",
            "Asset name",
            "BACnet device instance",
            "BACnet network",
            "IP address",
        ),
        duplicate_key_fields=("Asset ID", "BACnet device instance"),
        extra_checks=(
            _field_check("IP address", _validate_ip),
            _field_check("BACnet device instance", _validate_numeric),
            _field_check("BACnet network", _validate_numeric),
        ),
    ),
    "mqtt_register": ImportProfile(
        import_type="mqtt_register",
        description="Expected MQTT asset and topic register for MQTT discovery.",
        required_columns=(
            "Project/site",
            "System",
            "Expected topic",
            "Expected schema version",
            "Expected points",
            "Expected units",
            "Expected reporting interval",
            "Source protocol",
        ),
        # Asset ID / Asset name are one-of (_validate_asset_identity). Notes and
        # Payload type are optional: a blank Payload type means "check metadata,
        # state and pointset". Make/Model/GUID feed the UDMI metadata/state match
        # (udmi_validation expected_schedule); Site/Serial/Room/Firmware are
        # captured for the same comparison surface.
        optional_columns=(
            "Asset ID",
            "Asset name",
            "Payload type",
            "Notes",
            "Site",
            "Serial number",
            "Room",
            "GUID",
            "Make",
            "Model",
            "Firmware",
        ),
        duplicate_key_fields=("Asset ID", "Expected topic"),
        extra_checks=(
            _validate_asset_identity,
            _field_check("Expected topic", _validate_topic),
            _field_check("Expected topic", _validate_mqtt_asset_topic),
            _field_check("Expected reporting interval", _validate_positive_numeric),
            _validate_payload_type,
            _validate_mqtt_point_unit_pairs,
        ),
    ),
    "asset_validation": ImportProfile(
        import_type="asset_validation",
        description="Asset-level validation file for expected online and integration status.",
        required_columns=(
            "Project/site",
            "System",
            "Asset ID",
            "Asset name",
            "Source protocol",
            "Expected online status",
            "Expected topic or device reference",
            "Location",
        ),
        duplicate_key_fields=("Asset ID",),
    ),
    "bacnet_points": ImportProfile(
        import_type="bacnet_points",
        description="Expected BACnet point validation register.",
        required_columns=(
            "Asset ID",
            "Device instance",
            "BACnet network",
            "Object type",
            "Object instance",
            "Object name",
            "Expected point name",
            "Expected units",
            "Expected value type",
            "Required/optional flag",
        ),
        duplicate_key_fields=("Asset ID", "Object instance", "Expected point name"),
        extra_checks=(
            _field_check("Device instance", _validate_numeric),
            _field_check("BACnet network", _validate_numeric),
            _field_check("Object instance", _validate_numeric),
            _field_check("Expected units", _validate_units),
        ),
    ),
    "mqtt_points": ImportProfile(
        import_type="mqtt_points",
        description="Expected MQTT point extraction validation register.",
        required_columns=(
            "Asset ID",
            "Topic",
            "Payload type",
            "JSON path or field name",
            "Expected point name",
            "Expected units",
            "Expected value type",
            "Required/optional flag",
            "Expected reporting interval",
        ),
        duplicate_key_fields=("Asset ID", "JSON path or field name", "Expected point name"),
        extra_checks=(
            _field_check("Topic", _validate_topic),
            _field_check("Expected units", _validate_units),
            _field_check("Expected reporting interval", _validate_numeric),
        ),
    ),
    "mapping": ImportProfile(
        import_type="mapping",
        description="BACnet-to-MQTT mapping validation file.",
        required_columns=(
            "Asset ID",
            "BACnet device instance",
            "BACnet object type",
            "BACnet object instance",
            "BACnet object name",
            "BACnet units",
            "MQTT topic",
            "MQTT field/path",
            "MQTT units",
            "Tolerance",
            "Mapping required flag",
        ),
        duplicate_key_fields=("Asset ID", "BACnet object instance", "MQTT field/path"),
        extra_checks=(
            _field_check("BACnet device instance", _validate_numeric),
            _field_check("BACnet object instance", _validate_numeric),
            _field_check("MQTT topic", _validate_topic),
            _field_check("BACnet units", _validate_units),
            _field_check("MQTT units", _validate_units),
        ),
    ),
    "tolerances": ImportProfile(
        import_type="tolerances",
        description="Point-level tolerances used by comparison validation.",
        required_columns=("Asset ID", "Point name", "Tolerance"),
        duplicate_key_fields=("Asset ID", "Point name"),
        extra_checks=(_field_check("Tolerance", _validate_tolerance),),
    ),
}

EXAMPLE_ROWS: dict[ImportType, dict[str, str]] = {
    "ip_register": {
        "Project/site": "ElectraCom / Block B Plantroom",
        "System": "BMS",
        "Asset ID": "AHU-L03-017",
        "Asset name": "Level 3 AHU",
        "Expected IP address": "10.10.25.117",
        "Expected hostname": "ahu-l03-017",
        "Expected services/ports": "47808/udp, 443/tcp",
        "Ports that should not be enabled": "23/tcp, 21/tcp",
    },
    "bacnet_register": {
        "Project/site": "ElectraCom / Block B Plantroom",
        "System": "BMS",
        "Asset ID": "AHU-L03-017",
        "Asset name": "Level 3 AHU",
        "BACnet device instance": "1532117",
        "BACnet network": "1532",
        "IP address": "10.10.25.117",
    },
    "mqtt_register": {
        "Project/site": "ElectraCom / Block B Plantroom",
        "System": "BMS",
        "Asset ID": "MTR-ENERGY-009",
        "Asset name": "Energy Meter 9",
        # "prefix/#" encapsulates this asset's metadata/state/events topics.
        "Expected topic": "electracom/sct/1532/meter/009/#",
        # Blank Payload type => check metadata, state and pointset.
        "Payload type": "",
        "Expected schema version": "1.5.2",
        # Multiple points / units for one asset, comma-separated.
        "Expected points": "energy_sensor,power_sensor",
        "Expected units": "kwh,kw",
        "Expected reporting interval": "60",
        "Source protocol": "MQTT",
        "Notes": "",
        "Site": "Block B",
        "Serial number": "EM-009-SN-00421",
        "Room": "Level 3 plantroom",
        "GUID": "ifc://electracom/EM-1001001",
        "Make": "ExpectedCo",
        "Model": "Model-A",
        "Firmware": "1.4.2",
    },
    "asset_validation": {
        "Project/site": "ElectraCom / Block B Plantroom",
        "System": "BMS",
        "Asset ID": "AHU-L03-017",
        "Asset name": "Level 3 AHU",
        "Source protocol": "BACnet + MQTT",
        "Expected online status": "Online",
        "Expected topic or device reference": "1532117",
        "Location": "Level 3 plantroom",
    },
    "bacnet_points": {
        "Asset ID": "AHU-L03-017",
        "Device instance": "1532117",
        "BACnet network": "1532",
        "Object type": "analogInput",
        "Object instance": "300001",
        "Object name": "supply_air_temperature",
        "Expected point name": "supply_air_temperature_sensor",
        "Expected units": "degrees-celsius",
        "Expected value type": "number",
        "Required/optional flag": "required",
    },
    "mqtt_points": {
        "Asset ID": "AHU-L03-017",
        "Topic": "electracom/sct/1532/ahu/l03/events/pointset",
        "Payload type": "pointset",
        "JSON path or field name": "pointset.points.supply_air_temperature_sensor.present_value",
        "Expected point name": "supply_air_temperature_sensor",
        "Expected units": "degrees-celsius",
        "Expected value type": "number",
        "Required/optional flag": "required",
        "Expected reporting interval": "60",
    },
    "mapping": {
        "Asset ID": "AHU-L03-017",
        "BACnet device instance": "1532117",
        "BACnet object type": "analogInput",
        "BACnet object instance": "300001",
        "BACnet object name": "supply_air_temperature",
        "BACnet units": "degrees-celsius",
        "MQTT topic": "electracom/sct/1532/ahu/l03/events/pointset",
        "MQTT field/path": "pointset.points.supply_air_temperature_sensor.present_value",
        "MQTT units": "degrees-celsius",
        "Tolerance": "0.5",
        "Mapping required flag": "required",
    },
    "tolerances": {
        "Asset ID": "AHU-L03-017",
        "Point name": "supply_air_temperature_sensor",
        "Tolerance": "0.5",
    },
}


def _decode_csv_bytes(file_bytes: bytes) -> str:
    """Decode an uploaded CSV, tolerating the ways Excel actually saves them.

    Raises ValueError (which the route turns into a 400 with this text) rather
    than letting a raw codec message reach the operator. Best-effort by design:
    this is not encoding detection. A file that is neither UTF-8 nor UTF-16 is
    *assumed* to be Windows-1252 (Excel's "CSV (comma delimited)" save), so a
    cp1250 file decodes with a few wrong accents instead of erroring.
    """
    if file_bytes.startswith(b"PK\x03\x04"):
        raise ValueError(
            "This file is an Excel XLSX workbook renamed to .csv — upload it with an .xlsx "
            "extension or re-save it as 'CSV UTF-8 (comma delimited)'."
        )
    if file_bytes.startswith((codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE)):
        # Excel's "Unicode Text" save. The utf-16 codec consumes either BOM.
        text = file_bytes.decode("utf-16")
    else:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = file_bytes.decode("cp1252")
            except UnicodeDecodeError as error:
                # cp1252 leaves 0x81/0x8D/0x8F/0x90/0x9D unmapped, so this is
                # reachable for genuinely binary uploads.
                raise ValueError(
                    "CSV file is not readable text (tried UTF-8 and Windows-1252). Re-save it "
                    "from Excel as 'CSV UTF-8 (comma delimited)'."
                ) from error
    if "\x00" in text:
        # BOM-less UTF-16-LE ASCII is valid UTF-8 full of NULs, so it decodes
        # "successfully" into garbage columns — catch it here instead.
        raise ValueError(
            "CSV file contains binary data (it may be UTF-16 without a byte-order mark). "
            "Re-save it from Excel as 'CSV UTF-8 (comma delimited)'."
        )
    return text


class ImportService:
    """Import batches with database-backed metadata.

    Summary/accepted_rows/errors live in the database (ImportRepository); the
    uploaded original file is still written to disk under the imports root.
    """

    def __init__(self, engine: Engine | None = None) -> None:
        ensure_runtime_directories()
        self._repository = ImportRepository(engine if engine is not None else get_engine())

    def list_profiles(self) -> list[ImportProfileSummary]:
        return [profile.as_summary() for profile in PROFILES.values()]

    def build_template(self, import_type: ImportType, file_type: str) -> bytes:
        profile = PROFILES[import_type]
        example = EXAMPLE_ROWS[import_type]
        if file_type == "csv":
            return self._build_csv_template(profile, example)
        if file_type == "xlsx":
            return self._build_xlsx_template(profile, example)
        raise ValueError("Template format must be csv or xlsx.")

    def create_import(
        self,
        *,
        import_type: ImportType,
        file_name: str,
        file_bytes: bytes,
        project_id: str | None,
        site_id: str | None,
    ) -> tuple[ImportBatchSummary, ImportErrorReport]:
        profile = PROFILES[import_type]
        file_type = self._detect_file_type(file_name)
        rows = self._parse_rows(file_type=file_type, file_bytes=file_bytes)

        import_id = f"imp_{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}_{token_hex(4)}"
        stored_file_name = f"{import_id}_{Path(file_name).name}"
        stored_path = IMPORT_FILES_ROOT / stored_file_name
        stored_path.write_bytes(file_bytes)

        mapped_rows, missing_columns = self._canonicalize_rows(profile, rows)
        errors: list[ImportErrorRecord] = []
        warnings: list[ImportErrorRecord] = []
        accepted_rows: list[dict[str, str]] = []

        if missing_columns:
            errors.extend(
                ImportErrorRecord(
                    field=column,
                    code="missing_required_column",
                    message=f"Required column '{column}' is missing.",
                )
                for column in missing_columns
            )
        else:
            seen_keys: set[tuple[str, ...]] = set()
            # Cross-row mqtt_register state: identity -> topic roots of its
            # first errorless row (per batch, so nothing leaks across imports).
            mqtt_roots_by_identity: dict[str, frozenset[str]] = {}
            for row_number, row in enumerate(mapped_rows, start=2):
                row_errors = profile.validate_row(row, row_number)
                # Warnings never affect acceptance and are collected even for
                # rejected rows, so one re-upload can fix everything at once.
                warnings.extend(profile.collect_warnings(row, row_number))
                duplicate_key = tuple(row.get(field, "").strip() for field in profile.duplicate_key_fields)
                if duplicate_key and any(duplicate_key):
                    if duplicate_key in seen_keys:
                        row_errors.append(
                            ImportErrorRecord(
                                row_number=row_number,
                                code="duplicate_row",
                                message=f"Duplicate record detected for key fields {', '.join(profile.duplicate_key_fields)}.",
                            )
                        )
                    else:
                        seen_keys.add(duplicate_key)
                if import_type == "mqtt_register" and not row_errors:
                    conflict = _conflicting_asset_topic_error(row, row_number, mqtt_roots_by_identity)
                    if conflict is not None:
                        row_errors.append(conflict)
                if row_errors:
                    errors.extend(row_errors)
                else:
                    accepted_rows.append(row)

        status = self._status(
            total_rows=len(mapped_rows),
            accepted_rows=len(accepted_rows),
            missing_columns=missing_columns,
        )
        summary = ImportBatchSummary(
            import_id=import_id,
            import_type=import_type,
            file_name=Path(file_name).name,
            file_type=file_type,
            project_id=project_id,
            site_id=site_id,
            total_rows=len(mapped_rows),
            accepted_rows=len(accepted_rows),
            rejected_rows=max(len(mapped_rows) - len(accepted_rows), 0),
            status=status,
            missing_columns=missing_columns,
            warnings=warnings,
            stored_file_name=stored_file_name,
            created_at=datetime.now(UTC),
        )
        error_report = ImportErrorReport(import_id=import_id, errors=errors)

        self._repository.create(
            import_id=import_id,
            import_type=import_type,
            project_id=project_id,
            site_id=site_id,
            original_filename=Path(file_name).name,
            stored_file_path=str(stored_path),
            summary=summary.model_dump(mode="json"),
            accepted_rows=accepted_rows,
            errors=[error.model_dump(mode="json") for error in errors],
            created_at=summary.created_at,
        )

        return summary, error_report

    def get_import(self, import_id: str) -> ImportBatchSummary:
        return ImportBatchSummary.model_validate(self._repository.get_summary(import_id))

    def get_import_errors(self, import_id: str) -> ImportErrorReport:
        return ImportErrorReport.model_validate(self._repository.get_errors(import_id))

    def _detect_file_type(self, file_name: str) -> str:
        suffix = Path(file_name).suffix.lower()
        if suffix == ".csv":
            return "csv"
        if suffix == ".xlsx":
            return "xlsx"
        raise ValueError("Only CSV and XLSX imports are supported.")

    def _parse_rows(self, *, file_type: str, file_bytes: bytes) -> list[dict[str, str]]:
        if file_type == "csv":
            return self._parse_csv(file_bytes)
        if file_type == "xlsx":
            return self._parse_xlsx(file_bytes)
        raise ValueError(f"Unsupported file type: {file_type}")

    def _parse_csv(self, file_bytes: bytes) -> list[dict[str, str]]:
        text = _decode_csv_bytes(file_bytes)
        rows: list[dict[str, str]] = []
        try:
            # newline="" hands the line endings to the csv module rather than to
            # StringIO's universal-newline translation: without it a CR-only save
            # (classic Excel for Mac) raises csv.Error — which is NOT a ValueError,
            # so it escaped the route's 400 handler as a 500.
            reader = csv.DictReader(io.StringIO(text, newline=""))
            if reader.fieldnames is None:
                raise ValueError("CSV file is missing a header row.")
            if len(reader.fieldnames) == 1 and any(sep in reader.fieldnames[0] for sep in (";", "\t")):
                # A regional Excel save. Without this the whole header parses as
                # one column and every required column is reported missing, which
                # tells the operator nothing about the actual problem. We say so
                # rather than re-parsing with the guessed delimiter: auto-guessing
                # mis-splits quoted content and would quietly accept a format the
                # templates never promised.
                raise ValueError(
                    "CSV file is not comma-delimited (its header row uses ';' or tabs — a regional "
                    "Excel CSV save). Re-save it as 'CSV UTF-8 (comma delimited)'."
                )
            for raw_row in reader:
                row = {_normalize_header(key): str(value or "").strip() for key, value in raw_row.items() if key is not None}
                if any(value.strip() for value in row.values()):
                    rows.append(row)
        except csv.Error as error:
            # Residual csv failures (e.g. a field above csv.field_size_limit).
            # csv.Error is not a ValueError, so it must be converted here or the
            # route answers 500. The ValueErrors raised above pass through.
            raise ValueError(f"CSV file could not be parsed: {error}") from error
        return rows

    def _parse_xlsx(self, file_bytes: bytes) -> list[dict[str, str]]:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = [_normalize_header(value) for value in next(iterator)]
        except StopIteration as error:
            raise ValueError("XLSX file is empty.") from error
        rows: list[dict[str, str]] = []
        for values in iterator:
            row = {headers[index]: str(value or "").strip() for index, value in enumerate(values)}
            if any(value.strip() for value in row.values()):
                rows.append(row)
        return rows

    def _canonicalize_rows(self, profile: ImportProfile, rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[str]]:
        if not rows:
            return [], list(profile.required_columns)

        first_row = rows[0]
        header_lookup = {_normalize_key(header): header for header in first_row.keys()}
        missing_columns = [column for column in profile.required_columns if _normalize_key(column) not in header_lookup]

        mapped_rows: list[dict[str, str]] = []
        for row in rows:
            mapped_row: dict[str, str] = {}
            for header, value in row.items():
                mapped_row[header] = value.strip()
            # Canonicalise required AND optional columns so header casing/spacing
            # variants land under the profile's canonical names (optional ones
            # only when actually present in the file).
            for column in profile.template_columns:
                source_header = header_lookup.get(_normalize_key(column))
                if source_header is not None:
                    mapped_row[column] = row.get(source_header, "").strip()
            mapped_rows.append(mapped_row)

        return mapped_rows, missing_columns

    def _status(self, *, total_rows: int, accepted_rows: int, missing_columns: list[str]) -> str:
        if missing_columns or total_rows == 0 or accepted_rows == 0:
            return "rejected"
        if accepted_rows == total_rows:
            return "accepted"
        return "partial"

    def _build_csv_template(self, profile: ImportProfile, example: dict[str, str]) -> bytes:
        columns = list(profile.template_columns)
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=columns, lineterminator="\n")
        writer.writeheader()
        writer.writerow({column: example.get(column, "") for column in columns})
        return buffer.getvalue().encode("utf-8-sig")

    def _build_xlsx_template(self, profile: ImportProfile, example: dict[str, str]) -> bytes:
        columns = list(profile.template_columns)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Default import template"
        sheet.append(columns)
        sheet.append([example.get(column, "") for column in columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        header_fill = PatternFill(fill_type="solid", fgColor="DCEBFF")
        header_font = Font(bold=True, color="1D1D1F")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells]
            width = min(max(max(len(value) for value in values) + 2, 14), 42)
            sheet.column_dimensions[column_cells[0].column_letter].width = width

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
