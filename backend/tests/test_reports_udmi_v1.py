"""Versioned UDMI report contract coverage across every report renderer."""

from __future__ import annotations

import copy
import hashlib
import io
import json
import os
import unittest
import xml.etree.ElementTree as ElementTree
import zipfile
from datetime import UTC, datetime
from unittest import mock
from uuid import uuid4

from harness import ApiTestCase
from openpyxl import load_workbook
from pydantic import ValidationError

_API_KEY = "test-reports-udmi-v1-key"
_ENV_OVERRIDES = {
    "JOB_EXECUTION_MODE": "inline",
    "AUTH_MODE": "api_key",
    "API_KEY": _API_KEY,
}


def _metric_groups(
    *,
    assets: tuple[int, int, int, int, int],
    payloads: tuple[int, int, int, int],
    faults: tuple[int, int, int, int, int, int],
    issues: tuple[int, int],
) -> dict[str, dict[str, int]]:
    return {
        "asset_metrics": dict(
            zip(
                ("expected", "observed", "not_observed", "with_issues", "successfully_validated"),
                assets,
                strict=True,
            )
        ),
        "payload_metrics": dict(
            zip(
                ("expected", "received", "with_issues", "successfully_validated"),
                payloads,
                strict=True,
            )
        ),
        "fault_metrics": dict(
            zip(
                (
                    "payload_formatting_issues",
                    "missing_points",
                    "point_naming_issues",
                    "additional_points",
                    "stale_or_cadence",
                    "other_issues",
                ),
                faults,
                strict=True,
            )
        ),
        "issue_metrics": dict(zip(("blocking", "warning"), issues, strict=True)),
    }


_TOTALS = _metric_groups(
    assets=(3, 2, 1, 2, 1),
    payloads=(7, 5, 2, 4),
    faults=(1, 1, 1, 1, 1, 1),
    issues=(3, 3),
)

_V1_SUMMARY = {
    "schema_version": "1.0",
    **_TOTALS,
    "system_metrics": [
        {
            "system": "BMS",
            **_metric_groups(
                assets=(2, 2, 0, 1, 1),
                payloads=(6, 5, 1, 4),
                faults=(1, 1, 1, 1, 0, 0),
                issues=(2, 2),
            ),
        },
        {
            "system": "LTG",
            **_metric_groups(
                assets=(1, 0, 1, 1, 0),
                payloads=(1, 0, 1, 0),
                faults=(0, 0, 0, 0, 1, 1),
                issues=(1, 1),
            ),
        },
    ],
    "asset_results": [
        {
            "asset_id": "AHU-1",
            "system": "BMS",
            "observed": True,
            "expected_payloads": 3,
            "received_payloads": 3,
            "all_expected_payloads_received": True,
            "all_received_payloads_successfully_validated": True,
            "successfully_validated": True,
            "issue_count": 0,
            "blocking_issue_count": 0,
            "last_observed_at": "2026-07-22T10:20:30+00:00",
            "payload_results": [],
        },
        {
            "asset_id": "FCU-2",
            "system": "BMS",
            "observed": True,
            "expected_payloads": 3,
            "received_payloads": 2,
            "all_expected_payloads_received": False,
            # The received subset passed, but one expected payload is missing.
            # The schedule must use successfully_validated and show No.
            "all_received_payloads_successfully_validated": True,
            "successfully_validated": False,
            "issue_count": 3,
            "blocking_issue_count": 2,
            "last_observed_at": None,
            "payload_results": [],
        },
        {
            "asset_id": "LIGHT-3",
            "system": "LTG",
            "observed": False,
            "expected_payloads": 1,
            "received_payloads": 0,
            "all_expected_payloads_received": False,
            "all_received_payloads_successfully_validated": False,
            "successfully_validated": False,
            "issue_count": 3,
            "blocking_issue_count": 1,
            "last_observed_at": None,
            "payload_results": [],
        },
    ],
    "fault_rows": [
        {
            "issue_id": f"issue-{index}",
            "asset_id": "FCU-2" if index < 4 else "LIGHT-3",
            "system": "BMS" if index < 4 else "LTG",
            "payload_type": "pointset",
            "category": category,
            "severity": "high" if index % 2 == 0 else "low",
            "description": f"Recorded {category} evidence.",
            "point_name": "zone_air_temperature_sensor",
            "expected_value": "expected",
            "observed_value": "observed",
            "suggested_action": "Review the retained evidence and correct the source.",
            "raw_evidence_uri": f"evidence://issue-{index}",
        }
        for index, category in enumerate(
            (
                "payload_formatting_issues",
                "missing_points",
                "point_naming_issues",
                "additional_points",
                "stale_or_cadence",
                "other_issues",
            ),
            start=1,
        )
    ],
}


def _expected_detail_columns() -> tuple[str, ...]:
    return (
        "Issue ID",
        "Asset ID",
        "System",
        "Payload",
        "Category",
        "Point",
        "Expected",
        "Observed",
        "Mismatch",
        "Suggested Action",
        "Description",
    )


_SCOPABLE_SUMMARY = {
    "schema_version": "1.1",
    "asset_metrics": {
        "expected": 2,
        "observed": 2,
        "not_observed": 0,
        "with_issues": 2,
        "successfully_validated": 1,
        "unexpected": 1,
    },
    "payload_metrics": {
        "expected": 5,
        "received": 4,
        "not_received": 1,
        "with_issues": 2,
        "successfully_validated": 3,
    },
    "fault_metrics": {
        "payload_formatting_issues": 0,
        "missing_points": 1,
        "point_naming_issues": 1,
        "additional_points": 1,
        "stale_or_cadence": 0,
        "other_issues": 2,
    },
    "issue_metrics": {"blocking": 4, "warning": 1},
    "system_metrics": [
        {
            "system": "BMS",
            **_metric_groups(
                assets=(1, 1, 0, 1, 0),
                payloads=(3, 2, 1, 1),
                faults=(0, 1, 1, 0, 0, 1),
                issues=(3, 0),
            ),
        },
        {
            "system": "LTG",
            **_metric_groups(
                assets=(1, 1, 0, 1, 1),
                payloads=(2, 2, 1, 2),
                faults=(0, 0, 0, 1, 0, 0),
                issues=(0, 1),
            ),
        },
    ],
    "asset_results": [
        {
            "asset_id": "A-1",
            "system": "BMS",
            "observed": True,
            "expected_payloads": 3,
            "received_payloads": 2,
            "all_expected_payloads_received": False,
            "all_received_payloads_successfully_validated": False,
            "successfully_validated": False,
            "issue_count": 3,
            "blocking_issue_count": 3,
            "last_observed_at": "2026-07-23T10:02:00+00:00",
            "payload_results": [
                {
                    "payload_type": "state",
                    "expected": True,
                    "received": True,
                    "has_issues": False,
                    "blocking_issue_count": 0,
                    "successfully_validated": True,
                    "topic": "site/a-1/state",
                    "received_at": "2026-07-23T10:00:00+00:00",
                },
                {
                    "payload_type": "metadata",
                    "expected": True,
                    "received": False,
                    "has_issues": True,
                    "blocking_issue_count": 1,
                    "successfully_validated": False,
                    "topic": "site/a-1/metadata",
                    "received_at": None,
                },
                {
                    "payload_type": "pointset",
                    "expected": True,
                    "received": True,
                    "has_issues": True,
                    "blocking_issue_count": 1,
                    "successfully_validated": False,
                    "topic": "site/a-1/pointset",
                    "received_at": "2026-07-23T10:02:00+00:00",
                },
            ],
        },
        {
            "asset_id": "B-1",
            "system": "LTG",
            "observed": True,
            "expected_payloads": 2,
            "received_payloads": 2,
            "all_expected_payloads_received": True,
            "all_received_payloads_successfully_validated": True,
            "successfully_validated": True,
            "issue_count": 1,
            "blocking_issue_count": 0,
            "last_observed_at": "2026-07-23T10:03:00+00:00",
            "payload_results": [
                {
                    "payload_type": "state",
                    "expected": True,
                    "received": True,
                    "has_issues": True,
                    "blocking_issue_count": 0,
                    "successfully_validated": True,
                    "topic": "site/b-1/state",
                    "received_at": "2026-07-23T10:01:00+00:00",
                },
                {
                    "payload_type": "pointset",
                    "expected": True,
                    "received": True,
                    "has_issues": False,
                    "blocking_issue_count": 0,
                    "successfully_validated": True,
                    "topic": "site/b-1/pointset",
                    "received_at": "2026-07-23T10:03:00+00:00",
                },
            ],
        },
    ],
    "fault_rows": [
        {
            "issue_id": "a-metadata-missing",
            "asset_id": "A-1",
            "system": "BMS",
            "payload_type": "metadata",
            "category": "missing_points",
            "severity": "high",
            "description": "Metadata was not received.",
            "point_name": "manufacturer",
            "expected_value": "Acme",
            "observed_value": None,
            "suggested_action": "Publish metadata.",
            "raw_evidence_uri": "evidence://metadata",
        },
        {
            "issue_id": "a-point-name",
            "asset_id": "A-1",
            "system": "BMS",
            "payload_type": "pointset",
            "category": "point_naming_issues",
            "severity": "high",
            "description": "Point name differs from the register.",
            "point_name": "zone_temp",
            "expected_value": "zone_air_temperature_sensor",
            "observed_value": "zone_temp",
            "suggested_action": "Rename the point.",
            "raw_evidence_uri": "evidence://point",
        },
        {
            "issue_id": "a-asset-wide",
            "asset_id": "A-1",
            "system": "BMS",
            "payload_type": None,
            "category": "other_issues",
            "severity": "high",
            "description": "Asset-wide finding.",
            "point_name": None,
            "expected_value": None,
            "observed_value": None,
            "suggested_action": "Review the asset.",
            "raw_evidence_uri": "evidence://asset",
        },
        {
            "issue_id": "run-wide",
            "asset_id": None,
            "system": "Unspecified",
            "payload_type": None,
            "category": "other_issues",
            "severity": "high",
            "description": "Run-wide broker finding.",
            "point_name": None,
            "expected_value": None,
            "observed_value": None,
            "suggested_action": "Review broker evidence.",
            "raw_evidence_uri": "evidence://run",
        },
        {
            "issue_id": "b-state-note",
            "asset_id": "B-1",
            "system": "LTG",
            "payload_type": "state",
            "category": "additional_points",
            "severity": "low",
            "description": "State carries an extra field.",
            "point_name": "extra",
            "expected_value": None,
            "observed_value": "present",
            "suggested_action": "Review the field.",
            "raw_evidence_uri": "evidence://state",
        },
    ],
    "unexpected_devices": [
        {
            "id": "rogue-1",
            "topic_root": "site/rogue-1",
            "topics": ["site/rogue-1/state"],
            "last_seen": "2026-07-23T10:04:00+00:00",
        }
    ],
    "unexpected_devices_measured": True,
    "unexpected_devices_measurement_scope": "site/#",
}


class UdmiV1ReportTests(ApiTestCase):
    env = _ENV_OVERRIDES
    client_headers = {"X-API-Key": _API_KEY}

    @classmethod
    def before_client(cls) -> None:
        import atexit
        import shutil
        import tempfile
        from pathlib import Path

        cls._temp_runtime = tempfile.mkdtemp(prefix="sct-reports-udmi-v1-")
        atexit.register(shutil.rmtree, cls._temp_runtime, ignore_errors=True)
        secrets_root = Path(cls._temp_runtime) / "secrets"
        secrets_root.mkdir(parents=True, exist_ok=True)

        import app.services.reports_integrity as integrity_module

        cls._patcher = mock.patch.object(integrity_module, "SECRETS_ROOT", secrets_root)
        cls._patcher.start()

    @classmethod
    def tearDownClass(cls) -> None:
        super().tearDownClass()
        cls._patcher.stop()

    def _seed_run(
        self,
        *,
        project_id: str = "demo-project",
        site_id: str = "demo-site",
        job_type: str = "udmi_validation",
        status: str = "succeeded",
        summary: dict | None = _V1_SUMMARY,
        parameters: dict | None = None,
        issues: list[dict] | None = None,
        result_summary_extra: dict | None = None,
    ) -> str:
        from app.schemas.jobs import JobCreateRequest
        from app.services.run_service import RunService
        from lifecycle_helpers import finish_run

        service = RunService()
        run = service.create_job_run(
            JobCreateRequest(
                project_id=project_id,
                site_id=site_id,
                job_type=job_type,
                parameters={
                    # Report fixtures never contact a broker. Each source gets
                    # its own reserved identity so an active-run test cannot
                    # block unrelated renderer cases.
                    "broker_host": f"report-fixture-{uuid4().hex}.invalid",
                    **(parameters or {}),
                },
            ),
            expected_job_type=job_type,
        )
        frozen_summary = (
            {
                "validation_summary_v1": summary,
                **(result_summary_extra or {}),
            }
            if summary is not None
            else None
        )
        if status in {"succeeded", "failed", "cancelled"}:
            finish_run(
                service,
                run.run_id,
                status=status,
                summary=frozen_summary,
                issues=issues,
            )
        elif status == "running":
            owned = service.claim_owned_run(run.run_id)
            self.assertIsNotNone(owned)
            active_owned = getattr(self, "_active_owned_test_runs", {})
            active_owned[run.run_id] = owned
            self._active_owned_test_runs = active_owned
            if frozen_summary is not None:
                owned.update_result_summary(run.run_id, frozen_summary, merge=False)
            owned.update_run_status(
                run.run_id,
                status="running",
                stage="running",
                progress_percent=50,
            )
        else:
            if frozen_summary is not None:
                service.update_result_summary(run.run_id, frozen_summary)
            if issues is not None:
                service.replace_issues(run.run_id, issues)
        return run.run_id

    def _finish_active_test_run(self, run_id: str) -> None:
        from app.services.run_service import RunService

        active_owned = getattr(self, "_active_owned_test_runs", {})
        owned = active_owned.pop(run_id, None)
        if owned is None:
            owned = RunService().claim_owned_run(run_id)
        self.assertIsNotNone(owned)
        outcome = owned.update_run_status(
            run_id,
            status="cancelled",
            stage="test_cleanup",
            progress_percent=100,
        )
        self.assertNotEqual(outcome.get("status"), "ownership_lost")

    def _create_report(
        self,
        output_format: str,
        source_run_ids: list[str],
        *,
        title: str | None = "  Site & <A> Validation  ",
        report_type: str = "udmi_validation",
        udmi_scope: dict | None = None,
        udmi_report_variant: str = "technical",
    ) -> dict:
        payload: dict[str, object] = {
            "project_id": "demo-project",
            "site_id": "demo-site",
            "report_type": report_type,
            "output_format": output_format,
            "source_run_ids": source_run_ids,
        }
        if title is not None:
            payload["report_title"] = title
        if udmi_scope is not None:
            payload["udmi_scope"] = udmi_scope
        payload["udmi_report_variant"] = udmi_report_variant
        response = self.client.post("/api/v1/reports", json=payload)
        self.assertEqual(response.status_code, 200, response.text)
        return response.json()

    def _download(self, report_id: str):
        response = self.client.get(f"/api/v1/reports/{report_id}/download")
        self.assertEqual(response.status_code, 200, response.text)
        return response

    def test_title_defaults_trimming_and_validation(self) -> None:
        custom = self._create_report("zip", [], title="  Site A Validation  ")
        self.assertEqual(custom["report_title"], "Site A Validation")
        stored = self.client.get(f"/api/v1/reports/{custom['report_id']}").json()
        self.assertEqual(stored["report_title"], "Site A Validation")

        udmi_default = self._create_report("zip", [], title=None)
        self.assertEqual(udmi_default["report_title"], "UDMI Validation Report")
        general_default = self._create_report(
            "zip",
            [],
            title=None,
            report_type="evidence_pack",
        )
        self.assertEqual(general_default["report_title"], "Smart Commissioning Report")

        from app.schemas.jobs import ReportRequest

        for title in (
            "   ",
            "bad\ncontrol",
            "bad\ud800surrogate",
            "bad\ufffecharacter",
            "bad\uffffcharacter",
            "x" * 161,
        ):
            with self.subTest(title=repr(title)), self.assertRaises(ValidationError):
                ReportRequest(
                    project_id="demo-project",
                    site_id="demo-site",
                    report_type="udmi_validation",
                    report_title=title,
                )

    def test_report_formats_share_evidence_set_and_declare_full_scope(self) -> None:
        source_id = self._seed_run(
            result_summary_extra={
                "broker_capture_attempted": True,
                "capture_mode": "bounded",
                "capture_window_seconds": 3600,
                "capture_started_at": "2026-08-05T08:00:00+00:00",
                "capture_ended_at": "2026-08-05T09:00:00+00:00",
                "window_completed": True,
                "termination_reason": "window_elapsed",
            },
        )

        reports = [
            self._create_report(output_format, [source_id])
            for output_format in ("zip", "pdf", "docx", "xlsx")
        ]

        evidence_set_ids = {report["evidence_set_id"] for report in reports}
        self.assertEqual(len(evidence_set_ids), 1)
        evidence_set_id = next(iter(evidence_set_ids))
        self.assertRegex(evidence_set_id, r"^evidence_[0-9a-f]{24}$")
        from app.services.run_service import RunService

        for report in reports:
            manifest = RunService().get_run(report["report_id"]).result_summary[
                "artifact_manifest"
            ]
            self.assertEqual(manifest["evidence_set_id"], evidence_set_id)

        with zipfile.ZipFile(io.BytesIO(self._download(reports[0]["report_id"]).content)) as archive:
            validation = json.loads(archive.read("validation_summary.json"))

        self.assertEqual(validation["evidence_set_id"], evidence_set_id)
        self.assertEqual(validation["report_scope"]["scope_kind"], "full_source_run")
        self.assertEqual(validation["report_scope"]["source_run_ids"], [source_id])
        self.assertEqual(validation["filter_provenance"]["text"], "")

    def test_technical_report_preserves_run_provenance_and_register_digest(self) -> None:
        from app.services.report_artifacts import verify_signed_manifest
        from app.services.run_service import RunService

        with mock.patch.dict(
            os.environ,
            {
                "SMART_COMMISSIONING_SOURCE_COMMIT": "b" * 40,
                "SMART_COMMISSIONING_PORTABLE_EXE_SHA256": "c" * 64,
            },
        ):
            source_id = self._seed_run(
                parameters={
                    "register_import_id": "imp-provenance-1",
                    "register_import_filename": "approved-register.csv",
                    "register_revision": "rev-2026-08-06",
                    "register_sha256": "a" * 64,
                    "machine_reference": "commissioning-laptop-01",
                    "broker_reference": "broker-profile-01",
                    "operator_reference": "operator-01",
                    "topic_filter": "demo-site/#",
                },
                result_summary_extra={
                    "broker_capture_attempted": True,
                    "capture_mode": "bounded",
                    "capture_window_seconds": 3600,
                    "capture_started_at": "2026-08-05T08:00:00+00:00",
                    "capture_ended_at": "2026-08-05T09:00:00+00:00",
                    "capture_duration_seconds": 3600.0,
                    "window_completed": True,
                    "termination_reason": "window_elapsed",
                },
            )
        report = self._create_report("zip", [source_id])
        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            validation = json.loads(archive.read("validation_summary.json"))
            portable_provenance = json.loads(archive.read("provenance.json"))
            raw_manifest = json.loads(archive.read("raw_evidence/manifest.json"))

        provenance = validation["evidence_provenance"]["sources"][0]
        self.assertEqual(provenance["source_run_id"], source_id)
        self.assertEqual(provenance["application_version"], "0.1.52")
        self.assertEqual(provenance["source_commit"], "b" * 40)
        self.assertEqual(provenance["portable_exe_sha256"], "c" * 64)
        self.assertEqual(
            provenance["register"],
            {
                "filename": "approved-register.csv",
                "revision": "rev-2026-08-06",
                "sha256": "a" * 64,
                "import_id": "imp-provenance-1",
            },
        )
        self.assertEqual(provenance["capture"]["termination_reason"], "window_elapsed")
        self.assertEqual(provenance["scope"]["topic_filter"], "demo-site/#")

        self.assertEqual(report["source_run_ids"], [source_id])
        self.assertEqual(portable_provenance["schema_version"], "1.0")
        health = self.client.get("/api/v1/health")
        self.assertEqual(health.status_code, 200, health.text)
        self.assertEqual(portable_provenance["application_version"], health.json()["version"])
        self.assertEqual(portable_provenance["source_run_ids"], [source_id])
        self.assertEqual(portable_provenance["sources"][0]["source_run_id"], source_id)
        self.assertEqual(portable_provenance["sources"][0]["application_version"], "0.1.52")
        self.assertEqual(portable_provenance["sources"][0]["source_commit"], "b" * 40)
        self.assertEqual(portable_provenance["sources"][0]["portable_exe_sha256"], "c" * 64)
        self.assertEqual(
            [source["source_run_id"] for source in raw_manifest["source_runs"]],
            [source_id],
        )

        listed = self.client.get("/api/v1/reports")
        self.assertEqual(listed.status_code, 200, listed.text)
        listed_report = next(
            candidate
            for candidate in listed.json()["reports"]
            if candidate["report_id"] == report["report_id"]
        )
        self.assertEqual(listed_report["source_run_ids"], [source_id])

        manifest = RunService().get_run(report["report_id"]).result_summary["artifact_manifest"]
        self.assertEqual(manifest["schema_version"], "1.2")
        self.assertEqual(manifest["application_version"], health.json()["version"])
        self.assertEqual(manifest["source_run_ids"], [source_id])
        self.assertEqual(manifest["source_provenance"], portable_provenance["sources"])
        self.assertTrue(verify_signed_manifest(manifest))
        tampered = copy.deepcopy(manifest)
        tampered["source_provenance"][0]["source_commit"] = "d" * 40
        self.assertFalse(verify_signed_manifest(tampered))

    def test_historical_report_provenance_normalizes_valid_legacy_hashes(self) -> None:
        from types import SimpleNamespace

        from app.api.routes.reports import ReportProvenanceError, _report_provenance

        source_id = "run_historical_source"
        report_run = SimpleNamespace(
            parameters={
                "report_snapshot_v2": {
                    "renderer_version": "0.1.52",
                    "source_run_ids": [source_id],
                    "source_run_snapshots": [
                        {
                            "run_id": source_id,
                            "parameters": {
                                "application_version": "0.1.49",
                                "application_source_commit": "legacy-commit",
                                "exe_sha256": "A" * 64,
                            },
                        }
                    ],
                    "configuration_provenance": {},
                    "source_result_hashes": {},
                    "source_run_seals": {},
                }
            }
        )

        provenance = _report_provenance(report_run)
        self.assertEqual(provenance["source_run_ids"], [source_id])
        self.assertEqual(provenance["sources"][0]["portable_exe_sha256"], "a" * 64)

        report_run.parameters["report_snapshot_v2"]["source_run_snapshots"][0]["parameters"][
            "application_source_commit"
        ] = "x" * 129
        with self.assertRaisesRegex(ReportProvenanceError, "source_commit"):
            _report_provenance(report_run)

    def test_malformed_stored_provenance_returns_422_without_publishing_a_report(self) -> None:
        from app.api.routes.reports import ReportProvenanceError

        before = {
            report["report_id"] for report in self.client.get("/api/v1/reports").json()["reports"]
        }
        with mock.patch(
            "app.api.routes.reports._report_provenance",
            side_effect=ReportProvenanceError(
                "Stored report provenance field 'source_commit' exceeds 128 characters."
            ),
        ):
            response = self.client.post(
                "/api/v1/reports",
                json={
                    "project_id": "demo-project",
                    "site_id": "demo-site",
                    "report_type": "evidence_pack",
                    "output_format": "zip",
                    "source_run_ids": [],
                },
            )

        self.assertEqual(response.status_code, 422, response.text)
        self.assertIn("source_commit", response.json()["detail"])
        after = {
            report["report_id"] for report in self.client.get("/api/v1/reports").json()["reports"]
        }
        self.assertEqual(after, before)

    def test_client_report_product_is_metrics_only_and_uses_same_evidence_set(self) -> None:
        source_id = self._seed_run(
            result_summary_extra={
                "broker_capture_attempted": True,
                "capture_mode": "bounded",
                "capture_window_seconds": 3600,
                "window_completed": True,
                "termination_reason": "window_elapsed",
            }
        )
        technical = self._create_report("zip", [source_id], udmi_report_variant="technical")
        client = self._create_report("zip", [source_id], udmi_report_variant="client")
        self.assertEqual(technical["evidence_set_id"], client["evidence_set_id"])
        self.assertEqual(client["udmi_report_variant"], "client")
        default_response = self.client.post(
            "/api/v1/reports",
            json={
                "project_id": "demo-project",
                "site_id": "demo-site",
                "report_type": "udmi_validation",
                "output_format": "zip",
                "source_run_ids": [source_id],
            },
        )
        self.assertEqual(default_response.status_code, 200, default_response.text)
        self.assertEqual(default_response.json()["udmi_report_variant"], "client")

        with zipfile.ZipFile(io.BytesIO(self._download(client["report_id"]).content)) as archive:
            names = set(archive.namelist())
            client_metrics = json.loads(archive.read("client_metrics.json"))
            provenance = json.loads(archive.read("provenance.json"))
        self.assertIn("client_metrics.json", names)
        self.assertEqual(provenance["source_run_ids"], [source_id])
        self.assertEqual(provenance["sources"][0]["source_run_id"], source_id)
        self.assertNotIn("fault_details.json", names)
        self.assertNotIn("raw_evidence/records.jsonl", names)
        self.assertEqual(client_metrics["report_product"], "client_metrics")
        self.assertTrue(client_metrics["metrics"])

        for output_format in ("pdf", "docx", "xlsx"):
            with self.subTest(output_format=output_format):
                report = self._create_report(
                    output_format,
                    [source_id],
                    udmi_report_variant="client",
                )
                content = self._download(report["report_id"]).content
                if output_format == "xlsx":
                    workbook = load_workbook(io.BytesIO(content))
                    self.assertEqual(workbook.sheetnames, ["Client Metrics"])
                elif output_format == "pdf":
                    self.assertIn(b"Client Metrics", content)
                    self.assertNotIn(b"Faults in Detail", content)
                    self.assertRegex(content, rb"/Count\s+1\b")
                else:
                    with zipfile.ZipFile(io.BytesIO(content)) as archive:
                        document = archive.read("word/document.xml")
                    self.assertIn(b"Client Metrics", document)
                    self.assertNotIn(b"Faults in Detail", document)

    def test_technical_report_zip_contains_portable_redacted_raw_evidence(self) -> None:
        source_id = self._seed_run(
            result_summary_extra={
                "raw_evidence": {
                    "records": [
                        {
                            "asset_id": "asset-1",
                            "payload_type": "metadata",
                            "topic": "site/devices/asset-1/metadata",
                            "payload": {
                                "timestamp": "2026-08-05T08:00:00Z",
                                "credentials": "removed",
                            },
                            "payload_encoding": "json",
                            "payload_timestamp": "2026-08-05T08:00:00Z",
                            "broker_received_at": "2026-08-05T08:00:01+00:00",
                            "retained": False,
                            "content_sha256": "a" * 64,
                            "redaction_status": "redacted",
                        }
                    ],
                    "captured_record_count": 1,
                    "retained_bytes": 128,
                    "truncated": False,
                }
            }
        )
        report = self._create_report("zip", [source_id], udmi_report_variant="technical")

        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            manifest = json.loads(archive.read("raw_evidence/manifest.json"))
            payload_manifest = json.loads(archive.read("raw_evidence/payload_manifest.json"))
            finding_index = json.loads(archive.read("raw_evidence/finding_index.json"))
            records = [
                json.loads(line)
                for line in archive.read("raw_evidence/records.jsonl").splitlines()
                if line
            ]
            payload_bytes = archive.read(payload_manifest["entries"][0]["filename"])

        self.assertEqual(manifest["record_count"], 1)
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["source_run_id"], source_id)
        self.assertEqual(records[0]["asset_id"], "asset-1")
        self.assertEqual(manifest["evidence_id_field"], "evidence_id")
        self.assertTrue(records[0]["evidence_id"].startswith("raw-"))
        self.assertEqual(finding_index["record_id_field"], "evidence_id")
        self.assertEqual(records[0]["payload"]["credentials"], "********")
        self.assertEqual(records[0]["payload"]["timestamp"], "2026-08-05T08:00:00Z")
        self.assertEqual(manifest["payload_file_manifest"], "raw_evidence/payload_manifest.json")
        self.assertEqual(len(payload_manifest["entries"]), 1)
        entry = payload_manifest["entries"][0]
        self.assertEqual(entry["topic"], "site/devices/asset-1/metadata")
        self.assertEqual(entry["asset_id"], "asset-1")
        self.assertEqual(entry["payload_type"], "metadata")
        self.assertEqual(entry["timestamp"], "2026-08-05T08:00:00Z")
        self.assertEqual(entry["byte_count"], len(payload_bytes))
        self.assertEqual(entry["sha256"], hashlib.sha256(payload_bytes).hexdigest())
        self.assertEqual(json.loads(payload_bytes), records[0]["payload"])
        self.assertNotIn(b"removed", payload_bytes)

    def test_raw_evidence_payload_manifest_keeps_latest_topic_and_marks_non_json_metadata(self) -> None:
        source_id = self._seed_run(
            result_summary_extra={
                "raw_evidence": {
                    "records": [
                        {
                            "asset_id": "asset-1",
                            "payload_type": "state",
                            "topic": "site/asset-1/state",
                            "payload": {"version": "older"},
                            "payload_encoding": "json",
                            "payload_timestamp": "2026-08-05T09:00:00Z",
                            "broker_received_at": "2026-08-05T08:00:00Z",
                            "content_sha256": "a" * 64,
                        },
                        {
                            "asset_id": "asset-1",
                            "payload_type": "state",
                            "topic": "site/asset-1/state",
                            "payload": {"version": "latest"},
                            "payload_encoding": "json",
                            "payload_timestamp": "2026-08-05T07:00:00Z",
                            "broker_received_at": "2026-08-05T08:01:00Z",
                            "content_sha256": "b" * 64,
                        },
                        {
                            "asset_id": "asset-2",
                            "payload_type": "metadata",
                            "topic": "site/asset-2/metadata",
                            "payload": None,
                            "payload_encoding": "omitted_non_json",
                            "content_sha256": "c" * 64,
                            "redaction_status": "metadata_only_non_json",
                        },
                    ],
                    "captured_record_count": 3,
                    "retained_bytes": 128,
                    "truncated": False,
                }
            }
        )
        report = self._create_report("zip", [source_id], udmi_report_variant="technical")

        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            payload_manifest = json.loads(archive.read("raw_evidence/payload_manifest.json"))
            entries = payload_manifest["entries"]
            state = next(entry for entry in entries if entry["topic"] == "site/asset-1/state")
            metadata = next(entry for entry in entries if entry["topic"] == "site/asset-2/metadata")
            state_payload = json.loads(archive.read(state["filename"]))

        self.assertEqual(len(entries), 2)
        self.assertEqual(state["export_status"], "payload_json")
        self.assertEqual(state_payload, {"version": "latest"})
        self.assertEqual(state["broker_received_at"], "2026-08-05T08:01:00Z")
        self.assertEqual(state["payload_timestamp"], "2026-08-05T07:00:00Z")
        self.assertEqual(metadata["export_status"], "metadata_only_non_json")
        self.assertIsNone(metadata["filename"])
    def test_evidence_set_id_is_stable_when_source_runs_are_reordered(self) -> None:
        first_source = self._seed_run()
        second_source = self._seed_run()

        forward = self._create_report("zip", [first_source, second_source])
        reverse = self._create_report("zip", [second_source, first_source])

        self.assertEqual(forward["evidence_set_id"], reverse["evidence_set_id"])

    def test_cancelled_source_is_ineligible_for_field_acceptance(self) -> None:
        source_id = self._seed_run(
            status="cancelled",
            result_summary_extra={
                "broker_capture_attempted": True,
                "capture_mode": "indefinite",
                "capture_started_at": "2026-08-05T08:00:00+00:00",
                "capture_ended_at": "2026-08-05T08:01:00+00:00",
                "window_completed": False,
                "termination_reason": "cancelled",
            },
        )

        report = self._create_report("zip", [source_id])
        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            validation = json.loads(archive.read("validation_summary.json"))

        self.assertFalse(validation["scope_complete"])
        self.assertFalse(validation["acceptance_eligible"])
        self.assertEqual(validation["source_runs"][0]["termination_reason"], "cancelled")
        self.assertFalse(validation["source_runs"][0]["window_completed"])

    def test_source_run_scope_is_validated_before_report_creation(self) -> None:
        missing = self.client.post(
            "/api/v1/reports",
            json={
                "project_id": "demo-project",
                "site_id": "demo-site",
                "report_type": "udmi_validation",
                "source_run_ids": ["missing-run"],
            },
        )
        self.assertEqual(missing.status_code, 404)
        self.assertEqual(missing.json()["detail"], "Run not found.")

        other_scope = self._seed_run(project_id="other-project")
        wrong_scope = self.client.post(
            "/api/v1/reports",
            json={
                "project_id": "demo-project",
                "site_id": "demo-site",
                "report_type": "udmi_validation",
                "source_run_ids": [other_scope],
            },
        )
        self.assertEqual(wrong_scope.status_code, 404)
        self.assertEqual(wrong_scope.json()["detail"], "Run not found.")

    def test_udmi_sources_must_have_the_right_type_and_be_terminal(self) -> None:
        wrong_type = self._seed_run(job_type="bacnet_validation")
        response = self.client.post(
            "/api/v1/reports",
            json={
                "project_id": "demo-project",
                "site_id": "demo-site",
                "report_type": "udmi_validation",
                "source_run_ids": [wrong_type],
            },
        )
        self.assertEqual(response.status_code, 422)
        self.assertIn("must be a UDMI validation run", response.json()["detail"])

        for status in ("queued", "running"):
            with self.subTest(status=status):
                source_id = self._seed_run(status=status)
                try:
                    response = self.client.post(
                        "/api/v1/reports",
                        json={
                            "project_id": "demo-project",
                            "site_id": "demo-site",
                            "report_type": "udmi_validation",
                            "source_run_ids": [source_id],
                        },
                    )
                    self.assertEqual(response.status_code, 422)
                    self.assertIn("is not terminal", response.json()["detail"])
                    self.assertIn(status, response.json()["detail"])
                finally:
                    self._finish_active_test_run(source_id)

    def test_malformed_current_contract_fails_but_absent_legacy_contract_exports(self) -> None:
        malformed = {"schema_version": "1.1", "asset_metrics": {"expected": 1}}
        unknown = copy.deepcopy(_SCOPABLE_SUMMARY)
        unknown["schema_version"] = "2.0"

        for summary in (malformed, unknown):
            with self.subTest(schema_version=summary["schema_version"]):
                source_id = self._seed_run(summary=summary)
                response = self.client.post(
                    "/api/v1/reports",
                    json={
                        "project_id": "demo-project",
                        "site_id": "demo-site",
                        "report_type": "udmi_validation",
                        "output_format": "zip",
                        "source_run_ids": [source_id],
                    },
                )
                self.assertEqual(response.status_code, 422, response.text)
                self.assertIn("malformed or unsupported", response.json()["detail"])

        malformed_nested: list[tuple[str, dict]] = []
        bad_system = copy.deepcopy(_SCOPABLE_SUMMARY)
        bad_system["system_metrics"] = [None]
        malformed_nested.append(("system", bad_system))
        bad_asset = copy.deepcopy(_SCOPABLE_SUMMARY)
        bad_asset["asset_results"] = [None]
        malformed_nested.append(("asset", bad_asset))
        bad_fault = copy.deepcopy(_SCOPABLE_SUMMARY)
        bad_fault["fault_rows"] = [None]
        malformed_nested.append(("fault", bad_fault))
        bad_payload = copy.deepcopy(_SCOPABLE_SUMMARY)
        bad_payload["asset_results"][0]["payload_results"] = [
            {"payload_type": "state"}
        ]
        malformed_nested.append(("payload", bad_payload))

        for label, summary in malformed_nested:
            with self.subTest(nested=label):
                source_id = self._seed_run(summary=summary)
                response = self.client.post(
                    "/api/v1/reports",
                    json={
                        "project_id": "demo-project",
                        "site_id": "demo-site",
                        "report_type": "udmi_validation",
                        "output_format": "zip",
                        "source_run_ids": [source_id],
                    },
                )
                self.assertEqual(response.status_code, 422, response.text)
                self.assertIn("malformed", response.json()["detail"])

        legacy_source = self._seed_run(summary=None)
        legacy_report = self._create_report("zip", [legacy_source])
        with zipfile.ZipFile(io.BytesIO(self._download(legacy_report["report_id"]).content)) as archive:
            self.assertIn("validation_summary.json", archive.namelist())
            self.assertNotIn("asset_validation_schedule.json", archive.namelist())
            legacy_summary = json.loads(archive.read("validation_summary.json"))
        self.assertIn("columns", legacy_summary)
        self.assertNotIn("schema_version", legacy_summary)

    def test_filtered_scope_is_exact_recomputed_and_persisted_canonically(self) -> None:
        source_id = self._seed_run(
            summary=copy.deepcopy(_SCOPABLE_SUMMARY),
            parameters={
                "assets": [
                    {"expected_schedule": {"asset_id": "A-1", "project_site": "Site A"}},
                    {"expected_schedule": {"asset_id": "B-1", "project_site": "Site B"}},
                ]
            },
        )
        scope = {
            "schema_version": "1.0",
            "selected_payloads": [
                {
                    "source_run_id": source_id,
                    "asset_id": "A-1",
                    "payload_type": "pointset",
                },
                {
                    "source_run_id": source_id,
                    "asset_id": "A-1",
                    "payload_type": "metadata",
                },
            ],
            "unexpected_device_ids": [],
            "filters": {
                "text": "provenance label only",
                "verdict": "fail",
                "topic_contains": "a-1",
                "system": "BMS",
                "observation": "all",
                "category": "validation",
            },
        }
        report = self._create_report("zip", [source_id], udmi_scope=scope)

        from app.services.run_service import RunService

        stored = RunService().get_run(report["report_id"])
        selected = stored.parameters["udmi_scope"]["selected_payloads"]
        self.assertEqual(
            [row["payload_type"] for row in selected],
            ["metadata", "pointset"],
        )

        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            header = json.loads(archive.read("summary.json"))
            summary = json.loads(archive.read("validation_summary.json"))
            schedule = json.loads(archive.read("asset_validation_schedule.json"))
            matrix = json.loads(archive.read("fault_matrix.json"))
            details = json.loads(archive.read("fault_details.json"))
            findings = json.loads(archive.read("findings.json"))

        self.assertEqual(header["Project"], "Site A")
        self.assertEqual(header["Site"], "Site A")
        self.assertEqual(set(header), {"Project", "Site", "Report ID", "Evidence set ID", "Acceptance", "Generated"})
        self.assertEqual(
            summary["asset_metrics"],
            {
                "expected": 1,
                "observed": 1,
                "not_observed": 0,
                "with_issues": 1,
                "successfully_validated": 0,
                "unexpected": 0,
                "wrong_topic": 0,
            },
        )
        self.assertEqual(
            summary["payload_metrics"],
            {
                "expected": 2,
                "received": 1,
                "not_received": 1,
                "with_issues": 1,
                "successfully_validated": 0,
            },
        )
        self.assertEqual(summary["issue_metrics"], {"blocking": 2, "warning": 0})
        self.assertEqual(summary["filter_provenance"]["text"], "provenance label only")
        self.assertEqual([row["asset_id"] for row in schedule["rows"]], ["A-1"])
        self.assertEqual(
            [row["payload_type"] for row in schedule["rows"][0]["payload_results"]],
            ["metadata", "pointset"],
        )
        self.assertEqual(len(matrix["rows"]), 1)
        self.assertTrue(matrix["rows"][0]["missing_points"])
        self.assertTrue(matrix["rows"][0]["point_naming_issues"])
        self.assertFalse(matrix["rows"][0]["other_issues"])
        self.assertEqual(
            {row["issue_id"] for row in details["rows"]},
            {"a-metadata-missing", "a-point-name"},
        )
        self.assertEqual(
            {row["issue_id"] for row in findings},
            {row["issue_id"] for row in details["rows"]},
        )
        self.assertTrue(
            all(
                "source_run_id" not in row
                and "severity" not in row
                and "raw_evidence_uri" not in row
                for row in details["rows"]
            )
        )
        self.assertTrue(
            all(
                row["source_run_id"] == source_id
                and row["severity"]
                and row["raw_evidence_uri"]
                for row in findings
            )
        )

    def test_scoped_annotated_register_preserves_every_frozen_row_and_value(self) -> None:
        summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        summary["wrong_topic_assets"] = [
            {
                "asset_id": "A-1",
                "system": "BMS",
                "expected_topic_root": "site/floor-1/A-1",
                "actual_topic_root": "site/floor-0/A-1",
                "payloads": [
                    {
                        "payload_type": "state",
                        "expected_topic": "site/floor-1/A-1/state",
                        "actual_topic": "site/floor-0/A-1/state",
                    },
                    {
                        "payload_type": "pointset",
                        "expected_topic": "site/floor-1/A-1/pointset",
                        "actual_topic": "site/floor-0/A-1/pointset",
                    },
                ],
                "last_seen": "2026-07-23T10:02:00+00:00",
            }
        ]
        state_payload = summary["asset_results"][0]["payload_results"][0]
        state_payload["topic"] = "site/floor-1/A-1/state"
        state_payload["topics"] = [
            "site/floor-0/A-1/state",
            "site/floor-1/A-1/state",
        ]
        original_columns = [
            "Row marker",
            "Asset ID",
            "Payload type",
            "Observed in run",
            "Topic status",
            "Actual topic(s)",
            "Comment",
            "Source run ID",
            "Observed at",
        ]
        frozen_rows = [
            {
                "Row marker": "row-1",
                "Asset ID": "A-1",
                "Payload type": "state",
                "Observed in run": "source-observed-1",
                "Topic status": "source-status-1",
                "Actual topic(s)": "source-topic-1",
                "Comment": "source-comment-1",
                "Source run ID": "source-id-1",
                "Observed at": "source-time-1",
            },
            {
                "Row marker": "row-2",
                "Asset ID": "A-1",
                "Payload type": "metadata",
                "Observed in run": "source-observed-2",
                "Topic status": "source-status-2",
                "Actual topic(s)": "source-topic-2",
                "Comment": "source-comment-2",
                "Source run ID": "source-id-2",
                "Observed at": "source-time-2",
            },
            {
                "Row marker": "row-3",
                "Asset ID": "B-1",
                "Payload type": "state",
                "Observed in run": "source-observed-3",
                "Topic status": "source-status-3",
                "Actual topic(s)": "source-topic-3",
                "Comment": "source-comment-3",
                "Source run ID": "source-id-3",
                "Observed at": "source-time-3",
            },
        ]
        source_id = self._seed_run(
            summary=summary,
            parameters={
                "register_columns": original_columns,
                "register_rows": frozen_rows,
            },
        )
        report = self._create_report(
            "zip",
            [source_id],
            udmi_scope={
                "schema_version": "1.0",
                "selected_payloads": [
                    {
                        "source_run_id": source_id,
                        "asset_id": "A-1",
                        "payload_type": "state",
                    }
                ],
                "unexpected_device_ids": [],
                "filters": {},
            },
        )

        with zipfile.ZipFile(
            io.BytesIO(self._download(report["report_id"]).content)
        ) as archive:
            register = json.loads(archive.read("annotated_input_register.json"))
            wrong_topics = json.loads(archive.read("wrong_topic_assets.json"))

        self.assertEqual(register["columns"][: len(original_columns)], original_columns)
        self.assertEqual(
            register["columns"][len(original_columns) :],
            [f"{column} (report)" for column in original_columns[3:]],
        )
        self.assertEqual(
            [row["Row marker"] for row in register["rows"]],
            ["row-1", "row-2", "row-3"],
        )
        for exported, frozen in zip(register["rows"], frozen_rows, strict=True):
            self.assertEqual(
                {column: exported[column] for column in original_columns},
                frozen,
            )

        state_row, metadata_row, other_asset_row = register["rows"]
        self.assertEqual(state_row["Observed in run (report)"], "Yes")
        self.assertEqual(state_row["Topic status (report)"], "Wrong topic")
        self.assertEqual(
            state_row["Actual topic(s) (report)"],
            "site/floor-0/A-1/state, site/floor-1/A-1/state",
        )

        self.assertNotIn(
            "Not included in this report's selected payload scope.",
            state_row["Comment (report)"],
        )
        self.assertEqual(metadata_row["Observed in run (report)"], "No")
        self.assertEqual(metadata_row["Topic status (report)"], "Not evaluated")
        self.assertEqual(metadata_row["Actual topic(s) (report)"], "")
        self.assertIn(
            "Not included in this report's selected payload scope.",
            metadata_row["Comment (report)"],
        )
        self.assertIn(
            "Not included in this report's selected payload scope.",
            other_asset_row["Comment (report)"],
        )
        self.assertEqual(
            [
                payload["payload_type"]
                for payload in wrong_topics["rows"][0]["payloads"]
            ],
            ["state"],
        )

    def test_annotated_register_redacts_private_values_in_every_export_format(self) -> None:
        source_id = self._seed_run(
            parameters={
                "register_columns": ["Asset ID", "password", "private_key"],
                "register_rows": [
                    {
                        "Asset ID": "A-1",
                        "password": "fixture-secret-value",
                        "private_key": "-----BEGIN PRIVATE KEY----- fixture-key -----END PRIVATE KEY-----",
                    }
                ],
            }
        )
        for output_format in ("zip", "xlsx", "docx", "pdf"):
            with self.subTest(output_format=output_format):
                report = self._create_report(output_format, [source_id])
                content = self._download(report["report_id"]).content
                self.assertNotIn(b"fixture-secret-value", content)
                self.assertNotIn(b"fixture-key", content)
                if output_format == "zip":
                    with zipfile.ZipFile(io.BytesIO(content)) as archive:
                        register = json.loads(archive.read("annotated_input_register.json"))
                    self.assertEqual(register["rows"][0]["password"], "********")
                    self.assertEqual(register["rows"][0]["private_key"], "********")

    def test_partial_scope_keeps_topic_fault_without_failing_payload_content(self) -> None:
        summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        summary["wrong_topic_assets"] = [
            {
                "asset_id": "A-1",
                "system": "BMS",
                "expected_topic_root": "site/floor-1/A-1",
                "actual_topic_root": "site/floor-0/A-1",
                "payloads": [
                    {
                        "payload_type": "state",
                        "expected_topic": "site/floor-1/A-1/state",
                        "actual_topic": "site/floor-0/A-1/state",
                    }
                ],
                "last_seen": "2026-07-23T10:00:00+00:00",
            }
        ]
        summary["fault_rows"].append(
            {
                "issue_id": "a-state-topic-mismatch",
                "asset_id": "A-1",
                "system": "BMS",
                "payload_type": None,
                "topic_compliance_payload_type": "state",
                "category": "other_issues",
                "severity": "high",
                "description": "Registered state payload used the wrong topic.",
                "point_name": None,
                "expected_value": "site/floor-1/A-1/state",
                "observed_value": "site/floor-0/A-1/state",
                "suggested_action": "Correct the publisher topic.",
                "raw_evidence_uri": None,
            }
        )
        source_id = self._seed_run(summary=summary)
        report = self._create_report(
            "zip",
            [source_id],
            udmi_scope={
                "schema_version": "1.0",
                "selected_payloads": [
                    {
                        "source_run_id": source_id,
                        "asset_id": "A-1",
                        "payload_type": "state",
                    }
                ],
                "unexpected_device_ids": [],
                "filters": {},
            },
        )

        with zipfile.ZipFile(
            io.BytesIO(self._download(report["report_id"]).content)
        ) as archive:
            rendered = json.loads(archive.read("validation_summary.json"))
            schedule = json.loads(archive.read("asset_validation_schedule.json"))
            findings = json.loads(archive.read("findings.json"))

        self.assertEqual(rendered["asset_metrics"]["wrong_topic"], 1)
        self.assertEqual(rendered["asset_metrics"]["successfully_validated"], 0)
        self.assertEqual(rendered["issue_metrics"], {"blocking": 1, "warning": 0})
        self.assertEqual(
            [row["issue_id"] for row in findings],
            ["a-state-topic-mismatch"],
        )
        state = schedule["rows"][0]["payload_results"][0]
        self.assertEqual(state["payload_type"], "state")
        self.assertFalse(state["has_issues"])
        self.assertEqual(state["blocking_issue_count"], 0)
        self.assertTrue(state["successfully_validated"])
        self.assertEqual(schedule["rows"][0]["blocking_issue_count"], 1)
        self.assertFalse(schedule["rows"][0]["successfully_validated"])

    def test_empty_filtered_scope_is_valid_and_yields_empty_report(self) -> None:
        source_id = self._seed_run(summary=copy.deepcopy(_SCOPABLE_SUMMARY))
        report = self._create_report(
            "zip",
            [source_id],
            udmi_scope={
                "schema_version": "1.0",
                "selected_payloads": [],
                "unexpected_device_ids": [],
                "filters": {},
            },
        )
        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            header = json.loads(archive.read("summary.json"))
            summary = json.loads(archive.read("validation_summary.json"))
            schedule = json.loads(archive.read("asset_validation_schedule.json"))
            matrix = json.loads(archive.read("fault_matrix.json"))
            findings = json.loads(archive.read("findings.json"))
        self.assertEqual(header["Project"], "Not recorded")
        self.assertEqual(header["Site"], "Not recorded")
        self.assertEqual(summary["asset_metrics"]["expected"], 0)
        self.assertEqual(summary["payload_metrics"]["expected"], 0)
        self.assertEqual(summary["payload_metrics"]["not_received"], 0)
        self.assertEqual(schedule["rows"], [])
        self.assertEqual(matrix["rows"], [])
        self.assertEqual(findings, [])

    def test_full_expected_asset_scope_ignores_nonexpected_payload_evidence(self) -> None:
        summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        asset = next(row for row in summary["asset_results"] if row["asset_id"] == "A-1")
        state = next(
            row for row in asset["payload_results"] if row["payload_type"] == "state"
        )
        asset["payload_results"] = [
            state,
            {
                "payload_type": "metadata",
                "expected": False,
                "received": True,
                "has_issues": False,
                "blocking_issue_count": 0,
                "successfully_validated": True,
                "topic": "site/a-1/metadata",
                "received_at": "2026-07-23T10:01:00+00:00",
            },
        ]
        asset["expected_payloads"] = 1
        asset["received_payloads"] = 1
        asset["all_expected_payloads_received"] = True
        asset["all_received_payloads_successfully_validated"] = True
        asset["successfully_validated"] = False
        asset["issue_count"] = 1
        asset["blocking_issue_count"] = 1
        summary["asset_results"] = [asset]
        summary["fault_rows"] = [
            next(row for row in summary["fault_rows"] if row["issue_id"] == "a-asset-wide")
        ]
        summary["unexpected_devices"] = []
        summary["asset_metrics"]["unexpected"] = 0

        source_id = self._seed_run(summary=summary)
        report = self._create_report(
            "zip",
            [source_id],
            udmi_scope={
                "schema_version": "1.0",
                "selected_payloads": [
                    {
                        "source_run_id": source_id,
                        "asset_id": "A-1",
                        "payload_type": "state",
                    }
                ],
                "unexpected_device_ids": [],
                "filters": {},
            },
        )

        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            rendered = json.loads(archive.read("validation_summary.json"))
            details = json.loads(archive.read("fault_details.json"))

        self.assertEqual(rendered["asset_metrics"]["expected"], 1)
        self.assertEqual(rendered["asset_metrics"]["successfully_validated"], 0)
        self.assertEqual(rendered["payload_metrics"]["expected"], 1)
        self.assertEqual(rendered["issue_metrics"]["blocking"], 1)
        self.assertEqual(
            [row["issue_id"] for row in details["rows"]],
            ["a-asset-wide"],
        )

    def test_filtered_scope_rejects_unknown_legacy_and_ambiguous_references(self) -> None:
        source_id = self._seed_run(summary=copy.deepcopy(_SCOPABLE_SUMMARY))

        def request(scope: dict, sources: list[str] | None = None, report_type: str = "udmi_validation"):
            return self.client.post(
                "/api/v1/reports",
                json={
                    "project_id": "demo-project",
                    "site_id": "demo-site",
                    "report_type": report_type,
                    "source_run_ids": sources if sources is not None else [source_id],
                    "udmi_scope": scope,
                },
            )

        unknown = request(
            {
                "selected_payloads": [
                    {
                        "source_run_id": "another-run",
                        "asset_id": "A-1",
                        "payload_type": "state",
                    }
                ]
            }
        )
        self.assertEqual(unknown.status_code, 422)
        self.assertIn("not present", unknown.json()["detail"])

        non_expected_summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        b_asset = next(
            asset
            for asset in non_expected_summary["asset_results"]
            if asset["asset_id"] == "B-1"
        )
        b_asset["payload_results"].append(
            {
                "payload_type": "metadata",
                "expected": False,
                "received": True,
                "has_issues": False,
                "blocking_issue_count": 0,
                "successfully_validated": True,
                "topic": "site/b-1/metadata",
                "received_at": "2026-07-23T10:03:30+00:00",
            }
        )
        non_expected_id = self._seed_run(summary=non_expected_summary)
        non_expected = request(
            {
                "selected_payloads": [
                    {
                        "source_run_id": non_expected_id,
                        "asset_id": "B-1",
                        "payload_type": "metadata",
                    }
                ]
            },
            sources=[non_expected_id],
        )
        self.assertEqual(non_expected.status_code, 422)
        self.assertIn("expected payloads only", non_expected.json()["detail"])

        legacy = copy.deepcopy(_V1_SUMMARY)
        for asset in legacy["asset_results"]:
            asset.pop("payload_results", None)
        legacy_id = self._seed_run(summary=legacy)
        legacy_response = request(
            {"selected_payloads": []},
            sources=[legacy_id],
        )
        self.assertEqual(legacy_response.status_code, 422)
        self.assertIn("predates exact payload filtering", legacy_response.json()["detail"])

        no_source = request({"selected_payloads": []}, sources=[])
        self.assertEqual(no_source.status_code, 422)
        wrong_type = request(
            {"selected_payloads": []},
            sources=[source_id],
            report_type="evidence_pack",
        )
        self.assertEqual(wrong_type.status_code, 422)

        second_id = self._seed_run(summary=copy.deepcopy(_SCOPABLE_SUMMARY))
        ambiguous = request(
            {
                "selected_payloads": [],
                "unexpected_device_ids": ["rogue-1"],
            },
            sources=[source_id, second_id],
        )
        self.assertEqual(ambiguous.status_code, 422)
        self.assertIn("uniquely", ambiguous.json()["detail"])

    def test_unexpected_device_selection_only_affects_unexpected_metric(self) -> None:
        source_id = self._seed_run(summary=copy.deepcopy(_SCOPABLE_SUMMARY))
        report = self._create_report(
            "zip",
            [source_id],
            udmi_scope={
                "selected_payloads": [],
                "unexpected_device_ids": ["rogue-1"],
            },
        )
        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            summary = json.loads(archive.read("validation_summary.json"))
            schedule = json.loads(archive.read("asset_validation_schedule.json"))
            matrix = json.loads(archive.read("fault_matrix.json"))
            details = json.loads(archive.read("fault_details.json"))
        self.assertEqual(summary["asset_metrics"]["unexpected"], 1)
        self.assertEqual(summary["asset_metrics"]["expected"], 0)
        self.assertEqual(summary["payload_metrics"]["expected"], 0)
        self.assertEqual([row["id"] for row in summary["unexpected_devices"]], ["rogue-1"])
        self.assertEqual(schedule["rows"], [])
        self.assertEqual(matrix["rows"], [])
        self.assertEqual(details["rows"], [])

    def test_schema_1_payload_issue_counts_include_received_payloads_only(self) -> None:
        summary = {
            "schema_version": "1.0",
            **_metric_groups(
                assets=(1, 0, 1, 1, 0),
                payloads=(1, 0, 1, 0),
                faults=(0, 0, 0, 0, 0, 1),
                issues=(1, 0),
            ),
            "system_metrics": [
                {
                    "system": "BMS",
                    **_metric_groups(
                        assets=(1, 0, 1, 1, 0),
                        payloads=(1, 0, 1, 0),
                        faults=(0, 0, 0, 0, 0, 1),
                        issues=(1, 0),
                    ),
                }
            ],
            "asset_results": [
                {
                    "asset_id": "A-1",
                    "system": "BMS",
                    "observed": False,
                    "expected_payloads": 1,
                    "received_payloads": 0,
                    "all_expected_payloads_received": False,
                    "all_received_payloads_successfully_validated": False,
                    "successfully_validated": False,
                    "issue_count": 1,
                    "blocking_issue_count": 1,
                    "last_observed_at": None,
                    "payload_results": [
                        {
                            "payload_type": "state",
                            "expected": True,
                            "received": False,
                            "has_issues": True,
                            "blocking_issue_count": 1,
                            "successfully_validated": False,
                            "topic": "site/a-1/state",
                            "received_at": None,
                        }
                    ],
                }
            ],
            "fault_rows": [
                {
                    "issue_id": "state-not-received",
                    "asset_id": "A-1",
                    "system": "BMS",
                    "payload_type": "state",
                    "category": "other_issues",
                    "severity": "high",
                    "description": "State was not received.",
                }
            ],
        }
        source_id = self._seed_run(summary=summary)
        report = self._create_report("zip", [source_id])

        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            rendered = json.loads(archive.read("validation_summary.json"))

        self.assertEqual(
            rendered["payload_metrics"],
            {
                "expected": 1,
                "received": 0,
                "not_received": 1,
                "with_issues": 0,
                "successfully_validated": 0,
            },
        )
        self.assertEqual(rendered["system_metrics"][0]["payload_metrics"]["with_issues"], 0)
        self.assertEqual(rendered["system_metrics"][0]["payload_metrics"]["not_received"], 1)

    def test_retired_unexpected_fault_is_removed_from_old_persisted_summary(self) -> None:
        summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        summary["fault_metrics"]["other_issues"] += 1
        summary["issue_metrics"]["blocking"] += 1
        summary["fault_rows"].append(
            {
                "issue_id": "legacy-unexpected",
                "asset_id": "rogue-legacy",
                "system": "Unspecified",
                "payload_type": None,
                "category": "other_issues",
                "severity": "high",
                "description": "Legacy unexpected publisher finding.",
            }
        )
        source_id = self._seed_run(
            summary=summary,
            issues=[
                {
                    "issue_id": "legacy-unexpected",
                    "asset_id": "rogue-legacy",
                    "issue_type": "unexpected_device",
                    "severity": "high",
                    "description": "Legacy unexpected publisher finding.",
                }
            ],
        )
        report = self._create_report("zip", [source_id])

        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            rendered = json.loads(archive.read("validation_summary.json"))
            details = json.loads(archive.read("fault_details.json"))

        self.assertEqual(rendered["asset_metrics"]["unexpected"], 1)
        self.assertEqual(rendered["fault_metrics"]["other_issues"], 2)
        self.assertEqual(rendered["issue_metrics"]["blocking"], 4)
        self.assertNotIn(
            "legacy-unexpected",
            {row["issue_id"] for row in details["rows"]},
        )

    def test_report_snapshot_survives_source_mutation_and_integrity_is_stable(self) -> None:
        source_id = self._seed_run(
            summary=copy.deepcopy(_SCOPABLE_SUMMARY),
            parameters={
                "assets": [
                    {"expected_schedule": {"asset_id": "A-1", "project_site": "Site A"}},
                    {"expected_schedule": {"asset_id": "B-1", "project_site": "Site A"}},
                ]
            },
        )
        report = self._create_report("zip", [source_id])

        from app.services.reports_integrity import INTEGRITY_KEY
        from app.services.run_service import RunService

        service = RunService()
        stored_before_download = service.get_run(report["report_id"])
        self.assertIsInstance(
            stored_before_download.parameters.get("udmi_report_snapshot"),
            dict,
        )
        self.assertIn("artifact_manifest", stored_before_download.result_summary)
        self.assertIn(INTEGRITY_KEY, stored_before_download.result_summary)

        service.update_result_summary(
            source_id,
            {"validation_summary_v1": copy.deepcopy(_V1_SUMMARY)},
        )
        first = self._download(report["report_id"]).content
        with zipfile.ZipFile(io.BytesIO(first)) as archive:
            header = json.loads(archive.read("summary.json"))
            rendered = json.loads(archive.read("validation_summary.json"))
        self.assertEqual(header["Project"], "Site A")
        self.assertEqual(rendered["asset_metrics"]["expected"], 2)

        integrity_after_first = dict(
            service.get_run(report["report_id"]).result_summary[INTEGRITY_KEY]
        )
        service.update_result_summary(
            source_id,
            {"validation_summary_v1": copy.deepcopy(_SCOPABLE_SUMMARY)},
        )
        second = self._download(report["report_id"]).content
        integrity_after_second = dict(
            service.get_run(report["report_id"]).result_summary[INTEGRITY_KEY]
        )

        self.assertEqual(first, second)
        self.assertEqual(integrity_after_first, integrity_after_second)

    def test_stripped_modern_report_cannot_downgrade_to_legacy_download(self) -> None:
        from app.services.run_service import RunService

        report = self._create_report("zip", [], report_type="evidence_pack")
        service = RunService()
        from smart_commissioning_core.db.models import Run, RunResult, RunSeal
        from sqlalchemy import delete, select, update

        with service.engine.begin() as connection:
            parameters = dict(
                connection.scalar(
                    select(Run.parameters).where(Run.id == report["report_id"])
                )
            )
            parameters.pop("report_snapshot_v2", None)
            parameters.pop("report_snapshot_sha256", None)
            connection.execute(
                delete(RunResult).where(RunResult.run_id == report["report_id"])
            )
            connection.execute(
                delete(RunSeal).where(RunSeal.run_id == report["report_id"])
            )
            connection.execute(
                update(Run)
                .where(Run.id == report["report_id"])
                .values(
                    parameters=parameters,
                    result_summary={
                        "legacy_report_integrity": {
                            "classification": "missing",
                            "migration": "v0.1.26",
                            "silently_resigned": False,
                        }
                    },
                    result_sha256=None,
                    terminal_at=None,
                    owner_token=None,
                    attempt=0,
                    claimed_at=None,
                    heartbeat_at=None,
                    lease_expires_at=None,
                    state_version=0,
                )
            )

        response = self.client.get(f"/api/v1/reports/{report['report_id']}/download")
        self.assertEqual(response.status_code, 404, response.text)

    def test_modern_report_without_contract_never_enters_legacy_fallback(self) -> None:
        from app.services.run_service import RunService
        from smart_commissioning_core.db.models import ReportEvidenceContract
        from sqlalchemy import delete

        report = self._create_report("zip", [], report_type="evidence_pack")
        service = RunService()
        with service.engine.begin() as connection:
            connection.execute(
                delete(ReportEvidenceContract).where(
                    ReportEvidenceContract.run_id == report["report_id"]
                )
            )

        response = self.client.get(f"/api/v1/reports/{report['report_id']}/download")
        self.assertEqual(response.status_code, 404, response.text)

    def test_f6_classified_legacy_report_download_is_deterministic_and_read_only(self) -> None:
        from app.services.run_service import RunService
        from smart_commissioning_core.db.db_run_store import get_or_create_project_and_site
        from smart_commissioning_core.db.engine import session_factory
        from smart_commissioning_core.db.models import (
            ReportEvidenceContract,
            Run,
            RunResult,
            RunSeal,
        )
        from smart_commissioning_core.run_context import canonical_sha256
        from smart_commissioning_core.run_lifecycle import TerminalResultV1

        service = RunService()
        run_id = f"run_legacy_report_{uuid4().hex[:16]}"
        now = datetime(2026, 7, 25, 10, 0, tzinfo=UTC)
        parameters = {
            "output_format": "zip",
            "report_type": "evidence_pack",
            "source_run_ids": [],
            "report_title_custom": False,
            "report_title": "Smart Commissioning Report",
            "report_generated_at": now.isoformat(),
        }
        summary = {
            "legacy_report_integrity": {
                "classification": "missing",
                "migration": "v0.1.26",
                "silently_resigned": False,
            }
        }
        terminal = TerminalResultV1(
            status="succeeded",
            stage="report_ready",
            summary=summary,
        )
        result_sha256 = terminal.sha256()
        context_sha256 = canonical_sha256(
            {
                "schema_version": "legacy-0",
                "run_id": run_id,
                "project_id": "demo-project",
                "site_id": "demo-site",
                "job_type": "report_generation",
                "parameters": parameters,
                "execution_mode": "inline",
            }
        )
        with session_factory(service.engine).begin() as session:
            get_or_create_project_and_site(session, "demo-project", "demo-site")
            session.add(
                Run(
                    id=run_id,
                    project_id="demo-project",
                    site_id="demo-site",
                    job_type="report_generation",
                    status=terminal.status,
                    stage=terminal.stage,
                    progress_percent=100,
                    parameters=parameters,
                    result_summary=summary,
                    execution_mode="inline",
                    result_sha256=result_sha256,
                    terminal_at=now,
                    created_at=now,
                    updated_at=now,
                )
            )
            session.flush()
            session.add_all(
                [
                    ReportEvidenceContract(
                        run_id=run_id,
                        contract_version="legacy_pre_lifecycle",
                        project_id="demo-project",
                        site_id="demo-site",
                        classified_at=now,
                    ),
                    RunResult(
                        run_id=run_id,
                        schema_version=terminal.schema_version,
                        terminal_status=terminal.status,
                        terminal_stage=terminal.stage,
                        summary=summary,
                        result_payload=terminal.model_dump(mode="json"),
                        result_sha256=result_sha256,
                        created_at=now,
                    ),
                    RunSeal(
                        run_id=run_id,
                        terminal_status=terminal.status,
                        context_sha256=context_sha256,
                        result_sha256=result_sha256,
                        sealed_at=now,
                    ),
                ]
            )

        first = self._download(run_id).content
        second = self._download(run_id).content
        fresh = service.get_run(run_id)

        self.assertEqual(first, second)
        self.assertEqual(fresh.result_summary, summary)

    def test_same_asset_id_in_two_sources_does_not_cross_contaminate_scope(self) -> None:
        first_summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        second_summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        template = {
            "asset_id": "A-1",
            "system": "BMS",
            "payload_type": "state",
            "category": "other_issues",
            "severity": "low",
            "point_name": None,
            "expected_value": None,
            "observed_value": None,
            "suggested_action": "Review state evidence.",
            "raw_evidence_uri": "evidence://state-scope",
        }
        first_summary["fault_rows"].append(
            {**template, "issue_id": "first-state", "description": "First source."}
        )
        second_summary["fault_rows"].append(
            {**template, "issue_id": "second-state", "description": "Second source."}
        )
        first_id = self._seed_run(summary=first_summary)
        second_id = self._seed_run(summary=second_summary)
        report = self._create_report(
            "zip",
            [first_id, second_id],
            udmi_scope={
                "selected_payloads": [
                    {
                        "source_run_id": second_id,
                        "asset_id": "A-1",
                        "payload_type": "state",
                    }
                ]
            },
        )
        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            findings = json.loads(archive.read("findings.json"))
            summary = json.loads(archive.read("validation_summary.json"))
        self.assertEqual([row["issue_id"] for row in findings], ["second-state"])
        self.assertEqual(findings[0]["source_run_id"], second_id)
        self.assertEqual(summary["asset_metrics"]["expected"], 1)

    def test_run_wide_fault_never_creates_a_fault_matrix_asset(self) -> None:
        source_id = self._seed_run(summary=copy.deepcopy(_SCOPABLE_SUMMARY))
        report = self._create_report("zip", [source_id])
        with zipfile.ZipFile(io.BytesIO(self._download(report["report_id"]).content)) as archive:
            matrix = json.loads(archive.read("fault_matrix.json"))
            details = json.loads(archive.read("fault_details.json"))
        self.assertEqual({row["asset_id"] for row in matrix["rows"]}, {"A-1", "B-1"})
        self.assertIn(None, {row["asset_id"] for row in details["rows"]})

    def test_failed_and_cancelled_sources_are_prominent_in_every_renderer(self) -> None:
        failed = self._seed_run(status="failed")
        cancelled = self._seed_run(status="cancelled", summary=None)

        for output_format in ("zip", "pdf", "docx", "xlsx"):
            with self.subTest(output_format=output_format):
                report = self._create_report(output_format, [failed, cancelled])
                content = self._download(report["report_id"]).content
                if output_format == "zip":
                    with zipfile.ZipFile(io.BytesIO(content)) as archive:
                        report_summary = json.loads(archive.read("summary.json"))
                        validation = json.loads(archive.read("validation_summary.json"))
                    self.assertEqual(
                        set(report_summary),
                        {"Project", "Site", "Report ID", "Evidence set ID", "Acceptance", "Generated"},
                    )
                    self.assertEqual(validation["report_job_status"], "succeeded")
                    self.assertFalse(validation["scope_complete"])
                    self.assertEqual(validation["scope_status"], "incomplete")
                    self.assertEqual(
                        {row["status"] for row in validation["incomplete_source_runs"]},
                        {"failed", "cancelled"},
                    )
                elif output_format == "pdf":
                    self.assertIn(b"Validation Scope Incomplete", content)
                    self.assertIn(b"INCOMPLETE", content)
                    self.assertNotIn(b"Status: succeeded", content)
                elif output_format == "docx":
                    with zipfile.ZipFile(io.BytesIO(content)) as archive:
                        document = archive.read("word/document.xml")
                    self.assertIn(b"Validation Scope Incomplete", document)
                    self.assertIn(b"INCOMPLETE", document)
                    self.assertNotIn(b"Status: succeeded", document)
                else:
                    workbook = load_workbook(io.BytesIO(content))
                    executive = workbook["Executive Summary"]
                    metadata = {
                        executive.cell(row, 1).value: executive.cell(row, 2).value
                        for row in range(2, executive.max_row + 1)
                    }
                    self.assertIn("INCOMPLETE", metadata["Validation Scope Incomplete"])
                    self.assertNotIn("Status", metadata)
                    self.assertNotIn("Validation scope", metadata)

    def test_zip_contains_versioned_summary_schedule_and_fault_sections(self) -> None:
        source_id = self._seed_run()
        report = self._create_report("zip", [source_id])
        content = self._download(report["report_id"]).content
        self.assertNotIn(b"freshness", content.lower())

        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = set(archive.namelist())
            self.assertTrue(
                {
                    "summary.json",
                    "validation_summary.json",
                    "asset_validation_schedule.json",
                    "fault_matrix.json",
                    "fault_details.json",
                    "metric_definitions.json",
                }.issubset(names)
            )
            summary = json.loads(archive.read("validation_summary.json"))
            schedule = json.loads(archive.read("asset_validation_schedule.json"))
            matrix = json.loads(archive.read("fault_matrix.json"))
            details = json.loads(archive.read("fault_details.json"))
            definitions = json.loads(archive.read("metric_definitions.json"))

        self.assertEqual(summary["schema_version"], "1.0")
        self.assertEqual(summary["report_title"], "Site & <A> Validation")
        self.assertEqual(
            summary["asset_metrics"],
            {**_TOTALS["asset_metrics"], "unexpected": 0, "wrong_topic": 0},
        )
        self.assertEqual(
            summary["payload_metrics"],
            {**_TOTALS["payload_metrics"], "not_received": 2},
        )
        self.assertEqual(summary["overall_compliance"], "1/3 (33%)")
        self.assertEqual(summary["payloads_correct"], "4/7 (57%)")
        self.assertEqual(summary["payloads_incorrect"], "3/7 (43%)")
        self.assertEqual(summary["system_metrics"][0]["system"], "BMS")
        self.assertIn("T", summary["last_validation_run_at"])
        self.assertIn("T", summary["report_generated_at"])
        fcu = next(row for row in schedule["rows"] if row["asset_id"] == "FCU-2")
        self.assertFalse(fcu["all_expected_payloads_received"])
        self.assertFalse(fcu["successfully_validated"])
        matrix_categories = {
            category
            for row in matrix["rows"]
            for category, value in row.items()
            if category in _TOTALS["fault_metrics"] and value is True
        }
        self.assertEqual(matrix_categories, set(_TOTALS["fault_metrics"]))
        self.assertEqual({row["category"] for row in details["rows"]}, set(_TOTALS["fault_metrics"]))
        definition_by_metric = {row["metric"]: row["definition"] for row in definitions["rows"]}
        self.assertIn("divided by expected payloads", definition_by_metric["Payloads Correct %"])
        self.assertIn(
            "divided by expected payloads",
            definition_by_metric["Payloads Incorrect %"],
        )

    def test_pdf_is_landscape_complete_and_deterministic(self) -> None:
        source_id = self._seed_run()
        report = self._create_report("pdf", [source_id])
        first = self._download(report["report_id"]).content
        second = self._download(report["report_id"]).content
        self.assertEqual(first, second)
        self.assertIn(b"/MediaBox [0 0 842 595]", first)
        for text in (
            b"Site & <A> Validation",
            b"Executive Summary",
            b"Metrics by System",
            b"Asset Validation Schedule",
            b"Fault Matrix",
            b"Faults in Detail",
            b"Metric Definitions",
            b"Overall Compliance",
            b"1/3 \\(33%\\)",
            b"Payloads Correct %",
            b"4/7 \\(57%\\)",
            b"Payloads Incorrect %",
            b"3/7 \\(43%\\)",
            b"1/2 \\(50%\\)",
        ):
            self.assertIn(text, first)
        self.assertNotIn(b"freshness", first.lower())
        self.assertNotIn(b"Online", first)
        self.assertNotIn(b"Offline", first)
        self.assertLess(first.find(b"Metric Definitions"), first.find(b"Executive Summary"))
        self.assertNotIn(b"Output format:", first)
        self.assertNotIn(b"Source runs:", first)

        from app.services.report_pdf import PdfDocument

        long_title = "T" * 160
        document = PdfDocument(header_left="ELECTRACOM", header_right=long_title, landscape=True)
        document.add_paragraph("Body remains below the branding band.")
        long_header_pdf = document.render()
        self.assertNotIn(long_title.encode("ascii"), long_header_pdf)
        self.assertIn(b"\x85", long_header_pdf)

    def test_docx_is_landscape_escaped_complete_and_deterministic(self) -> None:
        source_id = self._seed_run()
        report = self._create_report("docx", [source_id])
        first = self._download(report["report_id"]).content
        second = self._download(report["report_id"]).content
        self.assertEqual(first, second)
        with zipfile.ZipFile(io.BytesIO(first)) as archive:
            document = archive.read("word/document.xml")
            header = archive.read("word/header1.xml")
        self.assertIn(b'w:orient="landscape"', document)
        self.assertIn(b'w:w="16838" w:h="11906"', document)
        self.assertIn(b"Site &amp; &lt;A&gt; Validation", document)
        self.assertIn(b"Site &amp; &lt;A&gt; Validation", header)
        self.assertIn(b"Asset Validation Schedule", document)
        self.assertIn(b"Faults in Detail", document)
        self.assertIn(b"Overall Compliance", document)
        self.assertIn(b"1/3 (33%)", document)
        self.assertIn(b"Payloads Correct %", document)
        self.assertIn(b"4/7 (57%)", document)
        self.assertIn(b"Payloads Incorrect %", document)
        self.assertIn(b"3/7 (43%)", document)
        self.assertNotIn(b"freshness", document.lower())
        self.assertLess(document.find(b"Metric Definitions"), document.find(b"Executive Summary"))
        self.assertNotIn(b">Source Run<", document)
        self.assertNotIn(b">Severity<", document)
        self.assertNotIn(b">Evidence URI<", document)
        self.assertIn(b'<w:jc w:val="center"/>', document)
        self.assertIn(b'<w:vAlign w:val="center"/>', document)
        self.assertIn(b"<w:insideH", document)
        self.assertIn(b"<w:insideV", document)
        self.assertEqual(document.count(b"<w:tr>"), document.count(b"<w:cantSplit/>"))
        self.assertIn(b"<w:keepNext/>", document)
        self.assertIn(
            b'<w:pgMar w:top="720" w:right="720" w:bottom="1080" w:left="720" '
            b'w:header="540" w:footer="540" w:gutter="0"/>',
            document,
        )

    def test_register_project_site_drives_all_udmi_report_headers(self) -> None:
        source_id = self._seed_run(
            summary=copy.deepcopy(_SCOPABLE_SUMMARY),
            parameters={
                "assets": [
                    {"expected_schedule": {"asset_id": "A-1", "project_site": "Site A"}}
                ]
            },
        )
        for output_format in ("zip", "pdf", "docx", "xlsx"):
            with self.subTest(output_format=output_format):
                report = self._create_report(output_format, [source_id])
                content = self._download(report["report_id"]).content
                if output_format == "zip":
                    with zipfile.ZipFile(io.BytesIO(content)) as archive:
                        metadata = json.loads(archive.read("summary.json"))
                    self.assertEqual(metadata["Project"], "Site A")
                    self.assertEqual(metadata["Site"], "Site A")
                    self.assertEqual(set(metadata), {"Project", "Site", "Report ID", "Evidence set ID", "Acceptance", "Generated"})
                elif output_format == "pdf":
                    self.assertIn(b"Project: Site A", content)
                    self.assertIn(b"Site: Site A", content)
                    self.assertNotIn(b"Project: demo-project", content)
                elif output_format == "docx":
                    with zipfile.ZipFile(io.BytesIO(content)) as archive:
                        document = archive.read("word/document.xml")
                    self.assertIn(b"Project: Site A", document)
                    self.assertIn(b"Site: Site A", document)
                    self.assertNotIn(b"Project: demo-project", document)
                else:
                    workbook = load_workbook(io.BytesIO(content))
                    executive = workbook["Executive Summary"]
                    metadata = {
                        executive.cell(row, 1).value: executive.cell(row, 2).value
                        for row in range(2, 7)
                    }
                    self.assertEqual(metadata["Project"], "Site A")
                    self.assertEqual(metadata["Site"], "Site A")
                    self.assertNotIn("Output format", metadata)

    def test_long_fault_detail_is_complete_wrapped_centered_and_gridded(self) -> None:
        summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        long_description = "LONG_START " + ("measured commissioning detail " * 80) + "LONG_END"
        summary["fault_rows"][1]["description"] = long_description
        source_id = self._seed_run(summary=summary)

        pdf_report = self._create_report("pdf", [source_id], title="Detail Test")
        pdf_content = self._download(pdf_report["report_id"]).content
        self.assertIn(b"LONG_START", pdf_content)
        self.assertIn(b"LONG_END", pdf_content)
        self.assertNotIn(b"\x85", pdf_content)
        self.assertNotIn(b"(Source Run)", pdf_content)
        self.assertNotIn(b"(Evidence URI)", pdf_content)

        docx_report = self._create_report("docx", [source_id], title="Detail Test")
        with zipfile.ZipFile(io.BytesIO(self._download(docx_report["report_id"]).content)) as archive:
            document = archive.read("word/document.xml")
        self.assertIn(long_description.encode("ascii"), document)

        xlsx_report = self._create_report("xlsx", [source_id], title="Detail Test")
        workbook = load_workbook(io.BytesIO(self._download(xlsx_report["report_id"]).content))
        details = workbook["Faults in Detail"]
        headers = [cell.value for cell in details[1]]
        self.assertEqual(headers, list(_expected_detail_columns()))
        description_column = headers.index("Description") + 1
        description_cell = next(
            details.cell(row, description_column)
            for row in range(2, details.max_row + 1)
            if details.cell(row, description_column).value == long_description
        )
        self.assertTrue(description_cell.alignment.wrap_text)
        self.assertEqual(description_cell.alignment.horizontal, "center")
        self.assertEqual(description_cell.alignment.vertical, "center")
        self.assertEqual(description_cell.border.left.style, "thin")

    def test_wrapped_metric_row_moves_whole_to_fresh_pdf_page(self) -> None:
        """A short row must not tear across pages merely to use spare space."""

        from app.services import report_pdf

        document = report_pdf.PdfDocument(landscape=True)
        builder = report_pdf._PageBuilder(
            page_height=document._page_height,
            margin=report_pdf._MARGIN,
            bottom_limit=report_pdf._BOTTOM_LIMIT,
            top_reserve=0.0,
        )
        # Enough room for the header and minimum row, but not this wrapped row.
        builder.y = builder.bottom_limit + 50
        document._layout_wrapped_table(
            builder,
            ("Metric", "Value"),
            (("Payloads With Issues", "inspection evidence " * 18),),
            (1.0, 1.0),
            report_pdf._BODY_SIZE,
            center_cells=False,
            draw_grid=True,
        )

        self.assertEqual(len(builder.pages), 2)
        first_page = b"\n".join(builder.pages[0])
        second_page = b"\n".join(builder.pages[1])
        self.assertNotIn(b"Payloads With Issues", first_page)
        self.assertIn(b"Payloads With Issues", second_page)
        self.assertIn(b"inspection evidence", second_page)

    def test_xlsx_print_layout_filters_styles_and_asset_verdict(self) -> None:
        source_id = self._seed_run()
        report = self._create_report("xlsx", [source_id])
        first = self._download(report["report_id"]).content
        second = self._download(report["report_id"]).content
        self.assertEqual(first, second)
        workbook = load_workbook(io.BytesIO(first))
        self.assertEqual(
            workbook.sheetnames,
            [
                "Metric Definitions",
                "Executive Summary",
                "Metrics by System",
                "Asset Validation Schedule",
                "Fault Matrix",
                "Faults in Detail",
            ],
        )
        for sheet in workbook.worksheets:
            self.assertEqual(sheet.page_setup.orientation, "landscape")
            self.assertEqual(sheet.page_setup.fitToWidth, 1)
            self.assertTrue(sheet.freeze_panes)

        assets = workbook["Asset Validation Schedule"]
        self.assertEqual(assets.freeze_panes, "A2")
        self.assertTrue(assets.auto_filter.ref)
        headers = {cell.value: cell.column for cell in assets[1]}
        fcu_row = next(
            row
            for row in range(2, assets.max_row + 1)
            if assets.cell(row, headers["Asset ID"]).value == "FCU-2"
        )
        self.assertEqual(assets.cell(fcu_row, headers["All Payloads Received"]).value, "No")
        self.assertEqual(assets.cell(fcu_row, headers["All Payloads Validated"]).value, "No")
        self.assertEqual(assets.cell(fcu_row, headers["Evidence Timestamp"]).value, "\N{EM DASH}")

        systems = workbook["Metrics by System"]
        system_headers = {cell.value: cell.column for cell in systems[1]}
        bms_row = next(
            row
            for row in range(2, systems.max_row + 1)
            if systems.cell(row, system_headers["System"]).value == "BMS"
        )
        self.assertEqual(systems.cell(bms_row, system_headers["Completion"]).value, "1/2 (50%)")
        executive = workbook["Executive Summary"]
        compliance_row = next(
            row
            for row in range(2, executive.max_row + 1)
            if executive.cell(row, 1).value == "Overall Compliance"
        )
        self.assertEqual(executive.cell(compliance_row, 2).value, "1/3 (33%)")
        supporting = {
            executive.cell(row, 1).value: executive.cell(row, 2).value
            for row in range(2, executive.max_row + 1)
        }
        self.assertEqual(supporting["Payloads Correct %"], "4/7 (57%)")
        self.assertEqual(supporting["Payloads Incorrect %"], "3/7 (43%)")
        all_values = " ".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertNotIn("freshness", all_values.casefold())
        self.assertNotIn("offline", all_values.casefold())

        from app.api.routes.reports import _completion, _payload_correctness

        self.assertEqual(_completion({"expected": 0, "successfully_validated": 0}), "N/A")
        self.assertEqual(
            _payload_correctness({"expected": 0, "successfully_validated": 0}),
            ("N/A", "N/A"),
        )

    def test_wrong_topic_and_annotated_register_render_across_report_formats(self) -> None:
        summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        summary["asset_metrics"]["wrong_topic"] = 1
        summary["wrong_topic_assets"] = [
            {
                "asset_id": "A-1",
                "system": "BMS",
                "expected_topic_root": "site/floor-1/A-1",
                "actual_topic_root": "site/floor-0/A-1",
                "payloads": [
                    {
                        "payload_type": "state",
                        "expected_topic": "site/floor-1/A-1/state",
                        "actual_topic": "site/floor-0/A-1/state",
                    }
                ],
                "last_seen": "2026-07-23T10:00:00+00:00",
            }
        ]
        summary["asset_results"][0]["payload_results"][0]["topic"] = (
            "site/floor-0/A-1/state"
        )
        source_id = self._seed_run(
            summary=summary,
            parameters={
                "register_import_id": "imp-sanitized-1",
                "register_import_filename": "commissioning-register.csv",
                "register_columns": [
                    "Project/site",
                    "System",
                    "Asset ID",
                    "Expected topic",
                    "Payload applicability",
                ],
                "register_rows": [
                    {
                        "Project/site": "Site A",
                        "System": "BMS",
                        "Asset ID": "A-1",
                        "Expected topic": "site/floor-1/A-1/#",
                        "Payload applicability": "state,metadata,pointset",
                    }
                ],
            },
        )

        xlsx = self._create_report("xlsx", [source_id], title="Topic Review")
        workbook = load_workbook(io.BytesIO(self._download(xlsx["report_id"]).content))
        self.assertIn("Wrong Topic Assets", workbook.sheetnames)
        self.assertIn("Annotated Input Register", workbook.sheetnames)
        wrong_sheet = workbook["Wrong Topic Assets"]
        wrong_headers = {cell.value: cell.column for cell in wrong_sheet[1]}
        self.assertEqual(
            wrong_sheet.cell(2, wrong_headers["Actual Topic"]).value,
            "site/floor-0/A-1/state",
        )
        register = workbook["Annotated Input Register"]
        register_headers = [cell.value for cell in register[1]]
        self.assertEqual(
            register_headers[:4],
            ["Project/site", "System", "Asset ID", "Expected topic"],
        )
        self.assertEqual(register.cell(2, register_headers.index("Observed in run") + 1).value, "Yes")
        self.assertEqual(register.cell(2, register_headers.index("Topic status") + 1).value, "Wrong topic")
        actual_topics = register.cell(
            2,
            register_headers.index("Actual topic(s)") + 1,
        ).value
        self.assertIn("site/floor-0/A-1/state", actual_topics)
        self.assertIn("site/a-1/pointset", actual_topics)

        zipped = self._create_report("zip", [source_id], title="Topic Review")
        with zipfile.ZipFile(io.BytesIO(self._download(zipped["report_id"]).content)) as archive:
            self.assertIn("wrong_topic_assets.json", archive.namelist())
            self.assertIn("annotated_input_register.json", archive.namelist())
            register_json = json.loads(archive.read("annotated_input_register.json"))
        self.assertEqual(register_json["rows"][0]["Topic status"], "Wrong topic")

        docx = self._create_report("docx", [source_id], title="Topic Review")
        with zipfile.ZipFile(io.BytesIO(self._download(docx["report_id"]).content)) as archive:
            document_xml = archive.read("word/document.xml")
        self.assertIn(b"Registered Assets on Wrong Topics", document_xml)
        self.assertIn(b"site/floor-0/A-1/state", document_xml)
        wrong_topic_heading = document_xml.rsplit(
            b"Registered Assets on Wrong Topics",
            maxsplit=1,
        )[0].rsplit(b"<w:p>", maxsplit=1)[-1]
        self.assertIn(b"<w:pageBreakBefore/>", wrong_topic_heading)

        pdf = self._create_report("pdf", [source_id], title="Topic Review")
        pdf_content = self._download(pdf["report_id"]).content
        self.assertIn(b"Registered Assets on Wrong Topics", pdf_content)
        self.assertIn(b"site/floor-0/A-1/state", pdf_content)

    def test_wrong_topic_totals_use_source_run_asset_identity_per_system(self) -> None:
        summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        summary["wrong_topic_assets"] = [
            {
                "asset_id": "A-1",
                "system": "BMS",
                "expected_topic_root": "site/floor-1/A-1",
                "actual_topic_root": "site/floor-0/A-1",
                "payloads": [
                    {
                        "payload_type": "state",
                        "expected_topic": "site/floor-1/A-1/state",
                        "actual_topic": "site/floor-0/A-1/state",
                    }
                ],
                "last_seen": "2026-07-23T10:00:00+00:00",
            }
        ]
        first_id = self._seed_run(summary=copy.deepcopy(summary))
        second_id = self._seed_run(summary=copy.deepcopy(summary))
        report = self._create_report("zip", [first_id, second_id])

        with zipfile.ZipFile(
            io.BytesIO(self._download(report["report_id"]).content)
        ) as archive:
            rendered = json.loads(archive.read("validation_summary.json"))
            wrong_topics = json.loads(archive.read("wrong_topic_assets.json"))

        self.assertEqual(rendered["asset_metrics"]["wrong_topic"], 2)
        bms = next(
            row for row in rendered["system_metrics"] if row["system"] == "BMS"
        )
        self.assertEqual(bms["asset_metrics"]["wrong_topic"], 2)
        self.assertEqual(
            {
                (row["source_run_id"], row["asset_id"])
                for row in wrong_topics["rows"]
            },
            {(first_id, "A-1"), (second_id, "A-1")},
        )

    def test_legacy_count_only_wrong_topic_metric_is_not_erased(self) -> None:
        summary = copy.deepcopy(_SCOPABLE_SUMMARY)
        summary["asset_metrics"]["wrong_topic"] = 1
        bms = next(
            row for row in summary["system_metrics"] if row["system"] == "BMS"
        )
        bms["asset_metrics"]["wrong_topic"] = 1
        summary.pop("wrong_topic_assets", None)
        source_id = self._seed_run(summary=summary)
        report = self._create_report("zip", [source_id])

        with zipfile.ZipFile(
            io.BytesIO(self._download(report["report_id"]).content)
        ) as archive:
            rendered = json.loads(archive.read("validation_summary.json"))
            members = set(archive.namelist())

        self.assertEqual(rendered["asset_metrics"]["wrong_topic"], 1)
        rendered_bms = next(
            row for row in rendered["system_metrics"] if row["system"] == "BMS"
        )
        self.assertEqual(rendered_bms["asset_metrics"]["wrong_topic"], 1)
        self.assertNotIn("wrong_topic_assets.json", members)

        full_scope = self._create_report(
            "zip",
            [source_id],
            udmi_scope={
                "schema_version": "1.0",
                "selected_payloads": [
                    {
                        "source_run_id": source_id,
                        "asset_id": asset_id,
                        "payload_type": payload_type,
                    }
                    for asset_id, payload_types in (
                        ("A-1", ("state", "metadata", "pointset")),
                        ("B-1", ("state", "pointset")),
                    )
                    for payload_type in payload_types
                ],
                "unexpected_device_ids": [],
                "filters": {},
            },
        )
        with zipfile.ZipFile(
            io.BytesIO(self._download(full_scope["report_id"]).content)
        ) as archive:
            scoped = json.loads(archive.read("validation_summary.json"))
        self.assertEqual(scoped["asset_metrics"]["wrong_topic"], 1)
        scoped_bms = next(
            row for row in scoped["system_metrics"] if row["system"] == "BMS"
        )
        self.assertEqual(scoped_bms["asset_metrics"]["wrong_topic"], 1)

        partial_scope = self._create_report(
            "zip",
            [source_id],
            udmi_scope={
                "schema_version": "1.0",
                "selected_payloads": [
                    {
                        "source_run_id": source_id,
                        "asset_id": "A-1",
                        "payload_type": "state",
                    }
                ],
                "unexpected_device_ids": [],
                "filters": {},
            },
        )
        with zipfile.ZipFile(
            io.BytesIO(self._download(partial_scope["report_id"]).content)
        ) as archive:
            partial = json.loads(archive.read("validation_summary.json"))
        self.assertEqual(partial["asset_metrics"]["wrong_topic"], 0)

    def test_udmi_xlsx_treats_untrusted_text_as_literal_cells(self) -> None:
        summary = copy.deepcopy(_V1_SUMMARY)
        summary["system_metrics"][0]["system"] = '=HYPERLINK("https://invalid/system","x")'
        summary["asset_results"][0]["system"] = '=HYPERLINK("https://invalid/system","x")'
        summary["asset_results"][0]["asset_id"] = "+cmd|' /C calc'!A0"
        summary["fault_rows"][0]["system"] = '=HYPERLINK("https://invalid/system","x")'
        summary["fault_rows"][0]["asset_id"] = "+cmd|' /C calc'!A0"
        summary["fault_rows"][0]["issue_id"] = "@SUM(1+1)"
        summary["fault_rows"][0]["description"] = "-2+3"
        summary["fault_rows"][0]["observed_value"] = (
            "observed\x01\ud800\ufffe\uffffvalue"
        )
        source_id = self._seed_run(summary=summary)
        title = '=HYPERLINK("https://invalid/title","x")'
        report = self._create_report("xlsx", [source_id], title=title)
        xlsx_content = self._download(report["report_id"]).content
        workbook = load_workbook(io.BytesIO(xlsx_content))

        executive = workbook["Executive Summary"]
        self.assertEqual(executive["A1"].value, f"'{title}")
        systems = workbook["Metrics by System"]
        assets = workbook["Asset Validation Schedule"]
        details = workbook["Faults in Detail"]
        guarded_values = [
            next(
                cell
                for row in systems.iter_rows()
                for cell in row
                if cell.value == f"'{summary['system_metrics'][0]['system']}"
            ),
            next(
                cell
                for row in assets.iter_rows()
                for cell in row
                if cell.value == f"'{summary['asset_results'][0]['asset_id']}"
            ),
            next(
                cell
                for row in details.iter_rows()
                for cell in row
                if cell.value == "'@SUM(1+1)"
            ),
            next(
                cell
                for row in details.iter_rows()
                for cell in row
                if cell.value == "'-2+3"
            ),
        ]
        self.assertTrue(all(cell.data_type == "s" for cell in guarded_values))
        detail_headers = {cell.value: cell.column for cell in details[1]}
        first_issue_row = next(
            row
            for row in range(2, details.max_row + 1)
            if details.cell(row, detail_headers["Issue ID"]).value == "'@SUM(1+1)"
        )
        self.assertEqual(
            details.cell(first_issue_row, detail_headers["Observed"]).value,
            "observed\\uD800value",
        )
        with zipfile.ZipFile(io.BytesIO(xlsx_content)) as archive:
            for name in archive.namelist():
                if not name.endswith(".xml"):
                    continue
                member = archive.read(name)
                self.assertNotIn(b"\x01", member)
                ElementTree.fromstring(member)

        docx_report = self._create_report("docx", [source_id], title="Control Character Test")
        docx_content = self._download(docx_report["report_id"]).content
        with zipfile.ZipFile(io.BytesIO(docx_content)) as archive:
            document = archive.read("word/document.xml")
        self.assertIn(b"observed\\uD800value", document)
        self.assertNotIn(b"\x01", document)
        ElementTree.fromstring(document)

        legacy_title = '=HYPERLINK("https://invalid/legacy-title","x")'
        legacy_report = self._create_report(
            "xlsx",
            [],
            title=legacy_title,
            report_type="evidence_pack",
        )
        legacy_workbook = load_workbook(
            io.BytesIO(self._download(legacy_report["report_id"]).content)
        )
        legacy_title_cell = legacy_workbook["Report Summary"]["B2"]
        self.assertEqual(legacy_title_cell.value, f"'{legacy_title}")
        self.assertEqual(legacy_title_cell.data_type, "s")


if __name__ == "__main__":
    unittest.main()
