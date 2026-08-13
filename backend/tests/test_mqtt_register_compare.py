"""MQTT discovery register-comparison tests.

Each real MQTT run binds the newest usable ``mqtt_register`` import when the
run is created. Results derive verdicts from that exact import, so later register
uploads cannot rewrite completed evidence.

HONESTY: no register means NO verdicts (never all-red); a dry/failed run that
observed nothing carries no comparison; and an expected-but-unobserved register
topic is reported as a count, never as proof a device is absent.

Boots the same inline TestClient harness as test_engines_api. Each test uses a
DISTINCT project_id because the harness database is shared process-wide.
"""

from datetime import UTC, datetime, timedelta
from io import BytesIO

from app.schemas.jobs import JobCreateRequest
from lifecycle_helpers import finish_run
from openpyxl import load_workbook
from smart_commissioning_core.db.repositories import ImportRepository
from test_engines_api import _EngineApiTestCase

# NOTE: app.api.routes.discovery is imported LAZILY inside the methods below, never
# at module top. Importing it at collection time constructs the module-level
# `service = RunService()` singleton, which caches get_engine() for whatever
# DATABASE_URL is set THEN — i.e. the default DB, before the harness points the
# process at its migrated temp database. That stale engine then poisons every
# test in the run (services read `import_records` from a DB that has no such
# table). test_engines_api imports it lazily for the same reason.

_SITE = "lab-site"


def _topic(topic: str) -> dict:
    return {"topic": topic, "last_payload": {"present_value": 1}, "message_count": 1, "attributes": {}}


class MqttRegisterCompareTests(_EngineApiTestCase):
    # -- helpers -------------------------------------------------------------

    def _new_run(self, project_id: str) -> str:
        response = self.client.post(
            "/api/v1/discovery/mqtt/runs",
            json={
                "project_id": project_id,
                "site_id": _SITE,
                "job_type": "mqtt_discovery",
                "parameters": {"dry_run": True, "broker_host": "mqtt.example.local"},
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        return response.json()["run_id"]

    def _succeeded_run(self, project_id: str, topics: list[dict]) -> str:
        """Create and finalize a real (non-dry) MQTT capture."""
        from app.api.routes import discovery as discovery_routes

        parameters: dict[str, object] = {
            "authorized": True,
            "broker_host": f"{project_id}.invalid",
        }
        discovery_routes._bind_mqtt_register(project_id, _SITE, parameters)
        run = discovery_routes.service.create_job_run(
            JobCreateRequest(
                project_id=project_id,
                site_id=_SITE,
                job_type="mqtt_discovery",
                parameters=parameters,
            ),
            expected_job_type="mqtt_discovery",
            requesting_principal="test-suite",
        )
        finish_run(
            discovery_routes.service,
            run.run_id,
            stage="capture_complete",
            summary={"topics_discovered": len(topics)},
            topics=topics,
        )
        return run.run_id

    def _seed_register(
        self,
        project_id: str,
        rows: list[dict],
        *,
        import_id: str,
        filename: str = "register.csv",
        created_at: datetime | None = None,
    ) -> None:
        from app.api.routes import discovery as discovery_routes

        ImportRepository(discovery_routes.service.engine).create(
            import_id=import_id,
            import_type="mqtt_register",
            project_id=project_id,
            site_id=_SITE,
            original_filename=filename,
            stored_file_path="",
            summary={},
            accepted_rows=rows,
            created_at=created_at,
        )

    @staticmethod
    def _by_topic(topics: list[dict]) -> dict[str, dict]:
        return {str(row.get("topic")): row for row in topics}

    @staticmethod
    def _register_rows() -> list[dict]:
        return [
            {"Asset ID": "AHU-1", "Expected topic": "demo-site/b1/ahu-1/#"},
            {"Asset ID": "FCU-2", "Expected topic": "demo-site/b1/fcu-2/state", "Payload type": ""},
        ]

    @staticmethod
    def _observed_topics() -> list[dict]:
        return [
            _topic("demo-site/b1/ahu-1/state"),
            _topic("demo-site/b1/fcu-2/metadata"),
            _topic("demo-site/rogue/x/state"),
        ]

    # -- tests ---------------------------------------------------------------

    def test_results_stamp_matched_and_unmatched(self) -> None:
        project = "reg-compare-basic"
        self._seed_register(project, self._register_rows(), import_id="imp_basic")
        run_id = self._succeeded_run(project, self._observed_topics())

        data = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        comparison = data["register_comparison"]
        self.assertTrue(comparison["register_available"])
        self.assertEqual(comparison["matched_count"], 2)
        self.assertEqual(comparison["unmatched_count"], 1)
        self.assertEqual(comparison["import_filename"], "register.csv")

        by_topic = self._by_topic(data["topics"])
        ahu = by_topic["demo-site/b1/ahu-1/state"]["attributes"]
        # The wildcard register row is cited via its wildcard (one # green-lights
        # every child), not as an individually listed topic.
        self.assertEqual(ahu["register_match"], "matched")
        self.assertEqual(ahu["register_matched_filter"], "demo-site/b1/ahu-1/#")
        self.assertEqual(ahu["register_asset_id"], "AHU-1")

        fcu = by_topic["demo-site/b1/fcu-2/metadata"]["attributes"]
        # Blank Payload type = whole asset: the metadata sibling is expected even
        # though the register row literally lists only .../state.
        self.assertEqual(fcu["register_match"], "matched")
        self.assertEqual(fcu["register_matched_filter"], "demo-site/b1/fcu-2/metadata")
        self.assertEqual(fcu["register_asset_id"], "FCU-2")

        rogue = by_topic["demo-site/rogue/x/state"]["attributes"]
        self.assertEqual(rogue["register_match"], "unmatched")

        unobserved = {entry["filter"] for entry in comparison["unobserved_filters"]}
        # FCU-2's underived siblings (state / pointset variants) were never seen.
        self.assertIn("demo-site/b1/fcu-2/state", unobserved)

    def test_no_register_means_no_verdicts(self) -> None:
        project = "reg-compare-none"
        run_id = self._succeeded_run(project, self._observed_topics())
        data = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        self.assertEqual(data["register_comparison"], {"register_available": False})
        for row in data["topics"]:
            self.assertNotIn("register_match", row.get("attributes") or {})

    def test_later_register_import_does_not_rewrite_completed_evidence(self) -> None:
        project = "reg-compare-retro"
        run_id = self._succeeded_run(project, self._observed_topics())
        first = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        self.assertEqual(first["register_comparison"], {"register_available": False})

        self._seed_register(project, self._register_rows(), import_id="imp_retro")
        second = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        self.assertEqual(first, second)

    def test_dry_run_carries_no_comparison(self) -> None:
        project = "reg-compare-dry"
        # A register exists, but a dry run sent no packets. Its sealed evidence
        # therefore stays empty and cannot carry a comparison verdict.
        self._seed_register(project, self._register_rows(), import_id="imp_dry")
        run_id = self._new_run(project)  # leaves dry_run=true in the summary
        data = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        self.assertIsNone(data["register_comparison"])
        self.assertEqual(data["topics"], [])

    def test_topics_endpoint_matches_results_stamps(self) -> None:
        project = "reg-compare-topics"
        self._seed_register(project, self._register_rows(), import_id="imp_topics")
        run_id = self._succeeded_run(project, self._observed_topics())
        data = self.client.get(f"/api/v1/discovery/runs/{run_id}/topics").json()
        self.assertTrue(data["register_comparison"]["register_available"])
        by_topic = self._by_topic(data["topics"])
        self.assertEqual(by_topic["demo-site/b1/ahu-1/state"]["attributes"]["register_match"], "matched")
        self.assertEqual(by_topic["demo-site/rogue/x/state"]["attributes"]["register_match"], "unmatched")

    def test_newest_import_with_accepted_rows_wins(self) -> None:
        project = "reg-compare-newest"
        now = datetime.now(UTC)
        # Older import HAS usable rows; newer import is fully rejected (0 rows) and
        # must be skipped, mirroring validation.py _expected_assets_from_register.
        self._seed_register(
            project,
            self._register_rows(),
            import_id="imp_old_good",
            filename="good-register.csv",
            created_at=now - timedelta(hours=1),
        )
        self._seed_register(
            project, [], import_id="imp_new_empty", filename="empty-register.csv", created_at=now
        )
        run_id = self._succeeded_run(project, self._observed_topics())
        data = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        self.assertEqual(data["register_comparison"]["import_filename"], "good-register.csv")
        by_topic = self._by_topic(data["topics"])
        self.assertEqual(by_topic["demo-site/b1/ahu-1/state"]["attributes"]["register_match"], "matched")

    def test_xlsx_export_carries_register_match_column(self) -> None:
        project = "reg-compare-xlsx"
        self._seed_register(project, self._register_rows(), import_id="imp_xlsx")
        run_id = self._succeeded_run(project, self._observed_topics())
        resp = self.client.get(f"/api/v1/discovery/runs/{run_id}/topics.xlsx")
        self.assertEqual(resp.status_code, 200, resp.text)
        rows = list(load_workbook(BytesIO(resp.content)).active.iter_rows(values_only=True))
        self.assertEqual(rows[0][-1], "Register Match")
        by_topic = {row[0]: row for row in rows[1:]}
        self.assertEqual(by_topic["demo-site/b1/ahu-1/state"][-1], "matched (demo-site/b1/ahu-1/#)")
        self.assertEqual(by_topic["demo-site/rogue/x/state"][-1], "not in register")
