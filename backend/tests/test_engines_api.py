"""API tests for the wired discovery/validation engines.

Boots the FastAPI app against a temporary SQLite database in api_key auth mode
(same shared SCT_TEST_DATABASE_URL / JOB_EXECUTION_MODE=inline pattern as
test_runs_api.py), and drives the engine routes end to end through the
TestClient.

HONESTY: there is NO real building network, BACnet device, or MQTT broker here.
The only real I/O exercised is an IP TCP-connect sweep against an ephemeral
loopback listener this test opens itself (127.0.0.1) — the honest, environment-
safe slice of the IP engine's real path. BACnet dry runs use the OFFLINE
SimulatedBacnetBackend; non-dry tests assert that live simulation is rejected or
patch the unvalidated bacpypes3 path. Validation/comparison/mqtt-config tests use
inline parameters / fakes. No assertion in this file depends on real hardware.
"""

import importlib.util
import socket
import unittest

from harness import ApiTestCase
from smart_commissioning_core.engines.bacnet_params import (
    MODE_FOREIGN_DEVICE,
    PARAM_BACNET_MODE,
    PARAM_BACNET_TARGETS,
    PARAM_BBMD_ADDRESS,
    PARAM_BBMD_PORT,
    PARAM_FD_TTL,
    TARGET_ADDRESS,
    TARGET_ASSET_ID,
    TARGET_ASSET_NAME,
    TARGET_DEVICE_INSTANCE,
    TARGET_NETWORK,
)

_API_KEY = "test-engines-api-key"

_ENV_OVERRIDES = {
    "JOB_EXECUTION_MODE": "inline",
    "AUTH_MODE": "api_key",
    "API_KEY": _API_KEY,
}

_AUTH = {"authorized": True}


class _EngineApiTestCase(ApiTestCase):
    env = _ENV_OVERRIDES
    client_headers = {"X-API-Key": _API_KEY}

    def _post(self, path: str, parameters: dict, job_type: str) -> dict:
        response = self.client.post(
            path,
            json={
                "project_id": "demo-project",
                "site_id": "demo-site",
                "job_type": job_type,
                "parameters": parameters,
            },
        )
        return response


class IpDiscoveryApiTests(_EngineApiTestCase):
    def test_inline_loopback_scan_discovers_persists_and_lists(self) -> None:
        # Open a real ephemeral loopback listener so the production default
        # asyncio connect probe finds an open port on 127.0.0.1.
        listener = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        listener.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        listener.bind(("127.0.0.1", 0))
        listener.listen(8)
        open_port = listener.getsockname()[1]
        try:
            response = self._post(
                "/api/v1/discovery/ip/runs",
                {
                    **_AUTH,
                    "cidr": "127.0.0.1/32",
                    "ports": [open_port],
                    "scan_max_concurrency": 4,
                    "scan_rate_limit_per_sec": 0,  # disable rate limiting for speed
                    "scan_connect_timeout_s": 2,
                },
                "ip_discovery",
            )
        finally:
            listener.close()

        self.assertEqual(response.status_code, 200, response.text)
        accepted = response.json()
        self.assertEqual(accepted["status"], "succeeded", "inline mode runs synchronously")
        run_id = accepted["run_id"]

        results = self.client.get(f"/api/v1/discovery/runs/{run_id}/results")
        self.assertEqual(results.status_code, 200, results.text)
        body = results.json()
        assets = body["discovered_assets"]
        self.assertEqual(len(assets), 1, "127.0.0.1 should be responsive on the open port")
        self.assertEqual(assets[0]["ip_address"], "127.0.0.1")
        self.assertIn(open_port, [p["port"] for p in assets[0]["observed_ports"]])
        # Structured rows were persisted via DiscoveryRepository.
        self.assertEqual(len(body["devices"]), 1)
        self.assertEqual(body["devices"][0]["device_type"], "ip_host")
        self.assertIn(open_port, body["devices"][0]["attributes"]["open_ports"])

    def test_dry_run_returns_plan_without_scanning(self) -> None:
        response = self._post(
            "/api/v1/discovery/ip/runs",
            {"cidr": "10.99.0.0/30", "ports": [80, 443], "dry_run": True},
            "ip_discovery",
        )
        self.assertEqual(response.status_code, 200, response.text)
        run_id = response.json()["run_id"]

        results = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        summary = results["result_summary"]
        self.assertTrue(summary["dry_run"])
        plan = summary["dry_run_plan"]
        self.assertEqual(plan["engine"], "ip_discovery")
        self.assertEqual(plan["target_count"], 4)  # 2 hosts x 2 ports
        self.assertEqual(summary["hosts_responsive"], 0)
        self.assertEqual(results["devices"], [], "dry run persists no devices")

    def test_unauthorized_real_scan_rejected_with_403(self) -> None:
        # No authorization and not a dry run => boundary 403 before any scan.
        response = self._post(
            "/api/v1/discovery/ip/runs",
            {"cidr": "10.0.0.0/30", "ports": [80]},
            "ip_discovery",
        )
        self.assertEqual(response.status_code, 403, response.text)
        self.assertIn("authoriz", response.json()["detail"].lower())

    def test_authorized_run_stamps_real_authorizer(self) -> None:
        # scan_authorization.authorized_by must name the REAL authenticated
        # principal (the shared-key admin in these tests), never a client-supplied
        # label; an operator-supplied note is preserved.
        response = self._post(
            "/api/v1/discovery/ip/runs",
            {
                **_AUTH,
                "cidr": "127.0.0.1/32",
                "ports": [9],  # discard port, closed -> connect refused fast
                "scan_connect_timeout_s": 1,
                "scan_rate_limit_per_sec": 0,
                "scan_authorization": {"authorized_by": "frontend-operator", "note": "floor 3 sign-off"},
            },
            "ip_discovery",
        )
        self.assertEqual(response.status_code, 200, response.text)
        run_id = response.json()["run_id"]
        record = self.client.get(f"/api/v1/discovery/runs/{run_id}").json()
        authz = record["parameters"]["scan_authorization"]
        self.assertTrue(authz["authorized"])
        self.assertEqual(authz["authorized_by"], "shared-key")
        self.assertEqual(authz["note"], "floor 3 sign-off")  # operator note preserved

    def test_dry_run_adds_no_scan_authorization(self) -> None:
        # A dry run needs no authorization, so the route must NOT stamp one.
        run_id = self._post(
            "/api/v1/discovery/ip/runs",
            {"cidr": "10.99.0.0/30", "ports": [80], "dry_run": True},
            "ip_discovery",
        ).json()["run_id"]
        record = self.client.get(f"/api/v1/discovery/runs/{run_id}").json()
        self.assertNotIn("scan_authorization", record["parameters"])

    def test_no_cidr_falls_back_to_imported_register_addresses(self) -> None:
        # Upload an IP register with NO hostname column + a duplicate address,
        # then run discovery with no cidr/range: the route must scan the register's
        # deduped Expected IP addresses (not fail) — the core bug this fixes.
        csv = (
            b"Project/site,System,Asset ID,Asset name,Expected IP address,Expected services/ports\n"
            b"M,ACS,A1,Cam,10.10.100.230,80/tcp\n"
            b"M,ACS,A2,Door,10.10.100.230,80/tcp\n"  # dup address -> deduped
            b"M,ACS,A3,NVR,10.10.100.74,80/tcp\n"
        )
        up = self.client.post(
            "/api/v1/imports",
            data={"import_type": "ip_register", "project_id": "demo-project", "site_id": "demo-site"},
            files={"file": ("reg.csv", csv, "text/csv")},
        )
        self.assertEqual((up.status_code, up.json()["accepted_rows"], up.json()["rejected_rows"]), (200, 3, 0), up.text)

        run = self._post(
            "/api/v1/discovery/ip/runs",
            {"authorized": True, "ports": [9], "scan_connect_timeout_s": 1, "scan_rate_limit_per_sec": 0},
            "ip_discovery",
        )
        self.assertEqual(run.status_code, 200, run.text)
        self.assertEqual(run.json()["status"], "succeeded")
        record = self.client.get(f"/api/v1/discovery/runs/{run.json()['run_id']}").json()
        self.assertEqual(sorted(record["parameters"]["addresses"]), ["10.10.100.230", "10.10.100.74"])

    def test_no_targets_and_no_register_rejected_with_400(self) -> None:
        response = self.client.post(
            "/api/v1/discovery/ip/runs",
            json={"project_id": "empty-p", "site_id": "empty-s", "job_type": "ip_discovery",
                  "parameters": {"authorized": True}},
        )
        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("scan target", response.json()["detail"].lower())


class BacnetDiscoveryApiTests(_EngineApiTestCase):
    def test_non_dry_run_rejects_simulated_backend(self) -> None:
        response = self._post(
            "/api/v1/discovery/bacnet/runs",
            {**_AUTH, "bacnet_backend": "simulated"},
            "bacnet_discovery",
        )
        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("only available for dry runs", response.json()["detail"])

    def test_unknown_bacnet_backend_returns_400(self) -> None:
        response = self._post(
            "/api/v1/discovery/bacnet/runs",
            {**_AUTH, "bacnet_backend": "not-a-backend"},
            "bacnet_discovery",
        )
        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("Unsupported BACnet backend", response.json()["detail"])

    def test_bacnet_dry_run_emits_no_broadcast(self) -> None:
        response = self._post(
            "/api/v1/discovery/bacnet/runs",
            {
                **_AUTH,
                "dry_run": True,
                "device_instance_low": 1000,
                "device_instance_high": 2000,
            },
            "bacnet_discovery",
        )
        self.assertEqual(response.status_code, 200, response.text)
        run_id = response.json()["run_id"]
        results = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        self.assertIn("dry_run_plan", results["result_summary"])
        self.assertEqual(results["result_summary"]["backend"], "simulated")
        self.assertEqual(results["devices"], [])

    def test_authorized_run_defaults_bacnet_backend_to_bacpypes3(self) -> None:
        # HONESTY: an authorized real run (no dry_run, no explicit override) is
        # persisted with bacnet_backend=bacpypes3 so BOTH the inline and worker
        # paths attempt REAL discovery. Deterministic — independent of whether
        # bacpypes3 is installed (the parameter is stamped before the run executes;
        # with an Auto source interface the run has no local_address so no socket
        # I/O occurs even when the dependency is present).
        response = self._post(
            "/api/v1/discovery/bacnet/runs",
            {**_AUTH},
            "bacnet_discovery",
        )
        self.assertEqual(response.status_code, 200, response.text)
        run_id = response.json()["run_id"]
        record = self.client.get(f"/api/v1/discovery/runs/{run_id}").json()
        self.assertEqual(record["parameters"]["bacnet_backend"], "bacpypes3")

    @unittest.skipIf(
        importlib.util.find_spec("bacpypes3") is not None,
        "bacpypes3 is installed; the honest-failure (missing-dep) path is unreachable",
    )
    def test_authorized_run_without_bacpypes3_fails_not_simulated(self) -> None:
        # The exe-today case: an authorized real run selects bacpypes3, which is
        # not installed, so the inline run must terminate FAILED and expose NO
        # simulated devices/points — never the "Acme Controls"/"Globex BMS" fakes.
        # local_address is explicit so this reaches the missing-dependency failure
        # instead of stopping first at the Source Interface guard.
        response = self._post(
            "/api/v1/discovery/bacnet/runs",
            {**_AUTH, "local_address": "192.0.2.10/24"},
            "bacnet_discovery",
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["status"], "failed")
        run_id = response.json()["run_id"]

        results = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        self.assertEqual(results["status"], "failed")
        self.assertNotIn("Source Interface", str(results.get("error_message")))
        self.assertEqual(results["devices"], [])
        self.assertEqual(results["points"], [])
        self.assertEqual(results["discovered_assets"], [])
        # The simulated backend label must never appear on a real authorized run.
        self.assertNotEqual(results["result_summary"].get("backend"), "simulated")
        blob = str(results).lower()
        self.assertNotIn("acme controls", blob)
        self.assertNotIn("globex", blob)


class BacnetTransportPlumbingApiTests(_EngineApiTestCase):
    """Server-side injection of the saved BACnet transport config + register targets.

    This IS the feature: the frontend posts ``{authorized: true}`` and nothing
    else for a BACnet run, so a key not injected here never reaches the engine at
    all — which is precisely how the saved BBMD / Foreign Device fields came to be
    validated, stored, and then silently ignored by every scan.

    The parameter keys are IMPORTED from the shared contract, never spelled as
    literals. The engine suite imports the same names, so a key renamed on one
    side of the seam breaks both suites instead of passing both and failing only
    against a real BBMD on the lab network.

    Every run here is a DRY RUN. The route resolves transport and targets
    identically for dry and real runs (deliberately — the plan must echo the
    transport a real scan would use, so it can be checked before a packet is
    sent), and a dry run performs no I/O, so nothing here depends on bacpypes3
    being installed or on a network being reachable.

    Each test uses its OWN project/site: the harness database is process-wide and
    is never reset between tests, so a config snapshot or register import written
    under a shared id would leak into unrelated tests and make these pass or fail
    on execution order.
    """

    def _save_bacnet_config(self, project_id: str, values: dict, site_id: str = "lab-site") -> None:
        """Persist a BACnet config snapshot straight through the service.

        Bypassing PUT /configuration's validation is not a shortcut here — it is
        the real case. A snapshot is validated only when it is saved and is never
        re-validated or re-defaulted on load, so a machine configured by an
        EARLIER release holds whatever that release allowed. It is also the only
        way to build the invalid-BBMD-Address case at all, now that validate()
        rejects that at the field.
        """
        from app.api.routes import discovery as discovery_routes

        config_service = discovery_routes.config_service
        snapshot = config_service.load(project_id, site_id, mask_secrets=False)
        snapshot.bacnet.values.update(values)
        config_service.save(snapshot, project_id=project_id, site_id=site_id)

    def _dry_run(self, project_id: str, site_id: str = "lab-site", parameters: dict | None = None):
        return self.client.post(
            "/api/v1/discovery/bacnet/runs",
            json={
                "project_id": project_id,
                "site_id": site_id,
                "job_type": "bacnet_discovery",
                "parameters": {"dry_run": True, **(parameters or {})},
            },
        )

    def _persisted_parameters(self, response) -> dict:
        """The parameters as PERSISTED on the run record — what the worker reads.

        Asserting on the stored record rather than the request dict is the point:
        the hosted worker path never sees the route's local dict, only this.
        """
        self.assertEqual(response.status_code, 200, response.text)
        record = self.client.get(f"/api/v1/discovery/runs/{response.json()['run_id']}")
        self.assertEqual(record.status_code, 200, record.text)
        return record.json()["parameters"]

    def _run_count(self) -> int:
        return len(self.client.get("/api/v1/discovery/runs").json()["runs"])

    # -- transport: Foreign Device enabled ---------------------------------

    def test_foreign_device_enabled_persists_typed_transport_parameters(self) -> None:
        self._save_bacnet_config(
            "fd-on",
            {
                "Foreign Device": "Enabled",
                "BBMD Address": "10.20.30.40",
                "BBMD UDP Port": "47809",
                "TTL": "120",
            },
        )
        parameters = self._persisted_parameters(self._dry_run("fd-on"))
        self.assertEqual(parameters[PARAM_BACNET_MODE], MODE_FOREIGN_DEVICE)
        self.assertEqual(parameters[PARAM_BBMD_ADDRESS], "10.20.30.40")
        # TYPED, not the config's strings: the config stores every value as text,
        # and the engine needs ints on the far side of the Dramatiq round-trip.
        self.assertEqual(parameters[PARAM_BBMD_PORT], 47809)
        self.assertEqual(parameters[PARAM_FD_TTL], 120)

    def test_foreign_device_enabled_soft_defaults_junk_port_and_ttl(self) -> None:
        # Port/TTL were only ever validated on save, so an old snapshot can hold
        # junk in them. Neither is worth blocking a lab scan over — unlike the
        # BBMD Address, which is load-bearing and fails loud below.
        self._save_bacnet_config(
            "fd-junk",
            {
                "Foreign Device": "Enabled",
                "BBMD Address": "10.20.30.40",
                "BBMD UDP Port": "not-a-port",
                "TTL": "",
            },
        )
        parameters = self._persisted_parameters(self._dry_run("fd-junk"))
        self.assertEqual(parameters[PARAM_BACNET_MODE], MODE_FOREIGN_DEVICE)
        self.assertEqual(parameters[PARAM_BBMD_PORT], 47808)
        self.assertEqual(parameters[PARAM_FD_TTL], 300)

    def test_run_parameters_override_saved_transport_config(self) -> None:
        # setdefault semantics, consistent with source_ip / qos on these routes.
        self._save_bacnet_config(
            "fd-override",
            {"Foreign Device": "Enabled", "BBMD Address": "10.20.30.40", "TTL": "300"},
        )
        parameters = self._persisted_parameters(
            self._dry_run(
                "fd-override",
                parameters={PARAM_BBMD_ADDRESS: "192.0.2.99", PARAM_BBMD_PORT: 47810, PARAM_FD_TTL: 60},
            )
        )
        self.assertEqual(parameters[PARAM_BBMD_ADDRESS], "192.0.2.99")
        self.assertEqual(parameters[PARAM_BBMD_PORT], 47810)
        self.assertEqual(parameters[PARAM_FD_TTL], 60)
        # Keys the operator did NOT override still come from the config.
        self.assertEqual(parameters[PARAM_BACNET_MODE], MODE_FOREIGN_DEVICE)

    # -- transport: the zero-regression path -------------------------------

    def test_default_install_injects_no_transport_parameters(self) -> None:
        # A project with no saved config gets the seeded defaults. Nothing may be
        # injected: with nothing configured, discovery must behave EXACTLY as it
        # does today (local broadcast), and the seeded BBMD Address is fictional
        # demo data that must never reach a run.
        parameters = self._persisted_parameters(self._dry_run("fd-default"))
        for key in (PARAM_BACNET_MODE, PARAM_BBMD_ADDRESS, PARAM_BBMD_PORT, PARAM_FD_TTL):
            self.assertNotIn(key, parameters)
        self.assertNotIn("10.10.25.20", str(parameters))

    def test_bbmd_enabled_alone_never_triggers_foreign_device_registration(self) -> None:
        # THE trigger-discipline guard. "BBMD" is a different, informational
        # setting that happened to be seeded Enabled alongside a fictional
        # address; gating on it (or on "a BBMD Address is present") would make
        # every default install register against a host that does not exist.
        # Only "Foreign Device" == Enabled counts.
        self._save_bacnet_config(
            "fd-off",
            {"BBMD": "Enabled", "BBMD Address": "10.10.25.20", "Foreign Device": "Disabled"},
        )
        parameters = self._persisted_parameters(self._dry_run("fd-off"))
        for key in (PARAM_BACNET_MODE, PARAM_BBMD_ADDRESS, PARAM_BBMD_PORT, PARAM_FD_TTL):
            self.assertNotIn(key, parameters)

    # -- transport: fail loud, fail early ----------------------------------

    def test_invalid_bbmd_address_returns_400_and_creates_no_run(self) -> None:
        # Reachable from a snapshot saved before validate() checked this field —
        # i.e. from the machine this release is being shipped to. The run must
        # never be created and must never quietly fall back to broadcast: that
        # would report a clean empty local scan for a scan the operator asked to
        # send through a BBMD, which is the exact bug being fixed.
        self._save_bacnet_config(
            "fd-garbage", {"Foreign Device": "Enabled", "BBMD Address": "not-an-ip"}
        )
        before = self._run_count()
        response = self._dry_run("fd-garbage")
        self.assertEqual(response.status_code, 400, response.text)
        detail = response.json()["detail"]
        self.assertIn("BBMD Address", detail)
        self.assertIn("not-an-ip", detail)
        self.assertIn("Configuration page", detail)  # actionable: names where to fix it
        self.assertEqual(self._run_count(), before, "a rejected request must leave no orphaned run")

    def test_blank_bbmd_address_returns_400_and_creates_no_run(self) -> None:
        self._save_bacnet_config(
            "fd-blank", {"Foreign Device": "Enabled", "BBMD Address": "   "}
        )
        before = self._run_count()
        response = self._dry_run("fd-blank")
        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("BBMD Address is empty", response.json()["detail"])
        self.assertEqual(self._run_count(), before, "a rejected request must leave no orphaned run")

    # -- targeting: the bacnet_register ------------------------------------

    def _import_register(self, project_id: str, csv: bytes, *, expect_accepted: int) -> None:
        response = self.client.post(
            "/api/v1/imports",
            data={"import_type": "bacnet_register", "project_id": project_id, "site_id": "lab-site"},
            files={"file": ("bacnet-register.csv", csv, "text/csv")},
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["accepted_rows"], expect_accepted, response.text)

    def test_register_import_becomes_deduped_bacnet_targets(self) -> None:
        # A1 and A3 are distinct rows to the IMPORT (its duplicate key is
        # Asset ID + instance) but name the SAME device — same address, same
        # instance. The route must probe it once.
        self._import_register(
            "reg-dedupe",
            b"Project/site,System,Asset ID,Asset name,BACnet device instance,BACnet network,IP address\n"
            b"M,HVAC,A1,AHU-1,101,1,10.10.100.10\n"
            b"M,HVAC,A2,AHU-2,102,1,10.10.100.11\n"
            b"M,HVAC,A3,AHU-1 duplicate entry,101,1,10.10.100.10\n",
            expect_accepted=3,
        )
        parameters = self._persisted_parameters(self._dry_run("reg-dedupe"))
        targets = parameters[PARAM_BACNET_TARGETS]
        self.assertEqual(
            [(t[TARGET_ADDRESS], t[TARGET_DEVICE_INSTANCE]) for t in targets],
            [("10.10.100.10", 101), ("10.10.100.11", 102)],
            "deduped on (address, device_instance), first-seen (register) order",
        )
        # Rich rows, not bare addresses: the register identity is what lets the
        # run report WHICH expected device stayed silent.
        self.assertEqual(targets[0][TARGET_ASSET_ID], "A1")
        self.assertEqual(targets[0][TARGET_ASSET_NAME], "AHU-1")
        # Typed on the far side of the Dramatiq round-trip, like the transport ints.
        self.assertEqual(targets[0][TARGET_NETWORK], 1)
        self.assertIsInstance(targets[0][TARGET_DEVICE_INSTANCE], int)

    def test_no_register_still_creates_a_broadcast_only_run(self) -> None:
        # Deliberately UNLIKE the IP route, which 400s with no targets because an
        # IP sweep then has nothing to do at all. Broadcast-only BACnet discovery
        # is a legitimate scan and must not be blocked for want of a register.
        response = self._dry_run("reg-absent")
        self.assertEqual(response.status_code, 200, response.text)
        self.assertNotIn(PARAM_BACNET_TARGETS, self._persisted_parameters(response))

    def test_operator_supplied_targets_win_over_the_register(self) -> None:
        self._import_register(
            "reg-override",
            b"Project/site,System,Asset ID,Asset name,BACnet device instance,BACnet network,IP address\n"
            b"M,HVAC,A1,AHU-1,101,1,10.10.100.10\n",
            expect_accepted=1,
        )
        chosen = [{TARGET_ADDRESS: "192.0.2.7", TARGET_DEVICE_INSTANCE: 999}]
        parameters = self._persisted_parameters(
            self._dry_run("reg-override", parameters={PARAM_BACNET_TARGETS: chosen})
        )
        self.assertEqual(parameters[PARAM_BACNET_TARGETS], chosen)

    def test_legacy_register_rows_are_skipped_not_fatal(self) -> None:
        # A register imported before the numeric/IP row validators existed can
        # hold unusable rows. They must cost the operator those rows, never the
        # whole scan: a 500 here would read to the field as "discovery is broken"
        # with nothing naming the register as the cause. Written through the
        # repository because the import route would (correctly) reject them now.
        from app.api.routes import discovery as discovery_routes
        from smart_commissioning_core.db.repositories import ImportRepository

        ImportRepository(discovery_routes.service.engine).create(
            import_id="imp_legacy_bacnet_register",
            import_type="bacnet_register",
            project_id="reg-legacy",
            site_id="lab-site",
            original_filename="legacy-register.csv",
            stored_file_path="",
            summary={},
            accepted_rows=[
                {
                    "IP address": "10.10.100.10",
                    "BACnet device instance": "101",
                    "BACnet network": "1",
                    "Asset ID": "A1",
                    "Asset name": "AHU-1",
                },
                # Unusable instance: could never be matched to a discovered
                # device, so it could never be reported as expected-but-silent.
                {"IP address": "10.10.100.11", "BACnet device instance": "not-a-number"},
                {"IP address": "", "BACnet device instance": "103"},  # nothing to probe
                {"BACnet device instance": "104"},  # no address at all
                "not-even-a-row",  # corrupt record
            ],
        )
        response = self._dry_run("reg-legacy")
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(
            self._persisted_parameters(response)[PARAM_BACNET_TARGETS],
            [
                {
                    TARGET_ADDRESS: "10.10.100.10",
                    TARGET_DEVICE_INSTANCE: 101,
                    TARGET_ASSET_ID: "A1",
                    TARGET_ASSET_NAME: "AHU-1",
                    TARGET_NETWORK: 1,
                }
            ],
        )


class MqttDiscoveryApiTests(_EngineApiTestCase):
    def test_mqtt_dry_run_describes_plan_without_connecting(self) -> None:
        response = self._post(
            "/api/v1/discovery/mqtt/runs",
            {"topic_prefix": "udmi", "dry_run": True, "broker_host": "mqtt.example.local"},
            "mqtt_discovery",
        )
        self.assertEqual(response.status_code, 200, response.text)
        run_id = response.json()["run_id"]
        results = self.client.get(f"/api/v1/discovery/runs/{run_id}/results").json()
        plan = results["result_summary"]["dry_run_plan"]
        self.assertEqual(plan["engine"], "mqtt_discovery")
        self.assertIn("udmi/#", plan["targets"])
        self.assertEqual(plan["broker_host"], "mqtt.example.local")
        # Credentials must never appear in the plan.
        self.assertNotIn("password", str(plan).lower())

    def test_mqtt_unauthorized_real_capture_rejected(self) -> None:
        response = self._post(
            "/api/v1/discovery/mqtt/runs",
            {"topic_prefix": "udmi", "broker_host": "mqtt.example.local"},
            "mqtt_discovery",
        )
        self.assertEqual(response.status_code, 403, response.text)

    def test_topics_view_endpoint_available(self) -> None:
        # A dry run produces no topics, but the endpoint must respond cleanly.
        run_id = self._post(
            "/api/v1/discovery/mqtt/runs",
            {"dry_run": True, "broker_host": "mqtt.example.local"},
            "mqtt_discovery",
        ).json()["run_id"]
        topics = self.client.get(f"/api/v1/discovery/runs/{run_id}/topics")
        self.assertEqual(topics.status_code, 200, topics.text)
        self.assertEqual(topics.json()["topics"], [])

    def _mqtt_dry_run_id(self) -> str:
        return self._post(
            "/api/v1/discovery/mqtt/runs",
            {"dry_run": True, "broker_host": "mqtt.example.local"},
            "mqtt_discovery",
        ).json()["run_id"]

    def test_topics_xlsx_export_empty_for_dry_run(self) -> None:
        # mq9nhbzu Excel export: a dry run has no topics, so the workbook must be
        # header-only (no fabricated rows) and carry the xlsx download headers.
        from io import BytesIO

        from openpyxl import load_workbook

        run_id = self._mqtt_dry_run_id()
        resp = self.client.get(f"/api/v1/discovery/runs/{run_id}/topics.xlsx")
        self.assertEqual(resp.status_code, 200, resp.text)
        self.assertEqual(
            resp.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", resp.headers["content-disposition"])
        self.assertIn(".xlsx", resp.headers["content-disposition"])
        sheet = load_workbook(BytesIO(resp.content)).active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(
            rows[0],
            ("Topic", "Asset", "Last Seen", "Message Count", "Latest Payload", "Register Match"),
        )
        self.assertEqual(len(rows), 1, "empty capture stays empty — no fabricated data rows")

    def test_topics_xlsx_export_includes_persisted_rows(self) -> None:
        from io import BytesIO

        from app.api.routes import discovery as discovery_routes
        from openpyxl import load_workbook

        run_id = self._mqtt_dry_run_id()
        discovery_routes._discovery_repository().replace_topics(
            run_id,
            [
                {
                    "topic": "demo-site/b1/ahu-1/state",
                    "last_payload": {"present_value": 22},
                    "message_count": 3,
                    "attributes": {"device_ref": "AHU-1"},
                }
            ],
        )
        resp = self.client.get(f"/api/v1/discovery/runs/{run_id}/topics.xlsx")
        self.assertEqual(resp.status_code, 200, resp.text)
        rows = list(load_workbook(BytesIO(resp.content)).active.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)
        data = rows[1]
        self.assertEqual(data[0], "demo-site/b1/ahu-1/state")
        self.assertEqual(data[1], "AHU-1")
        self.assertEqual(data[3], "3")
        self.assertIn("present_value", data[4])

    def test_topics_xlsx_topic_filter_narrows_rows(self) -> None:
        from io import BytesIO

        from app.api.routes import discovery as discovery_routes
        from openpyxl import load_workbook

        run_id = self._mqtt_dry_run_id()
        discovery_routes._discovery_repository().replace_topics(
            run_id,
            [
                {"topic": "demo-site/b1/ahu-1/state", "last_payload": {"a": 1}, "message_count": 1, "attributes": {}},
                {
                    "topic": "demo-site/b1/ahu-1/events/pointset",
                    "last_payload": {"b": 2},
                    "message_count": 1,
                    "attributes": {},
                },
            ],
        )
        resp = self.client.get(
            f"/api/v1/discovery/runs/{run_id}/topics.xlsx",
            params={"topic_filter": "demo-site/+/+/state"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        topics = [row[0] for row in load_workbook(BytesIO(resp.content)).active.iter_rows(min_row=2, values_only=True)]
        self.assertEqual(topics, ["demo-site/b1/ahu-1/state"])

    def test_topics_xlsx_404_for_unknown_run(self) -> None:
        resp = self.client.get("/api/v1/discovery/runs/run_does_not_exist/topics.xlsx")
        self.assertEqual(resp.status_code, 404, resp.text)

    def test_run_parameter_secrets_redacted_to_client_but_real_internally(self) -> None:
        # A broker password / inline private key passed as run parameters must be
        # redacted in the API response (any viewer) but kept server-side so
        # execution + rollback still read the real values.
        run_id = self._post(
            "/api/v1/discovery/mqtt/runs",
            {
                "dry_run": True,
                "broker_host": "mqtt.example.local",
                "password": "hunter2",
                "private_key": "-----BEGIN PRIVATE KEY-----\nabc\n-----END PRIVATE KEY-----",
            },
            "mqtt_discovery",
        ).json()["run_id"]
        got = self.client.get(f"/api/v1/discovery/runs/{run_id}")
        self.assertEqual(got.status_code, 200, got.text)
        params = got.json()["parameters"]
        self.assertEqual(params["password"], "********")
        self.assertEqual(params["private_key"], "********")
        self.assertEqual(params["broker_host"], "mqtt.example.local")  # non-secret untouched
        self.assertNotIn("hunter2", got.text)
        # Server-side attribute access (rollback/execution) keeps the real value.
        from app.api.routes import discovery as discovery_routes

        self.assertEqual(discovery_routes.service.get_run(run_id).parameters["password"], "hunter2")


class PointValidationApiTests(_EngineApiTestCase):
    def test_inline_point_validation_flags_mismatch(self) -> None:
        response = self._post(
            "/api/v1/validation/bacnet/runs",
            {
                "expected_points": [
                    {
                        "Expected point name": "SupplyAirTemp",
                        "Expected value": "20",
                        "Expected value type": "number",
                        "Required/optional flag": "required",
                    },
                    {
                        "Expected point name": "MissingPoint",
                        "Required/optional flag": "required",
                    },
                ],
                "observed_points": [
                    {"point_name": "SupplyAirTemp", "observed_value": {"value": 25.0}},
                ],
            },
            "bacnet_validation",
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["status"], "succeeded")
        run_id = response.json()["run_id"]

        run = self.client.get(f"/api/v1/validation/runs/{run_id}").json()
        summary = run["result_summary"]
        self.assertEqual(summary["missing"], 1)
        self.assertEqual(summary["out_of_tolerance"], 1)

        issues = self.client.get(f"/api/v1/validation/runs/{run_id}/issues").json()["issues"]
        issue_types = {issue["issue_type"] for issue in issues}
        self.assertIn("out_of_tolerance", issue_types)
        self.assertIn("missing_point", issue_types)


class ReportFindingsApiTests(_EngineApiTestCase):
    def test_report_zip_carries_source_run_findings(self) -> None:
        import json as _json
        import zipfile
        from io import BytesIO

        # A validation run that produces a real finding (a required point missing).
        run_id = self._post(
            "/api/v1/validation/bacnet/runs",
            {
                "expected_points": [
                    {"Expected point name": "MissingPoint", "Required/optional flag": "required"}
                ],
                "observed_points": [],
            },
            "bacnet_validation",
        ).json()["run_id"]
        # A report scoped to that run must carry its findings, not just metadata.
        report = self.client.post(
            "/api/v1/reports",
            json={
                "project_id": "demo-project",
                "site_id": "demo-site",
                "report_type": "issue_report",
                "output_format": "zip",
                "source_run_ids": [run_id],
            },
        )
        self.assertEqual(report.status_code, 200, report.text)
        report_id = report.json()["report_id"]
        download = self.client.get(f"/api/v1/reports/{report_id}/download")
        self.assertEqual(download.status_code, 200, download.text)
        archive = zipfile.ZipFile(BytesIO(download.content))
        self.assertIn("findings.json", archive.namelist())
        findings = _json.loads(archive.read("findings.json"))
        self.assertTrue(findings, "a scoped report must carry the source run's findings")
        self.assertEqual(findings[0]["Source Run"], run_id)
        self.assertIn("missing_point", {finding["Type"] for finding in findings})


class MappingComparisonApiTests(_EngineApiTestCase):
    def test_inline_mapping_comparison_detects_out_of_tolerance(self) -> None:
        response = self._post(
            "/api/v1/validation/mapping/runs",
            {
                "mapping_rows": [
                    {
                        "Asset ID": "AHU-1",
                        "BACnet object name": "SupplyAirTemp",
                        "MQTT field/path": "supply_air_temperature",
                        "MQTT topic": "udmi/AHU-1/pointset",
                        "Tolerance": "0.5",
                        "Mapping required flag": "required",
                    },
                ],
                "bacnet_observed": [
                    {"point_name": "SupplyAirTemp", "observed_value": {"value": 18.6}},
                ],
                "mqtt_observed": [
                    {"field": "supply_air_temperature", "observed_value": {"value": 21.0}},
                ],
            },
            "mapping_validation",
        )
        self.assertEqual(response.status_code, 200, response.text)
        run_id = response.json()["run_id"]
        issues = self.client.get(f"/api/v1/validation/runs/{run_id}/issues").json()["issues"]
        self.assertTrue(any(issue["issue_type"] == "out_of_tolerance" for issue in issues))


class CancelEndpointApiTests(_EngineApiTestCase):
    def test_cancel_sets_flag_and_404_for_missing(self) -> None:
        # Create a run, then request cancellation. We assert the cancel flag is
        # honoured by the store (a fresh inline run picks it up via the engine's
        # is_cancel_requested checker). We verify the plumbing deterministically:
        # the cancel endpoint returns the run, and the store reports cancellation.
        run_id = self._post(
            "/api/v1/discovery/ip/runs",
            {"cidr": "10.0.0.0/30", "dry_run": True},
            "ip_discovery",
        ).json()["run_id"]

        cancel = self.client.post(f"/api/v1/runs/{run_id}/cancel")
        self.assertEqual(cancel.status_code, 200, cancel.text)
        self.assertEqual(cancel.json()["run_id"], run_id)

        # The store now reports cancellation requested for this run.
        from app.services.run_service import RunService

        self.assertTrue(RunService().is_cancel_requested(run_id))

        missing = self.client.post("/api/v1/runs/run_00000000000000_deadbeef/cancel")
        self.assertEqual(missing.status_code, 404)

    def test_cancelled_flag_makes_subsequent_run_report_cancelled(self) -> None:
        # Deterministic cancellation: the point-validation engine checks
        # cancellation at 200-row chunk boundaries. We create a run with a LARGE
        # register, pre-set its cancel flag, then drive the inline processor so
        # the engine observes the flag and flips status to cancelled. This is
        # the same plumbing the worker/inline-fallback path uses.
        from app.services.engine_dispatch import make_cancel_checker
        from app.services.run_service import RunService
        from smart_commissioning_core.engines.point_validation import (
            process_bacnet_validation_run,
        )

        service = RunService()
        new_run = service.create_job_run(
            _bacnet_request_with_points(250),
            expected_job_type="bacnet_validation",
        )
        service.request_cancel(new_run.run_id)

        # The inline route path builds the checker from the store exactly like
        # this; with the flag pre-set the engine observes it at a chunk boundary.
        processed = process_bacnet_validation_run(
            new_run.run_id,
            dict(new_run.parameters),
            run_store=service,
            execution_mode="inline_local_fallback",
            is_cancelled=make_cancel_checker(service, new_run.run_id),
        )
        self.assertEqual(processed.status, "cancelled")
        self.assertTrue(processed.result_summary["cancelled"])


class MqttConfigRollbackApiTests(_EngineApiTestCase):
    def test_publish_captures_previous_then_rollback_republishes(self) -> None:
        # Forward publish (validate-only, no live broker) that records a prior
        # config value supplied in the request for rollback.
        publish = self._post(
            "/api/v1/validation/mqtt-config/runs",
            {
                "topic": "demo-site/b1/ahu-1/config",
                "payload": '{"pointset":{"points":{"sat":{"set_value":22}}}}',
                "confirmed": True,
                "previous_config_payload": {"pointset": {"points": {"sat": {"set_value": 18}}}},
            },
            "mqtt_config_publish",
        )
        self.assertEqual(publish.status_code, 200, publish.text)
        run_id = publish.json()["run_id"]

        run = self.client.get(f"/api/v1/validation/runs/{run_id}").json()
        previous = run["result_summary"]["previous_config"]
        self.assertTrue(previous["captured"])
        self.assertIn("18", previous["payload"])

        # Rollback republishes the captured previous value.
        rollback = self.client.post(f"/api/v1/validation/mqtt-config/runs/{run_id}/rollback")
        self.assertEqual(rollback.status_code, 200, rollback.text)

        rolled = self.client.get(f"/api/v1/validation/runs/{run_id}").json()
        self.assertTrue(rolled["result_summary"]["rollback"])

    def test_rollback_without_captured_value_returns_400(self) -> None:
        publish = self._post(
            "/api/v1/validation/mqtt-config/runs",
            {
                "topic": "demo-site/b1/ahu-1/config",
                "payload": "{}",
                "confirmed": True,
                # No previous_config_payload -> nothing captured to roll back to.
            },
            "mqtt_config_publish",
        )
        run_id = publish.json()["run_id"]
        rollback = self.client.post(f"/api/v1/validation/mqtt-config/runs/{run_id}/rollback")
        self.assertEqual(rollback.status_code, 400, rollback.text)

    def test_live_publish_without_authorization_rejected(self) -> None:
        before = len(self.client.get("/api/v1/validation/runs").json()["runs"])
        response = self._post(
            "/api/v1/validation/mqtt-config/runs",
            {
                "topic": "demo-site/b1/ahu-1/config",
                "payload": "{}",
                "confirmed": True,
                "use_live_broker": True,
                "broker_host": "mqtt.example.local",
            },
            "mqtt_config_publish",
        )
        self.assertEqual(response.status_code, 403, response.text)
        # REGRESSION: the run was created BEFORE the auth check, so a 403 left an
        # orphan stranded at 'queued' that the startup sweep never reclaims. The
        # rejection must leave no run behind.
        after = len(self.client.get("/api/v1/validation/runs").json()["runs"])
        self.assertEqual(after, before, "a rejected live publish must leave no orphaned run")


def _bacnet_request_with_points(count: int):
    from app.schemas.jobs import JobCreateRequest

    return JobCreateRequest.model_validate(
        {
            "project_id": "demo-project",
            "site_id": "demo-site",
            "job_type": "bacnet_validation",
            "parameters": {
                "expected_points": [
                    {"Expected point name": f"P{i}", "Required/optional flag": "optional"}
                    for i in range(count)
                ],
                "observed_points": [],
            },
        }
    )


if __name__ == "__main__":
    unittest.main()
